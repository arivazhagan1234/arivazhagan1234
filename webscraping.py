from selenium import webdriver   
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as Ec
from selenium.common.exceptions import TimeoutException 
from openpyxl import Workbook
from openpyxl import load_workbook
from pandas import DataFrame as df
import os
import time

sheet_location=r'C:/Users/Admins/Downloads/WebScrabed.xlsx'
csv_location=r'C:/Users/Admins/Downloads/WebScrabed.csv'
allproducts_xpath="//*[contains(@class,'a-section a-spacing-base desktop-grid-content-view')]"

price_xpath=".//span[contains(@class,'a-price')]//span[@class='a-offscreen']"
mrp_xpath=".//span[@data-a-strike='true']//span[@class='a-offscreen']"
discount_xpath=".//div[@data-cy='price-recipe']//span[contains(text(),'off')]"

delevery_xpath=".//div[@data-cy='delivery-block']//div[contains(text(),'FREE delivery')]"
fastdelevery_xpath=".//div[@data-cy='delivery-block']//div[contains(text(),'fastest delivery')]"

product_img_url_xpath=".//img[contains(@class,'s-image')]"
product_description_xpath=".//a"
availability_xpath="//div[@id='availabilityInsideBuyBox_feature_div']//span"
sold_by_xpath="//a[@data-csa-c-content-id='odf-desktop-merchant-info']"
brand_xpath="//tr[contains(@class,'po-brand')]//span[@class='a-size-base po-break-word']"

brand_header_xpath01="//a[contains(text(),'Brand')]"
productlabel_xpath=".//h2"
rating_xpath=".//span[@class='a-icon-alt']"
review_xpath=".//span[contains(@class,'s-underline-text')]"
url='https://www.amazon.in/'

#Store the all urls
pageurl=[]

scraping=["https://www.amazon.in/s?rh=n%3A976392031%2Cp_36%3A1318504031&dc&qid=1779346229&rnid=1318502031&ref=sr_nr_p_36_1"]

#creat a new excel or reuse existing exel sheet
def storedata(*args):
    row=list(args)
    print("type of result....", row)  
    if not os.path.exists(sheet_location):
        wb=Workbook()
        ws=wb.active
        ws.title='Web scraped'
        headers=['PAGE TITLE','PRODUCT LABEL','PRODUCT RATING','PRICE','MRP','DISCOUNT','DELIVERY','FAST DELIVERY','IMG URL','STOCK','SELLER','BRAND NAME','URL']
        ws.append(headers)
    else:
        wb=load_workbook(sheet_location)
        ws=wb.active
    ws.append(row)
    wb.save(sheet_location)

    newrow=df([row], columns=['PAGE TITLE', 'PRODUCT LABEL', 'PRODUCT RATING','PRICE','MRP','DISCOUNT', 'DELIVERY', 'FAST DELIVERY', 'IMG URL', 'STOCK','SELLER','BRAND NAME','URL'])
    #create new csv file if not have existing file or use existing file  
    if os.path.exists(csv_location):
        newrow.to_csv(csv_location, mode='a', header=False, index=False)
    else:
        newrow.to_csv(csv_location, mode='w',header=True, index=False)
       
#create new browser instance
driver=webdriver.Chrome()
driver.get(url)
wait=WebDriverWait(driver,20)

driver.set_page_load_timeout(120)
driver.implicitly_wait(10)
driver.set_script_timeout(30)

#page title
print(driver.title) 

#Find all the links
links=wait.until(Ec.presence_of_all_elements_located((By.XPATH, '//a[@href]')))
print("Total number of links........",len(links))

def pagelinks(links):
    for lnk in links:    
        linktext=lnk.text
        linkurl=lnk.get_attribute('href')
        if linkurl and (linkurl.startswith('http://') or linkurl.startswith('https://')):
            print(linktext, linkurl)
            pageurl.append((linkurl, linktext))        
pagelinks(links)    
print(pageurl)


def safe_productlabel(product, xpath):
    elems=product.find_elements(By.XPATH, xpath)
    
    if not elems:
        return "Not found"

    text = elems[0].get_attribute("textContent").strip()

    try:
        if 'â' in text:
            text = text.encode('latin1').decode('utf-8')
    except:
        pass

    return (text.replace('\u202f', ' ') .replace('\xa0', ' ' ) .replace('\n', ' ') .strip() )

def safe_product_img_url(product, xpath):
    elems=product.find_elements(By.XPATH, xpath)
    
    if not elems:
        return "Not found"

    img_url = elems[0].get_attribute("src").strip()
    return img_url

def product_avail(producturl, availability_xpath,sold_by_xpath,brand_xpath):
        
        product_stock = "Not found"
        seller = "Not found"
        brand_name= "Not found"
                
        try:
            driver.execute_script("window.open(arguments[0],'_blank');",producturl)
            driver.switch_to.window(driver.window_handles[1])
            product_title=driver.title
            print("product title is..........................", product_title)
            
            #availability
            try:
                products=wait.until(Ec.presence_of_all_elements_located((By.XPATH, availability_xpath)))
                product_stock=products[0].get_attribute("textContent").strip()
            except TimeoutException:
                print("Timeout occurred while fetching product availability")
                
            #seller
            try:
                sellers=wait.until(Ec.presence_of_all_elements_located((By.XPATH, sold_by_xpath)))
                seller=sellers[0].get_attribute("textContent").strip()
            except TimeoutException:
                print("Timeout occurred while fetching seller information")
            
            #brand name
            brand_names=driver.find_elements(By.XPATH, brand_xpath)
            print("brand_names lenght is ......................................", len(brand_names))
            if brand_names:
                brand_name=brand_names[0].get_attribute("textContent").strip()
            else:
                other_brand_names=driver.find_elements(By.XPATH, brand_header_xpath01)
                print("other_brand_names lenght is ......................................", len(other_brand_names))
                if other_brand_names:
                    brand_name=other_brand_names[0].get_attribute("textContent").strip()
                else:
                    print("Brand name not found")

        except TimeoutException:
                print("Timeout occurred while loading product page")
                driver.execute_script("window.stop();")    
        finally:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        return product_stock, seller, brand_name

#to get product label and discription
def products_description(allproducts_xpath,productlabel_xpath,rating_xpath,review_xpath,price_xpath, mrp_xpath, discount_xpath,delevery_xpath, fastdelevery_xpath,product_img_url_xpath,product_description_xpath,availability_xpath,sold_by_xpath,brand_xpath):
    productdata=[]
    
    allproducts=wait.until(Ec.presence_of_all_elements_located((By.XPATH, allproducts_xpath)))
    print("Total number of products........",len(allproducts))
    
    for product in allproducts:
        label=safe_productlabel(product, productlabel_xpath)
        rating=safe_productlabel(product, rating_xpath)
        review=safe_productlabel(product, review_xpath)
        final_rating=f'{rating} {review}'

        price=safe_productlabel(product, price_xpath)
        mrp=safe_productlabel(product, mrp_xpath)
        discount=safe_productlabel(product, discount_xpath)
        delevery=safe_productlabel(product, delevery_xpath)
        fastdelevery=safe_productlabel(product, fastdelevery_xpath)     
        product_img_url=safe_product_img_url(product, product_img_url_xpath)

        elems=product.find_elements(By.XPATH, product_description_xpath)
        print("elems is avail......................................", elems)
        if elems:
            product_url=elems[0].get_attribute("href")
            product_stock,seller,brand_name=product_avail(product_url,availability_xpath,sold_by_xpath,brand_xpath) 
        else:
            product_stock="Not found"
            seller="Not found"
            brand_name="Not found"

        productdata.append((label, final_rating, price, mrp, discount, delevery, fastdelevery, product_img_url, product_stock, seller,brand_name))
        print("Label:", label, "| final_rating:", final_rating, "| Price:", price, "| MRP:", mrp, "| Discount:", discount, "| Delivery:", delevery, "| Fast Delivery:", fastdelevery, "| Image URL:", product_img_url, "| Availability:", product_stock, "| Seller Name:", seller, "| Brand Name:", brand_name)
        time.sleep(2)
    return productdata

#store page title,url,
def store_details(scraping,allproducts_xpath,productlabel_xpath,rating_xpath,review_xpath, price_xpath, mrp_xpath, discount_xpath,delevery_xpath, fastdelevery_xpath,product_img_url_xpath,product_description_xpath,availability_xpath,sold_by_xpath,brand_xpath):

    for url in scraping:
        try:
            driver.get(url)
            driver.set_page_load_timeout(120)
            driver.implicitly_wait(20)
            driver.maximize_window()

            pagetitle=driver.title
            print("page url is..........................", url)  
            productsdes=products_description(allproducts_xpath,productlabel_xpath,rating_xpath,review_xpath,price_xpath, mrp_xpath, discount_xpath,delevery_xpath, fastdelevery_xpath,product_img_url_xpath,product_description_xpath,availability_xpath,sold_by_xpath,brand_xpath)
            for label,final_rating,price,mrp,discount,delevery,fastdelevery,product_img_url,product_stock,seller,brand_name in productsdes: 
                storedata(pagetitle,label,final_rating,price,mrp,discount,delevery,fastdelevery,product_img_url,product_stock,seller,brand_name,url)
        except TimeoutException:

            print("Time out url is ......",url)
            driver.execute_script("window.stop();")
store_details(scraping,allproducts_xpath,productlabel_xpath, rating_xpath, review_xpath,price_xpath, mrp_xpath, discount_xpath,delevery_xpath, fastdelevery_xpath,product_img_url_xpath,product_description_xpath,availability_xpath,sold_by_xpath,brand_xpath)
                                                                        