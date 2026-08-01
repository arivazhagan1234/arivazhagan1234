#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PIIA v2.10 — Fase 1 — Script único executável
Plataforma de Inteligência Imobiliária Angola

Uso:
    python piia.py --setup
    python piia.py --scraping                  → AngoCasa: Luanda inteira, compra + arrendamento
    python piia.py --scraping --fontes todas     → Multi-fonte: AngoCasa + AngoImóveis + MIA + RE/MAX + Zenki + REE Gera
    python piia.py --scraping --bairro Talatona --max-paginas 5
    python piia.py --scraping --por-bairros --max-paginas 1
    python piia.py --excel
    python piia.py                        (executa tudo)
    python piia.py --debug                (verbose + salva HTML)

Requisitos:
    pip install requests beautifulsoup4 lxml openpyxl
"""

# ─────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────
import os, re, sys, json, time, math, random, shutil
import hashlib, sqlite3, logging, argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from logging.handlers import RotatingFileHandler
from urllib.parse import urljoin, urlparse, urlunparse, urlencode, parse_qsl
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                   Border, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─────────────────────────────────────────────────────────────────
# RAIZ DO PROJECTO
# ─────────────────────────────────────────────────────────────────
# Detecta automaticamente a pasta onde este script está salvo.
# Ex.: se o arquivo estiver em:
# C:\Users\claud\OneDrive\Desktop\Pessoal\Projectos\Imobiliário\Script Imobiliário.py
# então PROJECT_ROOT será:
# C:\Users\claud\OneDrive\Desktop\Pessoal\Projectos\Imobiliário
PROJECT_ROOT = Path(__file__).resolve().parent

# ─────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────
CFG: Dict = {
    "db":           PROJECT_ROOT / "database" / "piia.db",
    "data_raw":     PROJECT_ROOT / "data" / "raw",
    "excel_dir":    PROJECT_ROOT / "reports" / "excel",
    "html_debug":   PROJECT_ROOT / "html_debug",
    "logs_dir":     PROJECT_ROOT / "logs",
    "backups_dir":  PROJECT_ROOT / "backups",
    "base_url":     "https://www.angocasa.com",
    "items_pag":    50,
    "max_paginas":  200,
    "delay_min":    2.5,
    "delay_max":    6.0,
    "max_retries":  3,
    "timeout":      30,
    "salvar_html":  False,
    "cambio":       828.0,
    "reprocessar_horas": 168,  # evita reprocessar o mesmo anúncio em retomadas/interrupções (7 dias)
    "apenas_negocio": "todos",
    "max_anuncios_por_fonte": None,  # limite opcional para testes rápidos por fonte/negócio
    "bairros": [
        {"nome":"Talatona",        "slug":"talatona",                        "lat":-8.9262,"lon":13.1893},
        {"nome":"Lar do Patriota", "slug":"benfica-inclui-lar-do-patriota-", "lat":-8.9120,"lon":13.2150},
        {"nome":"Benfica",         "slug":"benfica-inclui-lar-do-patriota-", "lat":-8.9000,"lon":13.2200},
        {"nome":"Nova Vida",       "slug":"nova-vida-golfe",                 "lat":-8.9650,"lon":13.2200},
        {"nome":"Kilamba",         "slug":"kilamba",                         "lat":-8.9980,"lon":13.2820},
        {"nome":"Alvalade",        "slug":"alvalade",                        "lat":-8.8270,"lon":13.2450},
        {"nome":"Maianga",         "slug":"maianga",                         "lat":-8.8200,"lon":13.2350},
        {"nome":"Ingombota",       "slug":"ingombota",                       "lat":-8.8280,"lon":13.2320},
        {"nome":"Maculusso",       "slug":"maculusso",                       "lat":-8.8230,"lon":13.2430},
        {"nome":"Morro Bento",     "slug":"morro-bento",                     "lat":-8.8400,"lon":13.2500},
        {"nome":"Miramar",         "slug":"miramar",                         "lat":-8.8120,"lon":13.2390},
        {"nome":"Ilha de Luanda",  "slug":"ilha-do-cabo",                    "lat":-8.8000,"lon":13.2100},
        {"nome":"Coqueiros",       "slug":"coqueiros",                       "lat":-8.8300,"lon":13.2280},
        {"nome":"Camama",          "slug":"camama",                          "lat":-8.9430,"lon":13.2580},
        {"nome":"Viana",           "slug":"viana",                           "lat":-8.9040,"lon":13.3740},
        {"nome":"Zango",           "slug":"zango",                           "lat":-8.9200,"lon":13.4200},
    ],
}


# ─────────────────────────────────────────────────────────────────
# FONTES MULTI-SITE
# ─────────────────────────────────────────────────────────────────
# Observação importante:
# - AngoCasa usa coletor dedicado, com seletores já conhecidos.
# - As demais fontes usam coletor genérico configurável. Em sites que alterem HTML,
#   basta ajustar caminhos e seletores no bloco abaixo, sem mexer na BD/Excel.
FONTES_EXTRA: Dict[str, Dict] = {
    "angoimoveis": {
        "nome": "AngoImóveis",
        "base": "https://angoimoveis.com",
        "paths": {
            "compra": ["/comprar", "/venda", "/imoveis-a-venda", "/imoveis", "/"],
            "arrendamento": ["/arrendar", "/alugar", "/imoveis-para-arrendar", "/imoveis", "/"],
        },
        "page_params": ["page", "pagina", "paged"],
        "page_path_patterns": ["/page/{pagina}/", "/pagina/{pagina}/"],
        "link_keywords": ["imovel", "imoveis", "apartamento", "vivenda", "moradia", "terreno", "armazem", "escritorio", "loja"],
    },
    "mia": {
        "nome": "MIA",
        "base": "https://mia.africa",
        "dedicado": True,
        "paths": {
            "compra": ["/comprar", "/venda", "/properties", "/imoveis", "/"],
            "arrendamento": ["/arrendar", "/alugar", "/rent", "/properties", "/imoveis", "/"],
        },
        "page_params": ["page", "pagina"],
        "page_path_patterns": ["/page/{pagina}/", "/pagina/{pagina}/"],
        "link_keywords": ["property", "properties", "imovel", "imoveis", "apartamento", "vivenda", "terreno", "rent", "sale"],
    },
    "remax": {
        "nome": "RE/MAX Angola",
        "base": "https://www.remax-multitrust.co.ao",
        "dedicado": True,
        "paths": {
            "compra": ["/listings?ListingClass=-1&TransactionTypeUID=261"],
            "arrendamento": ["/listings?ListingClass=-1&TransactionTypeUID=260"],
        },
        "page_params": [],
        "page_path_patterns": [],
        "link_keywords": ["angariacoes", "listing", "property", "imovel", "apartamento", "vivenda", "terreno"],
    },
    "zenki": {
        "nome": "Zenki Real Estate",
        "base": "https://zenkirealestate.com",
        "dedicado": True,
        "paths": {
            "compra": ["/properties", "/property", "/imoveis", "/for-sale", "/"],
            "arrendamento": ["/properties", "/property", "/for-rent", "/rent", "/"],
        },
        "page_params": ["page", "paged"],
        "page_path_patterns": ["/page/{pagina}/"],
        "link_keywords": ["property", "properties", "imovel", "apartamento", "office", "warehouse", "retail", "land"],
    },
    "ree_gera": {
        "nome": "REE Gera Imobiliária",
        "base": "https://ree-gera.com",
        "dedicado": True,
        "paths": {
            "compra": ["/comprar", "/venda", "/imoveis", "/properties", "/"],
            "arrendamento": ["/arrendar", "/alugar", "/imoveis", "/properties", "/"],
        },
        "page_params": ["page", "paged", "pagina"],
        "page_path_patterns": ["/page/{pagina}/", "/pagina/{pagina}/"],
        "link_keywords": ["imovel", "imoveis", "apartamento", "vivenda", "moradia", "terreno", "arrendar", "vender"],
    },
}

FONTES_DISPONIVEIS = ["angocasa"] + list(FONTES_EXTRA.keys())


def carregar_env():
    env = PROJECT_ROOT / ".env"
    if not env.exists(): return
    for ln in env.read_text(encoding="utf-8").splitlines():
        ln = ln.strip()
        if not ln or ln.startswith("#") or "=" not in ln: continue
        k, _, v = ln.partition("=")
        k, v = k.strip(), v.strip()
        m = {"TAXA_CAMBIO_AKZ_USD":("cambio",float),
             "SCRAPING_DELAY_MIN": ("delay_min",float),
             "SCRAPING_DELAY_MAX": ("delay_max",float),
             "MAX_PAGINAS":        ("max_paginas",int),
             "SALVAR_HTML_DEBUG":  ("salvar_html",lambda x:x.upper()=="TRUE"),
             "REPROCESSAR_HORAS":   ("reprocessar_horas",int)}
        if k in m:
            try: CFG[m[k][0]] = m[k][1](v)
            except: pass

# ─────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────
def configurar_log(debug=False):
    CFG["logs_dir"].mkdir(parents=True, exist_ok=True)
    lf  = CFG["logs_dir"] / f"piia_{datetime.now():%Y%m%d}.log"
    fmt = logging.Formatter("%(asctime)s [%(levelname)-8s] %(name)s — %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
    root = logging.getLogger()
    root.setLevel(logging.DEBUG if debug else logging.INFO)
    if not root.handlers:
        fh = RotatingFileHandler(lf, maxBytes=10*1024*1024, backupCount=30, encoding="utf-8")
        fh.setFormatter(fmt); root.addHandler(fh)
        ch = logging.StreamHandler(sys.stdout)
        ch.setFormatter(fmt); root.addHandler(ch)

log = logging.getLogger("piia")

# ─────────────────────────────────────────────────────────────────
# SETUP DE PASTAS
# ─────────────────────────────────────────────────────────────────
def criar_estrutura():
    pastas = [
        "config","database","data/raw","data/processed","data/cache",
        "reports/excel","reports/pdf","reports/dashboards",
        "market_reports/PDFs Mercado","market_reports/Noticias","market_reports/Processados",
        "project_sources/Infraestruturas","project_sources/Projetos Publicos",
        "project_sources/Projetos Privados","project_sources/Noticias",
        "screenshots","html_debug","logs","tests","backups","src",
    ]
    for p in pastas:
        full = PROJECT_ROOT / hp
        full.mkdir(parents=True, exist_ok=True)
        log.info(f"  ✔ {p}")
    env_ex = PROJECT_ROOT / ".env.example"
    if not env_ex.exists():
        env_ex.write_text(
            "TAXA_CAMBIO_AKZ_USD=828.0\n"
            "SCRAPING_DELAY_MIN=2.5\n"
            "SCRAPING_DELAY_MAX=6.0\n"
            "MAX_PAGINAS=200\n"
            "SALVAR_HTML_DEBUG=FALSE\n", encoding="utf-8")
    log.info("✅ Estrutura criada.")


# ─────────────────────────────────────────────────────────────────
# BASE DE DADOS
# ─────────────────────────────────────────────────────────────────
class DB:
    def __init__(self):
        p = CFG["db"]
        p.parent.mkdir(parents=True, exist_ok=True)
        self.path = str(p)
        self._init()

    def _con(self):
        c = sqlite3.connect(self.path, timeout=30)
        c.row_factory = sqlite3.Row
        c.execute("PRAGMA journal_mode=WAL")
        c.execute("PRAGMA foreign_keys=ON")
        return c

    def _init(self):
        conn = self._con()
        try:
            c = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS anuncios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT NOT NULL UNIQUE, id_angocasa TEXT, hash_conteudo TEXT,
                tipo_negocio TEXT, pais TEXT DEFAULT 'Angola',
                provincia TEXT DEFAULT 'Luanda', bairro TEXT, municipio TEXT,
                lat REAL, lon REAL, headline TEXT, titulo_anuncio TEXT,
                tipo_imovel TEXT, tipologia TEXT, estado TEXT,
                area_m2 REAL, area_util_m2 REAL, area_bruta_m2 REAL, area_lote_m2 REAL,
                preco_original REAL, moeda TEXT DEFAULT 'AKZ',
                preco_akz REAL, preco_usd REAL, preco_m2_akz REAL,
                quartos INTEGER, wc INTEGER, mobilado INTEGER DEFAULT 0,
                garagem INTEGER DEFAULT 0, piscina INTEGER DEFAULT 0,
                gerador INTEGER DEFAULT 0, reservatorio INTEGER DEFAULT 0,
                seguranca INTEGER DEFAULT 0, churrasqueira INTEGER DEFAULT 0,
                ar_condicionado INTEGER DEFAULT 0, internet INTEGER DEFAULT 0,
                condominio_fechado INTEGER DEFAULT 0,
                publicado_em TEXT, descricao TEXT,
                data_extracao TEXT DEFAULT CURRENT_TIMESTAMP,
                data_ultima_atualizacao TEXT DEFAULT CURRENT_TIMESTAMP,
                ativo INTEGER DEFAULT 1, n_atualizacoes INTEGER DEFAULT 0)""")
            for i in ["CREATE INDEX IF NOT EXISTS idx_b ON anuncios(bairro)",
                      "CREATE INDEX IF NOT EXISTS idx_n ON anuncios(tipo_negocio)",
                      "CREATE INDEX IF NOT EXISTS idx_t ON anuncios(tipo_imovel)",
                      "CREATE INDEX IF NOT EXISTS idx_p ON anuncios(preco_akz)",
                      "CREATE INDEX IF NOT EXISTS idx_a ON anuncios(ativo)"]:
                c.execute(i)
            c.execute("""CREATE TABLE IF NOT EXISTS historico_precos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                anuncio_id INTEGER NOT NULL REFERENCES anuncios(id),
                url TEXT NOT NULL, id_angocasa TEXT, bairro TEXT,
                tipo_negocio TEXT, tipo_imovel TEXT, tipologia TEXT,
                preco_original REAL, moeda TEXT,
                preco_akz REAL, preco_usd REAL, preco_m2_akz REAL,
                preco_akz_ant REAL, variacao_akz REAL, variacao_pct REAL,
                evento TEXT, data_registo TEXT DEFAULT CURRENT_TIMESTAMP)""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_hp ON historico_precos(url)")
            c.execute("""CREATE TABLE IF NOT EXISTS anuncios_removidos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, anuncio_id INTEGER,
                url TEXT, id_angocasa TEXT, bairro TEXT, tipo_negocio TEXT,
                preco_akz REAL, dias_no_mercado INTEGER,
                data_remocao TEXT DEFAULT CURRENT_TIMESTAMP)""")
            c.execute("""CREATE TABLE IF NOT EXISTS erros_scraping (
                id INTEGER PRIMARY KEY AUTOINCREMENT, url TEXT,
                tipo_erro TEXT, mensagem TEXT, tentativas INTEGER DEFAULT 1,
                data_erro TEXT DEFAULT CURRENT_TIMESTAMP)""")
            c.execute("""CREATE TABLE IF NOT EXISTS runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                inicio TEXT, fim TEXT, status TEXT DEFAULT 'em_curso',
                modo TEXT, n_coletados INTEGER DEFAULT 0,
                n_novos INTEGER DEFAULT 0, n_atualizados INTEGER DEFAULT 0,
                n_preco_alt INTEGER DEFAULT 0, n_removidos INTEGER DEFAULT 0,
                n_erros INTEGER DEFAULT 0, bairros_proc TEXT, detalhes TEXT)""")
            c.execute("""CREATE TABLE IF NOT EXISTS scraping_checkpoints (
                fonte TEXT NOT NULL,
                tipo_negocio TEXT NOT NULL,
                contexto TEXT DEFAULT '',
                pagina INTEGER DEFAULT 1,
                atualizado_em TEXT DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (fonte, tipo_negocio, contexto)
            )""")

            # Migração leve para versões anteriores do PIIA.
            # Permite usar a mesma base SQLite sem apagar dados antigos.
            self._garantir_colunas(c, "anuncios", {
                "fonte": "TEXT DEFAULT 'AngoCasa'",
                "id_fonte": "TEXT",
                "anunciante": "TEXT",
                "telefone_whatsapp": "TEXT",
                "url_imagem": "TEXT",
                "segmento_imobiliario": "TEXT",
                "email_agente": "TEXT",
                "id_zenki": "TEXT",
                "imovel_corporativo": "INTEGER DEFAULT 0",
                "preco_disponivel": "INTEGER DEFAULT 0",
                "url_brochura": "TEXT",
                "pdf_brochura": "TEXT",
                "hash_dedupe": "TEXT",
                "headline_completa": "TEXT",
                "raw_json": "TEXT",
            })
            self._garantir_colunas(c, "erros_scraping", {
                "fonte": "TEXT",
                "pagina": "INTEGER",
            })
            conn.commit()
            log.info("✅ Base de dados OK.")
        finally:
            conn.close()


    def _garantir_colunas(self, cursor, tabela: str, colunas: Dict[str, str]):
        """Adiciona colunas novas sem destruir a base existente."""
        cursor.execute(f"PRAGMA table_info({tabela})")
        existentes = {r[1] for r in cursor.fetchall()}
        for nome, definicao in colunas.items():
            if nome not in existentes:
                cursor.execute(f"ALTER TABLE {tabela} ADD COLUMN {nome} {definicao}")
                log.info(f"  Migração SQLite: coluna adicionada {tabela}.{nome}")


    def _reg_preco(self, c, d, aid, p_ant, evento):
        p = d.get("preco_akz")
        va = vp = None
        if p_ant and p:
            va = float(p) - float(p_ant)
            vp = (va / float(p_ant)) * 100
        c.execute("""INSERT INTO historico_precos
            (anuncio_id,url,id_angocasa,bairro,tipo_negocio,tipo_imovel,tipologia,
             preco_original,moeda,preco_akz,preco_usd,preco_m2_akz,
             preco_akz_ant,variacao_akz,variacao_pct,evento)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (aid,d.get("url"),d.get("id_angocasa"),d.get("bairro"),
             d.get("tipo_negocio"),d.get("tipo_imovel"),d.get("tipologia"),
             d.get("preco_original"),d.get("moeda"),p,d.get("preco_usd"),
             d.get("preco_m2_akz"),p_ant,va,vp,evento))

    def upsert(self, d):
        conn = self._con()
        try:
            c = conn.cursor()
            ignore = {"data_extracao","data_ultima_atualizacao","hash_conteudo","n_atualizacoes"}
            h_novo = hashlib.md5(json.dumps(
                {k:v for k,v in d.items() if k not in ignore},
                sort_keys=True, ensure_ascii=False, default=str).encode()).hexdigest()
            c.execute("SELECT id,hash_conteudo,preco_akz FROM anuncios WHERE url=?",(d["url"],))
            ex = c.fetchone()
            agora = datetime.now().isoformat()
            if not ex:
                d["hash_conteudo"]=h_novo; d["data_extracao"]=agora
                d["data_ultima_atualizacao"]=agora; d["n_atualizacoes"]=0
                cols=", ".join(d.keys()); phs=", ".join(["?"]*len(d))
                c.execute(f"INSERT INTO anuncios ({cols}) VALUES ({phs})",list(d.values()))
                aid=c.lastrowid
                self._reg_preco(c,d,aid,None,"novo")
                conn.commit(); return aid,"novo"
            aid=ex["id"]; p_ant=ex["preco_akz"]
            if ex["hash_conteudo"]==h_novo: return aid,"sem_alteracao"
            p_novo=d.get("preco_akz")
            ev=("preco_alterado" if p_ant and p_novo and abs(float(p_ant)-float(p_novo))>1
                else "atualizado")
            d["hash_conteudo"]=h_novo; d["data_ultima_atualizacao"]=agora
            upd=[k for k in d if k not in ("url","data_extracao")]
            sc=", ".join(f"{k}=?" for k in upd)
            c.execute(f"UPDATE anuncios SET {sc},n_atualizacoes=n_atualizacoes+1 WHERE url=?",
                      [d[k] for k in upd]+[d["url"]])
            if ev=="preco_alterado": self._reg_preco(c,d,aid,p_ant,ev)
            conn.commit(); return aid,ev
        except:
            conn.rollback(); raise
        finally:
            conn.close()

    def marcar_inativos(self, bairro, negocio, urls_ativas, completa):
        if not completa or not urls_ativas: return
        conn=self._con()
        try:
            c=conn.cursor()
            phs=",".join(["?"]*len(urls_ativas))
            c.execute(f"""SELECT id,url,id_angocasa,preco_akz,data_extracao
                FROM anuncios WHERE bairro=? AND tipo_negocio=?
                AND url NOT IN ({phs}) AND ativo=1""",
                [bairro,negocio]+list(urls_ativas))
            rem=c.fetchall()
            for r in rem:
                dias=None
                try:
                    dias=(datetime.now()-datetime.fromisoformat(r["data_extracao"])).days
                except: pass
                c.execute("""INSERT INTO anuncios_removidos
                    (anuncio_id,url,id_angocasa,bairro,tipo_negocio,preco_akz,dias_no_mercado)
                    VALUES(?,?,?,?,?,?,?)""",
                    (r["id"],r["url"],r["id_angocasa"],bairro,negocio,r["preco_akz"],dias))
            c.execute(f"""UPDATE anuncios SET ativo=0
                WHERE bairro=? AND tipo_negocio=?
                AND url NOT IN ({phs}) AND ativo=1""",
                [bairro,negocio]+list(urls_ativas))
            conn.commit()
            if rem: log.info(f"  ⬇  {len(rem)} removidos ({bairro}/{negocio})")
        finally:
            conn.close()

    def reg_erro(self, url, tipo, msg):
        conn=self._con()
        try:
            c=conn.cursor()
            c.execute("INSERT INTO erros_scraping(url,tipo_erro,mensagem) VALUES(?,?,?)",(url,tipo,msg))
            conn.commit()
        finally: conn.close()

    def anuncio_processado_recente(self, url, horas=None):
        """
        Verifica se um anúncio já foi processado recentemente.
        Usado para retomar execuções interrompidas sem baixar novamente
        detalhes que já foram gravados no SQLite.
        """
        horas = horas if horas is not None else CFG.get("reprocessar_horas", 12)
        conn = self._con()
        try:
            c = conn.cursor()
            c.execute("SELECT data_ultima_atualizacao FROM anuncios WHERE url=? AND ativo=1", (url,))
            row = c.fetchone()
            if not row or not row["data_ultima_atualizacao"]:
                return False
            try:
                dt = datetime.fromisoformat(row["data_ultima_atualizacao"])
                return (datetime.now() - dt).total_seconds() < horas * 3600
            except Exception:
                return False
        finally:
            conn.close()

    def obter_todos(self, filtros=None):
        conn=self._con(); q="SELECT * FROM anuncios WHERE ativo=1"; p=[]
        try:
            if filtros:
                for f in ("bairro","tipo_negocio","tipo_imovel","tipologia"):
                    if filtros.get(f): q+=f" AND {f}=?"; p.append(filtros[f])
                if filtros.get("preco_min"): q+=" AND preco_akz>=?"; p.append(filtros["preco_min"])
                if filtros.get("preco_max"): q+=" AND preco_akz<=?"; p.append(filtros["preco_max"])
            q+=" ORDER BY data_extracao DESC"
            c=conn.cursor(); c.execute(q,p)
            return [dict(r) for r in c.fetchall()]
        finally: conn.close()

    def obter_historico(self):
        conn=self._con()
        try:
            c=conn.cursor()
            c.execute("SELECT * FROM historico_precos ORDER BY data_registo DESC")
            return [dict(r) for r in c.fetchall()]
        finally: conn.close()

    def obter_alteracoes_semana(self):
        conn=self._con()
        try:
            c=conn.cursor()
            lim=(datetime.now()-timedelta(days=7)).isoformat()
            c.execute("""SELECT hp.*,a.headline,a.area_m2,a.bairro as bairro_an
                FROM historico_precos hp LEFT JOIN anuncios a ON hp.anuncio_id=a.id
                WHERE hp.evento='preco_alterado' AND hp.data_registo>=?
                ORDER BY hp.variacao_pct ASC""",(lim,))
            return [dict(r) for r in c.fetchall()]
        finally: conn.close()

    def stats(self):
        conn=self._con()
        try:
            c=conn.cursor(); r={}
            for k,q in [
                ("total",         "SELECT COUNT(*) FROM anuncios WHERE ativo=1"),
                ("compra",        "SELECT COUNT(*) FROM anuncios WHERE ativo=1 AND tipo_negocio='compra'"),
                ("arrendamento",  "SELECT COUNT(*) FROM anuncios WHERE ativo=1 AND tipo_negocio='arrendamento'"),
                ("bairros",       "SELECT COUNT(DISTINCT bairro) FROM anuncios WHERE ativo=1"),
                ("hist_precos",   "SELECT COUNT(*) FROM historico_precos"),
                ("preco_alt",     "SELECT COUNT(*) FROM historico_precos WHERE evento='preco_alterado'"),
                ("removidos",     "SELECT COUNT(*) FROM anuncios_removidos"),
                ("erros",         "SELECT COUNT(*) FROM erros_scraping"),
            ]:
                c.execute(q); r[k]=c.fetchone()[0]
            c.execute("SELECT MAX(data_extracao) FROM anuncios")
            r["ultima_extracao"]=c.fetchone()[0]
            return r
        finally: conn.close()

    def iniciar_run(self,modo):
        conn=self._con()
        try:
            c=conn.cursor()
            c.execute("INSERT INTO runs(inicio,status,modo) VALUES(?,?,?)",
                      (datetime.now().isoformat(),"em_curso",modo))
            conn.commit(); return c.lastrowid
        finally: conn.close()

    def finalizar_run(self,rid,dados):
        dados["fim"]=datetime.now().isoformat()
        dados["status"]=dados.get("status","concluido")
        conn=self._con()
        try:
            c=conn.cursor()
            sc=", ".join(f"{k}=?" for k in dados)
            c.execute(f"UPDATE runs SET {sc} WHERE id=?",list(dados.values())+[rid])
            conn.commit()
        finally: conn.close()

    def obter_checkpoint(self, fonte: str, tipo_negocio: str, contexto: str = "") -> int:
        """Retorna a próxima página a processar para retomada segura."""
        conn = self._con()
        try:
            c = conn.cursor()
            c.execute("""SELECT pagina FROM scraping_checkpoints
                         WHERE fonte=? AND tipo_negocio=? AND contexto=?""",
                      (fonte, tipo_negocio, contexto or ""))
            row = c.fetchone()
            return int(row["pagina"]) if row and row["pagina"] else 1
        except Exception:
            return 1
        finally:
            conn.close()

    def salvar_checkpoint(self, fonte: str, tipo_negocio: str, pagina: int, contexto: str = ""):
        """Grava a próxima página a processar em caso de interrupção."""
        conn = self._con()
        try:
            c = conn.cursor()
            c.execute("""INSERT INTO scraping_checkpoints(fonte,tipo_negocio,contexto,pagina,atualizado_em)
                         VALUES(?,?,?,?,?)
                         ON CONFLICT(fonte,tipo_negocio,contexto)
                         DO UPDATE SET pagina=excluded.pagina, atualizado_em=excluded.atualizado_em""",
                      (fonte, tipo_negocio, contexto or "", int(pagina), datetime.now().isoformat()))
            conn.commit()
        finally:
            conn.close()

    def limpar_checkpoint(self, fonte: str, tipo_negocio: str, contexto: str = ""):
        """Remove checkpoint quando a coleta do bloco foi concluída."""
        conn = self._con()
        try:
            c = conn.cursor()
            c.execute("DELETE FROM scraping_checkpoints WHERE fonte=? AND tipo_negocio=? AND contexto=?",
                      (fonte, tipo_negocio, contexto or ""))
            conn.commit()
        finally:
            conn.close()

    def backup(self):
        d=CFG["backups_dir"]; d.mkdir(parents=True,exist_ok=True)
        dst=d/f"piia_backup_{datetime.now():%Y%m%d_%H%M%S}.db"
        shutil.copy2(self.path,str(dst)); log.info(f"✅ Backup: {dst.name}")


# ─────────────────────────────────────────────────────────────────
# NORMALIZADORES
# ─────────────────────────────────────────────────────────────────
def norm_num(txt):
    if not txt: return None
    s = re.sub(r"[^\d,.]","",str(txt).strip())
    if not re.search(r"\d",s): return None
    if "," in s and "." in s:
        s=s.replace(".","").replace(",",".")
    elif "," in s and s.index(",") < len(s)-4:
        s=s.replace(",","")
    elif "," in s:
        s=s.replace(",",".")
    elif "." in s and "." in s:
        s=s.replace(".","")
    try: return float(s)
    except: return None

def norm_area(txt):
    if not txt: return None
    m=re.search(r"([\d\s.,]+)\s*m[²2]",str(txt),re.I)
    return norm_num(m.group(1)) if m else None

def norm_area_flex(txt):
    """Extrai área mesmo quando o site informa apenas o número, sem m²."""
    if not txt:
        return None
    area = norm_area(txt)
    if area:
        return area
    n = norm_num(txt)
    # Evitar capturar IDs, anos ou telefones como área. Áreas imobiliárias plausíveis.
    if n and 5 <= float(n) <= 5_000_000:
        return float(n)
    return None

def norm_tipologia(txt):
    if not txt: return None
    m=re.search(r"\b([TVtv]\d+)\b",str(txt))
    return m.group(1).upper() if m else None

def norm_tipo_url(url):
    u=url.lower()
    for t in ("apartamento","vivenda","terreno","loja","escritorio","armazem","hotel","quinta"):
        if f"/{t}/" in u: return t
    return "outro"

def norm_negocio(txt):
    t=(txt or "").lower()
    if any(p in t for p in ("arrendar","arrend")): return "arrendamento"
    if any(p in t for p in ("comprar","vend","compra")): return "compra"
    return t.strip()

def akz2usd(v):
    return round(float(v)/CFG["cambio"],2) if v else None

def usd2akz(v):
    return round(float(v)*CFG["cambio"],2) if v else None

def tem_kw(txt, kws):
    if not txt: return 0
    t=txt.lower()
    return 1 if any(k in t for k in kws) else 0


def limpar_texto(txt, limite=None):
    """Normaliza espaços e remove quebras duplicadas."""
    if txt is None:
        return ""
    s = re.sub(r"\s+", " ", str(txt)).strip()
    return s[:limite].strip() if limite else s

def dominio_base(url):
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return ""

def url_absoluta(base, href):
    if not href:
        return ""
    u = urljoin(base, href.strip())
    if u.startswith("http://"):
        u = u.replace("http://", "https://", 1)
    parsed = urlparse(u)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", parsed.query, ""))

def detectar_moeda(txt, default="AKZ"):
    t = (txt or "").upper()
    if any(x in t for x in ["USD", "US$", "$", "DÓLAR", "DOLAR"]):
        return "USD"
    if any(x in t for x in ["AKZ", "AOA", " KZ", "KZ ", "KWANZA"]):
        return "AKZ"
    return default

def extrair_preco_texto(txt):
    """Extrai primeiro preço monetário plausível de um texto."""
    if not txt:
        return None, "AKZ"
    texto = limpar_texto(txt)
    padroes = [
        r"(?:USD|US\$|\$|AKZ|AOA|KZ)\s*[\d\s.,]{3,}",
        r"[\d\s.,]{3,}\s*(?:USD|US\$|\$|AKZ|AOA|KZ)",
    ]
    candidatos = []
    for p in padroes:
        candidatos += re.findall(p, texto, flags=re.I)
    if not candidatos:
        return None, "AKZ"
    # Preferir candidato que não pareça telefone; valores imobiliários geralmente têm >= 4 dígitos.
    for cand in candidatos:
        n = norm_num(cand)
        if n and n >= 1000:
            return n, detectar_moeda(cand)
    return None, "AKZ"

def extrair_telefone(txt):
    if not txt:
        return None
    t = limpar_texto(txt)
    pats = [
        r"\+244\s?9\d{2}\s?\d{3}\s?\d{3}",
        r"\b9\d{2}\s?\d{3}\s?\d{3}\b",
        r"\+244\s?\d{3}\s?\d{3}\s?\d{3}",
    ]
    achados=[]
    for p in pats:
        achados += re.findall(p, t)
    if not achados:
        return None
    limpos=[]
    for a in achados:
        x=re.sub(r"\s+", "", a)
        if x not in limpos:
            limpos.append(x)
    return "; ".join(limpos[:3])

def detectar_bairro(txt):
    t=(txt or "").lower()
    for b in CFG["bairros"]:
        if b["nome"].lower() in t or b["slug"].lower().replace("-", " ") in t:
            return b["nome"], b.get("lat"), b.get("lon")
    return None, None, None

def detectar_tipo_imovel(txt, url=""):
    t=((txt or "") + " " + (url or "")).lower()
    mapa = [
        ("apartamento", ["apartamento", "apartment", "flat", " t1", " t2", " t3", " t4"]),
        ("vivenda", ["vivenda", "moradia", "house", "villa", "vila"]),
        ("terreno", ["terreno", "lote", "land", "plot"]),
        ("escritorio", ["escritório", "escritorio", "office"]),
        ("armazem", ["armazém", "armazem", "warehouse"]),
        ("loja", ["loja", "retail", "shop", "store"]),
        ("quinta", ["quinta", "farm"]),
    ]
    for tipo, kws in mapa:
        if any(k in t for k in kws):
            return tipo
    return norm_tipo_url(url)

def detectar_negocio(txt, url="", fallback=""):
    t=((txt or "") + " " + (url or "") + " " + (fallback or "")).lower()
    if any(k in t for k in ["arrendar", "arrendamento", "alugar", "renda", "rent", "for-rent", "to-rent"]):
        return "arrendamento"
    if any(k in t for k in ["venda", "vender", "comprar", "compra", "sale", "for-sale"]):
        return "compra"
    return fallback or ""


def extrair_headline_completa(soup: BeautifulSoup, fallback: str = "") -> str:
    """Extrai a headline mais completa possível da página de detalhe."""
    candidatos = []
    for sel in ("h1", ".text h2", ".text h3", "h2", "h3"):
        obj = soup.select_one(sel)
        if obj:
            txt = " ".join(obj.stripped_strings)
            txt = limpar_texto(txt)
            if txt:
                candidatos.append(txt)
    for sel in ("meta[property='og:title']", "meta[name='twitter:title']"):
        obj = soup.select_one(sel)
        if obj and obj.get("content"):
            txt = limpar_texto(obj.get("content"))
            if txt:
                candidatos.append(txt)
    if fallback:
        candidatos.append(limpar_texto(fallback))
    if not candidatos:
        return ""
    # Preferir o texto mais informativo, sem duplicações triviais.
    unicos = list(dict.fromkeys(candidatos))
    return max(unicos, key=len)



def derivar_tipologia(texto: str, tipo_imovel: str = "", quartos=None) -> Optional[str]:
    """
    Deriva tipologia a partir de headline/descrição, sem sobrescrever valores já extraídos.
    Retorna Tn para apartamento e Vn para vivenda; para comercial/terreno retorna None.
    """
    tipo = (tipo_imovel or "").lower().strip()
    txt = limpar_texto(texto or "")
    low = txt.lower()

    # T3, T 3, T-3, V4, V 5, T3+1, tipologia T2.
    m = re.search(r"\b([TtVv])\s*[-\s]?\s*(\d{1,2})(?:\s*\+\s*\d{1,2})?\b", txt)
    if m:
        prefixo = m.group(1).upper()
        n = int(m.group(2))
        if 0 <= n <= 20:
            return f"{prefixo}{n}"

    if re.search(r"\b(studio|st[úu]dio|est[úu]dio|kitchenette|kitnet|t0)\b", low):
        return "T0"

    n = None
    m = re.search(r"(\d{1,2})\s*(?:quartos?|dormit[óo]rios?|su[ií]tes?|bedrooms?)\b", low)
    if m:
        n = int(m.group(1))
    elif quartos not in (None, "", 0, "0"):
        try:
            n = int(float(quartos))
        except Exception:
            n = None

    if n is not None and 0 <= n <= 20 and tipo in ("apartamento", "vivenda"):
        return ("T" if tipo == "apartamento" else "V") + str(n)
    return None


def derivar_quartos_texto(tipologia, texto: str, tipo_imovel: str = "") -> Optional[int]:
    """Deriva quartos apenas para imóveis residenciais."""
    tipo = (tipo_imovel or "").lower().strip()
    if tipo not in ("apartamento", "vivenda"):
        return None
    if tipologia:
        m = re.search(r"\d+", str(tipologia))
        if m:
            n = int(m.group())
            return n if 0 <= n <= 20 else None
    m = re.search(r"(\d{1,2})\s*(?:quartos?|dormit[óo]rios?|su[ií]tes?|bedrooms?)\b", (texto or "").lower())
    if m:
        n = int(m.group(1))
        return n if 0 <= n <= 20 else None
    return None


def derivar_area_texto(texto: str) -> Optional[float]:
    """Extrai a primeira área plausível em m² do texto."""
    txt = texto or ""
    # Evita apanhar preços: exige unidade m2/m².
    m = re.search(r"([\d\.\s,]{1,12})\s*m\s*[²2]\b", txt, re.I)
    if m:
        v = norm_num(m.group(1))
        if v and 5 <= float(v) <= 1_000_000:
            return float(v)
    return None

def classificar_segmento_imobiliario(tipo_imovel: str, texto: str = "") -> str:
    t = ((tipo_imovel or "") + " " + (texto or "")).lower()
    if any(k in t for k in ["hotel", "guest house", "hospedagem", "hotelaria"]):
        return "Hotelaria"
    if any(k in t for k in ["misto", "mixed-use", "mixed use", "uso misto"]):
        return "Misto"
    if any(k in t for k in ["industrial", "armazem", "armazém", "warehouse", "fábrica", "fabrica"]):
        return "Industrial"
    if any(k in t for k in ["terreno", "lote", "land", "plot"]):
        return "Terreno"
    if any(k in t for k in ["escritorio", "escritório", "loja", "retail", "comercial", "office", "shop"]):
        return "Comercial"
    return "Residencial"


def gerar_hash_dedupe(d: dict) -> str:
    """Hash anti-duplicação entre fontes diferentes."""
    partes = [
        limpar_texto((d.get("headline") or d.get("titulo_anuncio") or "").lower(), 180),
        limpar_texto(d.get("bairro") or "").lower(),
        limpar_texto(d.get("tipo_negocio") or "").lower(),
        limpar_texto(d.get("tipo_imovel") or "").lower(),
        str(int(float(d.get("preco_akz") or 0))) if d.get("preco_akz") else "0",
        str(int(float(d.get("area_m2") or 0))) if d.get("area_m2") else "0",
    ]
    base = "|".join(partes)
    return hashlib.sha1(base.encode("utf-8", errors="ignore")).hexdigest()


def enriquecer_campos_padrao(d: dict) -> dict:
    """Completa campos novos sem quebrar fontes antigas."""
    headline = limpar_texto(d.get("headline") or d.get("titulo_anuncio") or "")
    d["headline"] = headline
    d["headline_completa"] = limpar_texto(d.get("headline_completa") or headline)
    if not d.get("titulo_anuncio"):
        d["titulo_anuncio"] = d["headline_completa"]
    texto = " ".join([str(d.get(k) or "") for k in ("headline", "headline_completa", "titulo_anuncio", "descricao", "tipo_imovel")])

    # Derivação central — melhora tipologia/quartos/área para todas as fontes sem sobrescrever valores válidos.
    if not d.get("tipologia"):
        d["tipologia"] = derivar_tipologia(texto, d.get("tipo_imovel"), d.get("quartos"))
    if not d.get("quartos"):
        q_der = derivar_quartos_texto(d.get("tipologia"), texto, d.get("tipo_imovel"))
        if q_der is not None:
            d["quartos"] = q_der
    try:
        area_atual = float(d.get("area_m2") or 0)
    except Exception:
        area_atual = 0
    if area_atual <= 0:
        a_der = derivar_area_texto(texto)
        if a_der:
            d["area_m2"] = a_der
            if not d.get("area_util_m2"):
                d["area_util_m2"] = a_der
    try:
        if d.get("preco_akz") and d.get("area_m2") and float(d["area_m2"]) > 0 and not d.get("preco_m2_akz"):
            d["preco_m2_akz"] = round(float(d["preco_akz"]) / float(d["area_m2"]), 2)
    except Exception:
        pass

    if not d.get("segmento_imobiliario"):
        d["segmento_imobiliario"] = classificar_segmento_imobiliario(d.get("tipo_imovel"), texto)
    if d.get("imovel_corporativo") is None or "imovel_corporativo" not in d:
        d["imovel_corporativo"] = 1 if (d.get("tipo_imovel") in ("escritorio", "armazem", "loja")) else 0
    if d.get("preco_disponivel") is None or "preco_disponivel" not in d:
        d["preco_disponivel"] = 1 if (d.get("preco_original") or d.get("preco_akz") or d.get("preco_usd")) else 0
    d.setdefault("email_agente", None)
    d.setdefault("id_zenki", None)
    d.setdefault("url_brochura", None)
    d.setdefault("pdf_brochura", None)
    d["hash_dedupe"] = d.get("hash_dedupe") or gerar_hash_dedupe(d)
    return d


# ─────────────────────────────────────────────────────────────────
# SCRAPER DEDICADO — RE/MAX ANGOLA
# ─────────────────────────────────────────────────────────────────
class ScraperRemax:
    """
    Coletor dedicado para RE/MAX Angola / Multitrust.

    Versão corrigida: Azure Cognitive Search com campos aninhados em `content/*`.
    - filtro correto: content/TransactionTypeUID
    - documentos em response['value'][i]['content']
    - URL pt-AO em ShortLinks[LanguageCode == 'pt-AO']
    - cache do JSON da API para evitar abrir detalhe quando os dados já vêm completos.
    """

    BASE = "https://www.remax-multitrust.co.ao"
    API_SEARCH = f"{BASE}/search/listing-search/docs/search"
    PAGE_SIZE = 24
    IMG_CDN = "https://cdn.gryphtech.com/userimages"

    NEGOCIOS = {
        "compra": 261,
        "arrendamento": 260,
    }

    TIPO_MAP = {
        "apartamento": ["apartamento", "flat", "studio", "apartment"],
        "vivenda": ["moradia", "vivenda", "villa", "house", "detached"],
        "terreno": ["terreno", "land", "lot", "plot", "terrenos"],
        "escritorio": ["escritório", "escritorio", "escritórios", "escritorios", "office"],
        "armazem": ["armazém", "armazem", "armazéns", "armazens", "warehouse", "industrial"],
        "loja": ["loja", "lojas", "retail", "shop", "store", "estabelecimento"],
        "quinta": ["quinta", "farm"],
        "outro": ["prédio", "predio", "hotel", "building", "investimentos"],
    }

    def __init__(self, db):
        self.db = db
        self.ses = self._sessao()
        self._doc_cache = {}

    def _sessao(self):
        s = requests.Session()
        s.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
            "Referer": f"{self.BASE}/listings",
            "Origin": self.BASE,
            "Content-Type": "application/json",
        })
        return s

    def _delay(self):
        time.sleep(random.uniform(CFG["delay_min"], CFG["delay_max"]))

    def _normalizar_url_remax(self, href: str) -> str:
        if not href:
            return ""
        href = str(href).strip()
        if href.startswith("//"):
            href = "https:" + href
        if href.startswith("http"):
            return href.replace("http://", "https://", 1)
        return url_absoluta(self.BASE, "/" + href.lstrip("/"))

    def _content_val(self, content: dict, *keys, default=None):
        """Busca chaves alternativas no JSON da RE/MAX."""
        for key in keys:
            if key in content and content.get(key) not in (None, ""):
                return content.get(key)
        # fallback case-insensitive
        low = {str(k).lower(): k for k in content.keys()}
        for key in keys:
            k = low.get(str(key).lower())
            if k is not None and content.get(k) not in (None, ""):
                return content.get(k)
        return default


    def _iterar_json(self, obj):
        """Percorre dict/list da API RE/MAX para localizar campos aninhados."""
        if isinstance(obj, dict):
            yield obj
            for v in obj.values():
                yield from self._iterar_json(v)
        elif isinstance(obj, list):
            for item in obj:
                yield from self._iterar_json(item)

    def _deep_val(self, content: dict, *keys, default=None):
        """Busca chaves em qualquer nível do JSON, preservando a primeira ocorrência útil."""
        if not isinstance(content, dict):
            return default
        wanted = {str(k).lower() for k in keys}
        for node in self._iterar_json(content):
            low = {str(k).lower(): k for k in node.keys()}
            for key in wanted:
                real = low.get(key)
                if real is not None and node.get(real) not in (None, "", [], {}):
                    return node.get(real)
        return default

    def _deep_int(self, content: dict, *keys) -> Optional[int]:
        v = self._deep_val(content, *keys)
        if isinstance(v, dict):
            for kk in ("Value", "value", "Count", "count", "Number", "number"):
                if kk in v:
                    v = v.get(kk)
                    break
        if isinstance(v, list):
            v = next((x for x in v if x not in (None, "", {}, [])), None)
        try:
            if v not in (None, ""):
                return int(float(str(v).replace(",", ".")))
        except Exception:
            pass
        return None

    def _deep_float(self, content: dict, *keys) -> Optional[float]:
        v = self._deep_val(content, *keys)
        if isinstance(v, dict):
            for kk in ("Value", "value", "Amount", "amount", "Total", "total"):
                if kk in v:
                    v = v.get(kk)
                    break
        if isinstance(v, list):
            v = next((x for x in v if x not in (None, "", {}, [])), None)
        try:
            if v not in (None, ""):
                return float(str(v).replace(" ", "").replace(",", "."))
        except Exception:
            pass
        return None

    def _json_texto_remax(self, content: dict, limite: int = 12000) -> str:
        """Concatena textos úteis do JSON sem depender de nomes exactos dos campos."""
        partes = []
        def walk(obj):
            if len(" ".join(partes)) > limite:
                return
            if isinstance(obj, dict):
                for k, v in obj.items():
                    lk = str(k).lower()
                    # Evitar despejar arrays grandes de multimédia/base64/links técnicos.
                    if any(x in lk for x in ("multimedia", "image", "photo", "thumbnail", "shortlinks")):
                        continue
                    walk(v)
            elif isinstance(obj, list):
                for item in obj[:50]:
                    walk(item)
            elif isinstance(obj, (str, int, float)):
                txt = limpar_texto(obj)
                if txt and len(txt) <= 500:
                    partes.append(txt)
        walk(content)
        return limpar_texto(" ".join(partes), limite=limite)

    def _url_partes_remax(self, url: str) -> dict:
        """Extrai partes semântpicas da URL /pt-ao/angariacoes/tipo/negocio/bairro/.../id."""
        parts = [p for p in urlarse(url).path.strip("/").split("/") if p]
        out = {"tipo_slug": "", "negocio_slug": "", "bairro_slug": "", "slug_titulo": "", "mls_slug": ""}
        try:
            idx = parts.index("angariacoes")
            out["tipo_slug"] = parts[idx + 1] if len(parts) > idx + 1 else ""
            out["negocio_slug"] = parts[idx + 2] if len(parts) > idx + 2 else ""
            out["bairro_slug"] = parts[idx + 3] if len(parts) > idx + 3 else ""
            out["mls_slug"] = parts[-1] if parts else ""
            meio = parts[idx + 4:-1]
            out["slug_titulo"] = " ".join(meio).replace("-", " ").strip()
        except ValueError:
            pass
        return out

    def _title_case_slug(self, txt: str) -> str:
        txt = limpar_texto(str(txt or "").replace("-", " ").replace("_", " "))
        mapa = {"ingombotas": "Ingombota", "escritorios": "Escritório", "lojas": "Loja"}
        return mapa.get(txt.lower(), txt.title()) if txt else ""

    def _titulo_parece_id(self, txt: str) -> bool:
        t = limpar_texto(txt)
        if not t:
            return True
        # Ex.: "126100110 122" ou "126100110-122"
        if re.fullmatch(r"\d{5,}\s*[- ]\s*\d+", t):
            return True
        nums = re.sub(r"\D", "", t)
        return len(nums) >= max(6, int(len(t.replace(" ", "")) * 0.7))

    def _tipo_label_remax(self, tipo_imovel: str, url: str = "") -> str:
        labels = {
            "apartamento": "Apartamento", "vivenda": "Moradia", "terreno": "Terreno",
            "escritorio": "Escritório", "armazem": "Armazém", "loja": "Loja",
            "quinta": "Quinta", "outro": "Imóvel",
        }
        if tipo_imovel in labels:
            return labels[tipo_imovel]
        tipo_slug = self._url_partes_remax(url).get("tipo_slug", "")
        return self._title_case_slug(tipo_slug) or "Imóvel"

    def _montar_headline_remax(self, content: dict, url: str, tipo_imovel: str, negocio: str,
                               bairro: str, quartos: Optional[int], tipologia: Optional[str], desc: str) -> str:
        """Cria headline comercial quando a API devolve apenas MLSID/ref."""
        campos_titulo = [
            "Title", "ListingTitle", "PropertyTitle", "MarketingTitle", "MarketingHeadline", "Headline",
            "Name", "ListingName", "PropertyName", "DisplayTitle", "WebTitle", "SeoTitle", "ShortTitle",
            "Caption", "PublicTitle", "AddressLine1", "DisplayAddress", "FullAddress", "StreetAddress",
        ]
        candidatos = []
        for k in campos_titulo:
            v = self._deep_val(content, k)
            if isinstance(v, str):
                vv = limpar_texto(v)
                if vv and not self._titulo_parece_id(vv) and len(vv) >= 5:
                    candidatos.append(vv)
        partes_url = self._url_partes_remax(url)
        if partes_url.get("slug_titulo") and not self._titulo_parece_id(partes_url["slug_titulo"]):
            candidatos.append(self._title_case_slug(partes_url["slug_titulo"]))
        # Evitar descrições longas como headline.
        candidatos = [c for c in dict.fromkeys(candidatos) if len(c) <= 140]
        if candidatos:
            return max(candidatos, key=len)
        tipo_label = self._tipo_label_remax(tipo_imovel, url)
        tip_txt = tipologia or ""
        if not tip_txt and quartos and tipo_imovel in ("apartamento", "vivenda"):
            tip_txt = ("T" if tipo_imovel == "apartamento" else "V") + str(quartos)
        neg_txt = "para arrendamento" if negocio == "arrendamento" else "à venda"
        loc = bairro if bairro and bairro != "Não informado" else self._title_case_slug(partes_url.get("bairro_slug", ""))
        base = " ".join([tipo_label, tip_txt, neg_txt]).strip()
        return limpar_texto(f"{base} em {loc}" if loc else base)

    def _as_content(self, doc) -> dict:
        """Normaliza response['value'][i]['content']; aceita dict ou string JSON."""
        if not isinstance(doc, dict):
            return {}
        content = doc.get("content") or doc.get("Content") or doc
        if isinstance(content, str):
            try:
                content = json.loads(content)
            except Exception:
                return {}
        return content if isinstance(content, dict) else {}

    def _url_de_doc(self, content: dict) -> Optional[str]:
        links = self._content_val(content, "ShortLinks", "shortLinks", default=[]) or []
        if isinstance(links, str):
            try:
                links = json.loads(links)
            except Exception:
                links = []
        if isinstance(links, list):
            # preferir pt-AO
            for lnk in links:
                if isinstance(lnk, dict):
                    lang = str(lnk.get("LanguageCode") or lnk.get("languageCode") or "")
                    short = lnk.get("ShortLink") or lnk.get("shortLink") or lnk.get("Url") or lnk.get("url") or ""
                    if lang.lower() == "pt-ao" and short:
                        return self._normalizar_url_remax(short)
            for lnk in links:
                if isinstance(lnk, dict):
                    short = lnk.get("ShortLink") or lnk.get("shortLink") or lnk.get("Url") or lnk.get("url") or ""
                    if short and ("angariacoes" in short or "listing" in short.lower()):
                        return self._normalizar_url_remax(short)
        # fallback: procurar em qualquer string do documento
        for v in content.values():
            if isinstance(v, str) and "/angariacoes/" in v:
                return self._normalizar_url_remax(v)
        return ""

    def _tipo_imovel_de_url(self, url: str, content: Optional[dict] = None) -> str:
        content = content or {}
        texto = " ".join(str(x or "") for x in [
            self._content_val(content, "PropertyTypeName", "PropertyType", "PropertyCategory", "ListingTypeName"),
            self._content_val(content, "Title", "ListingTitle", "Headline"),
            urlparse(url).path.replace("-", " "),
        ]).lower()
        for canonical, aliases in self.TIPO_MAP.items():
            if any(a.lower() in texto for a in aliases):
                return canonical
        return detectar_tipo_imovel(texto, url)

    def _negocio_de_uid(self, uid) -> str:
        try:
            uid = int(uid)
        except (TypeError, ValueError):
            return ""
        return "compra" if uid == 261 else ("arrendamento" if uid == 260 else "")

    def _url_imagem(self, content: dict) -> Optional[str]:
        multimedia = self._content_val(content, "Multimedia", "multimedia", "Images", "images", default=[]) or []
        if isinstance(multimedia, str):
            try:
                multimedia = json.loads(multimedia)
            except Exception:
                multimedia = []
        if isinstance(multimedia, list):
            imgs = [m for m in multimedia if isinstance(m, dict) and (m.get("FileName") or m.get("filename") or m.get("Url") or m.get("url"))]
            if imgs:
                def ordem(m):
                    try:
                        return int(m.get("Order") or m.get("order") or 9999)
                    except Exception:
                        return 9999
                imgs = sorted(imgs, key=ordem)
                first = imgs[0]
                direct = first.get("Url") or first.get("url")
                if direct:
                    return self._normalizar_url_remax(direct)
                fname = first.get("FileName") or first.get("filename")
                region_id = self._content_val(content, "RegionId", "RegionID", "regionId", default="")
                if fname and region_id:
                    return f"{self.IMG_CDN}/{region_id}/LargeWM/{fname}"
        return None

    def _salvar_debug_json(self, data, pagina: int, negocio: str):
        try:
            d = CFG["data_raw"]
            d.mkdir(parents=True, exist_ok=True)
            f = d / f"remax_api_debug_{negocio}_p{pagina}_{datetime.now():%Y%m%d_%H%M%S}.json"
            f.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            log.debug(f"[RE/MAX] JSON debug salvo: {f.name}")
        except Exception as exc:
            log.debug(f"[RE/MAX] Não foi possível salvar JSON debug: {exc}")

    def _payload(self, transaction_uid: int, pagina: int) -> dict:
        skip = max(0, int(pagina) - 1) * self.PAGE_SIZE
        return {
            "search": "angola luanda",
            "top": self.PAGE_SIZE,
            "skip": skip,
            "count": True,
            "queryType": "simple",
            "filter": (
                f"content/TransactionTypeUID eq {int(transaction_uid)}"
                " and content/IsFindable eq true"
                " and content/IsViewable eq true"
            ),
        }

    def _payloads_azure_search(self, transaction_uid: int, pagina: int) -> List[dict]:
        """Mantido por compatibilidade; tenta variações seguras se a primeira falhar."""
        base = self._payload(transaction_uid, pagina)
        return [
            base,
            {**base, "search": "*"},
            {**base, "filter": f"content/TransactionTypeUID eq {int(transaction_uid)}"},
            {**base, "filter": (
                f"content/TransactionTypeUID eq {int(transaction_uid)}"
                " and content/CountryCode eq 'AO'"
            )},
        ]

    def _post_api(self, payload: dict, pagina: int, negocio: str) -> Optional[dict]:
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.post(self.API_SEARCH, json=payload, timeout=CFG["timeout"])
                if r.status_code >= 400:
                    corpo = limpar_texto(r.text, limite=800)
                    log.warning(f"[RE/MAX] HTTP {r.status_code} | tentativa {tentativa} | body: {corpo}")
                    if CFG.get("salvar_html"):
                        try:
                            CFG["data_raw"].mkdir(parents=True, exist_ok=True)
                            (CFG["data_raw"] / f"remax_http_{r.status_code}_{negocio}_p{pagina}.txt").write_text(r.text, encoding="utf-8")
                        except Exception:
                            pass
                    if tentativa < CFG["max_retries"]:
                        time.sleep(2 * tentativa)
                    continue
                data = r.json()
                if CFG.get("salvar_html") or pagina == 1:
                    self._salvar_debug_json(data, pagina, negocio)
                return data
            except requests.RequestException as e:
                log.warning(f"[RE/MAX] Rede {e} | tentativa {tentativa}")
            except ValueError as e:
                log.warning(f"[RE/MAX] Resposta não JSON | tentativa {tentativa}: {e}")
            if tentativa < CFG["max_retries"]:
                time.sleep(2 * tentativa)
        return None

    def _get_html(self, url: str) -> Optional[BeautifulSoup]:
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.get(url, timeout=CFG["timeout"])
                r.raise_for_status()
                if CFG["salvar_html"]:
                    slug = re.sub(r"[^\w]", "_", url)[-100:]
                    d = CFG["html_debug"]
                    d.mkdir(parents=True, exist_ok=True)
                    (d / f"remax_{slug}.html").write_text(r.text, encoding="utf-8")
                return BeautifulSoup(r.text, "lxml")
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 404:
                    return None
                log.warning(f"[RE/MAX] HTTP {e} | tentativa {tentativa}")
            except requests.RequestException as e:
                log.warning(f"[RE/MAX] Rede {e} | tentativa {tentativa}")
            if tentativa < CFG["max_retries"]:
                time.sleep(2 * tentativa)
        return None

    def _extrair_total_paginas(self, data) -> int:
        try:
            total = int(data.get("@odata.count") or 0)
            if total:
                return max(1, math.ceil(total / self.PAGE_SIZE))
        except Exception:
            pass
        try:
            n = len(data.get("value") or [])
            return 1 if n < self.PAGE_SIZE else CFG["max_paginas"]
        except Exception:
            return 1

    def _extrair_links_api(self, transaction_uid: int, pagina: int, negocio: str) -> Tuple[List[str], int]:
        urls = []
        total_paginas = 1
        for payload in self._payloads_azure_search(transaction_uid, pagina):
            data = self._post_api(payload, pagina, negocio)
            if not data:
                continue
            total_paginas = max(total_paginas, self._extrair_total_paginas(data))
            for doc in data.get("value") or []:
                content = self._as_content(doc)
                if not content:
                    continue
                u = self._url_de_doc(content)
                if not u or "/angariacoes/" not in u:
                    continue
                # Filtro defensivo por negócio, caso algum payload alternativo traga misturado.
                uid_doc = self._content_val(content, "TransactionTypeUID", "transactionTypeUID")
                if uid_doc and self._negocio_de_uid(uid_doc) and self._negocio_de_uid(uid_doc) != negocio:
                    continue
                urls.append(u)
                self._doc_cache[u] = content
            urls = list(dict.fromkeys(urls))
            if urls:
                log.debug(f"[RE/MAX] Payload OK: {len(urls)} URLs na pág {pagina}")
                break
        return urls, total_paginas

    def _extrair_links_html_fallback(self, transaction_uid: int) -> List[str]:
        soup = self._get_html(f"{self.BASE}/listings?ListingClass=-1&TransactionTypeUID={transaction_uid}")
        if not soup:
            return []
        html = str(soup).replace("\\/", "/").replace("&amp;", "&")
        urls = []
        for patt in [
            r"https?://www\.remax-multitrust\.co\.ao/pt-ao/angariacoes/[^\"\'\<\>\s]+",
            r"/pt-ao/angariacoes/[^\"\'\<\>\s]+",
        ]:
            for m in re.findall(patt, html, flags=re.I):
                urls.append(self._normalizar_url_remax(m))
        for a in soup.select("a[href*='/angariacoes/']"):
            href = a.get("href") or ""
            urls.append(self._normalizar_url_remax(href))
        return list(dict.fromkeys([u for u in urls if u and "/angariacoes/" in u]))

    def _normalizar_doc_api(self, url: str, content: dict, negocio_fallback: str) -> Optional[dict]:
        """Normaliza um documento da API RE/MAX para o modelo canónico PIIA.

        Correção v2.11:
        - Não usa MLSID como headline.
        - Reconstrói headline quando a API só devolve referência.
        - Deduz tipologia a partir de quartos e tipo de imóvel.
        - Lê quartos/WC/área/preço por busca recursiva no JSON.
        """
        if not content:
            return None

        partes_url = self._url_partes_remax(url)
        texto_json = self._json_texto_remax(content)

        uid = self._deep_val(content, "TransactionTypeUID", "transactionTypeUID")
        negocio = self._negocio_de_uid(uid) or detectar_negocio(url, fallback=negocio_fallback) or negocio_fallback
        tipo_imovel = self._tipo_imovel_de_url(url, content)

        mls = self._deep_val(content, "MLSID", "MlsId", "ListingId", "ListingID", "Reference", "ReferenceNumber", "Id", "id")
        id_fonte = str(mls).strip() if mls else (partes_url.get("mls_slug") or hashlib.md5(url.encode("utf-8")).hexdigest()[:16])

        desc = limpar_texto(self._deep_val(
            content, "Description", "PublicDescription", "Remarks", "MarketingDescription", "DescriptionPT", "LongDescription", default=""
        ), limite=5000)
        if not desc:
            desc = limpar_texto(texto_json, limite=5000)

        # Preço e moeda
        preco_original = self._deep_float(
            content, "ListingPrice", "Price", "ListPrice", "AskingPrice", "SalePrice", "RentPrice",
            "DisplayPrice", "OriginalPrice", "MonthlyPrice", "PriceValue"
        )
        # RE/MAX Angola publica preços em AOA/Kz por defeito.
        # Só tratar como USD quando o campo de moeda indicar explicitamente dólar.
        moeda_raw = str(self._deep_val(content, "Currency", "CurrencyCode", "PriceCurrency", "CurrencyISO", default="") or "").upper()
        if any(k in moeda_raw for k in ("USD", "US$", "DOLAR", "DÓLAR", "DOLLAR")):
            moeda = "USD"
        else:
            moeda = "AKZ"
        if preco_original and preco_original <= 0:
            preco_original = None
        if moeda == "USD":
            preco_usd = preco_original
            preco_akz = usd2akz(preco_usd) if preco_usd else None
        else:
            preco_akz = preco_original
            preco_usd = akz2usd(preco_akz) if preco_akz else None

        # Área — evitar gravar 0 como área válida.
        area_m2 = self._deep_float(
            content, "TotalArea", "LivingArea", "TotalAreaSQM", "AreaSQM", "GrossArea", "NetArea",
            "BuiltArea", "PropertyArea", "Area", "LotSize", "LandArea", "Size"
        )
        if area_m2 is not None and area_m2 <= 0:
            area_m2 = None
        unit_txt = str(self._deep_val(content, "AreaUnit", "MeasureUnit", "TotalAreaUnit", "Unit", default="") or "").lower()
        if area_m2 and ("ft" in unit_txt or "sqft" in unit_txt):
            area_m2 = round(area_m2 * 0.092903, 2)
        preco_m2 = round(preco_akz / area_m2, 2) if preco_akz and area_m2 and area_m2 > 0 else None

        # Quartos e WC — vários nomes possíveis no índice RE/MAX.
        quartos = self._deep_int(
            content, "Bedrooms", "Bedroom", "NumberOfBedrooms", "TotalBedrooms", "BedroomsTotal",
            "BedRooms", "Beds", "Rooms", "TotalRooms", "NumberOfRooms", "NoOfBedrooms"
        )
        wc = self._deep_int(
            content, "Bathrooms", "Bathroom", "NumberOfBathrooms", "TotalBathrooms", "BathroomsTotal",
            "Baths", "WCs", "WC", "Toilets", "NumberOfToilets", "NoOfBathrooms"
        )

        # Tipologia: primeiro texto explícito; depois Tn/Vn por quartos.
        tipologia = norm_tipologia(" ".join(str(x or "") for x in [
            self._deep_val(content, "Typology", "Tipologia", "PropertyTypology", "BedroomType", "UnitType", default=""),
            texto_json[:1500],
            urlparse(url).path.replace("-", " ")
        ]))
        if quartos is None and tipologia:
            m = re.search(r"\d+", tipologia)
            quartos = int(m.group()) if m else None
        if not tipologia and quartos and tipo_imovel in ("apartamento", "vivenda"):
            tipologia = ("T" if tipo_imovel == "apartamento" else "V") + str(quartos)

        loc_text = " ".join(str(x or "") for x in [
            self._deep_val(content, "Location", "Address", "FullAddress", "DisplayAddress", "District", "City", "Region", "Neighborhood", "Zone"),
            texto_json[:2000],
            urlparse(url).path.replace("-", " "),
        ])
        bairro, lat, lon = detectar_bairro(loc_text)
        if not bairro:
            bairro = limpar_texto(self._deep_val(content, "District", "Neighborhood", "Zone", "City", "Region", default=""))
        if not bairro:
            bairro = self._title_case_slug(partes_url.get("bairro_slug", "")) or "Não informado"
        try:
            lat_v = self._deep_val(content, "Latitude", "Lat")
            lon_v = self._deep_val(content, "Longitude", "Lng", "Lon")
            lat = float(lat_v) if lat_v not in (None, "") else lat
            lon = float(lon_v) if lon_v not in (None, "") else lon
        except Exception:
            pass

        headline = self._montar_headline_remax(content, url, tipo_imovel, negocio, bairro, quartos, tipologia, desc)
        headline_completa = headline
        if area_m2 and tipo_imovel in ("terreno", "armazem", "loja", "escritorio") and str(int(area_m2)) not in headline:
            headline_completa = limpar_texto(f"{headline} — {area_m2:g} m²")

        agent = self._deep_val(content, "AgentName", "AssociateName", "ListingAgentName", "BrokerName", "OfficeName", "ConsultantName", "AgentFullName")
        phone = extrair_telefone(" ".join(str(x or "") for x in [
            self._deep_val(content, "AgentPhone", "AssociatePhone", "Phone", "MobilePhone", "OfficePhone", "CellPhone", "WhatsApp"),
            texto_json[:4000],
        ]))
        email = self._deep_val(content, "AgentEmail", "AssociateEmail", "Email", "ConsultantEmail")
        img_url = self._url_imagem(content)

        texto_flags = (headline + " " + desc + " " + texto_json[:6000]).lower()
        return {
            "fonte": "RE/MAX Angola",
            "id_fonte": id_fonte,
            "url": url,
            "id_angocasa": None,
            "tipo_negocio": negocio,
            "pais": "Angola",
            "provincia": "Luanda",
            "bairro": bairro,
            "lat": lat,
            "lon": lon,
            "headline": headline,
            "headline_completa": headline_completa,
            "titulo_anuncio": headline,
            "tipo_imovel": tipo_imovel,
            "tipologia": tipologia,
            "estado": limpar_texto(self._deep_val(content, "PropertyStatus", "Status", "ListingStatus", default="")) or None,
            "area_m2": area_m2,
            "area_util_m2": area_m2,
            "area_bruta_m2": area_m2,
            "area_lote_m2": area_m2 if tipo_imovel == "terreno" else None,
            "preco_original": preco_original,
            "moeda": moeda,
            "preco_akz": preco_akz,
            "preco_usd": preco_usd,
            "preco_m2_akz": preco_m2,
            "quartos": quartos,
            "wc": wc,
            "mobilado": tem_kw(texto_flags, ["mobilado", "mobiliado", "furnished"]),
            "garagem": tem_kw(texto_flags, ["garagem", "estacionamento", "parking", "garage"]),
            "piscina": tem_kw(texto_flags, ["piscina", "pool"]),
            "gerador": tem_kw(texto_flags, ["gerador", "generator"]),
            "reservatorio": tem_kw(texto_flags, ["reservatório", "reservatorio", "tanque"]),
            "seguranca": tem_kw(texto_flags, ["segurança", "seguranca", "vigilância", "vigilancia", "security", "portaria"]),
            "churrasqueira": tem_kw(texto_flags, ["churrasqueira", "barbecue", "bbq"]),
            "ar_condicionado": tem_kw(texto_flags, ["ar condicionado", "ar-condicionado", "air conditioning", "a/c"]),
            "internet": tem_kw(texto_flags, ["internet", "fibra", "wifi", "wi-fi"]),
            "condominio_fechado": tem_kw(texto_flags, ["condomínio fechado", "condominio fechado", "condomínio privado", "gated", "portaria"]),
            "publicado_em": self._deep_val(content, "DateListed", "CreatedDate", "ModifiedDate", "ListingDate", "PublishDate"),
            "descricao": desc,
            "anunciante": limpar_texto(agent) if agent else None,
            "telefone_whatsapp": phone,
            "email_agente": limpar_texto(email) if email else None,
            "url_imagem": img_url,
            "raw_json": json.dumps({"fonte": "RE/MAX Angola", "id_fonte": id_fonte, "content": content}, ensure_ascii=False, default=str)[:15000],
        }

    def extrair_detalhe(self, url: str, negocio_fallback: str) -> Optional[dict]:
        content = self._doc_cache.get(url)
        if content:
            dados_api = self._normalizar_doc_api(url, content, negocio_fallback)
            if dados_api:
                return dados_api

        # Fallback antigo por HTML, caso a API traga link mas poucos dados.
        soup = self._get_html(url)
        if not soup:
            return None
        self._delay()
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        main = soup.find("main") or soup
        text = limpar_texto(main.get_text(" ", strip=True), limite=18000)
        h1 = soup.select_one("h1") or soup.select_one("h2")
        headline = limpar_texto(h1.get_text(" ", strip=True)) if h1 else limpar_texto(urlparse(url).path.strip("/").split("/")[-1].replace("-", " ").title())
        preco_extra, moeda = extrair_preco_texto(text)
        if moeda == "USD":
            preco_usd = preco_extra
            preco_akz = usd2akz(preco_usd) if preco_usd else None
        else:
            preco_akz = preco_extra
            preco_usd = akz2usd(preco_akz) if preco_akz else None
        area_m2 = norm_area(text)
        preco_m2 = round(preco_akz / area_m2, 2) if preco_akz and area_m2 else None
        bairro, lat, lon = detectar_bairro(text + " " + headline)
        tipo_imovel = self._tipo_imovel_de_url(url, {})
        tipologia = norm_tipologia(text + " " + headline)
        telefone = extrair_telefone(text)
        return {
            "fonte": "RE/MAX Angola", "id_fonte": hashlib.md5(url.encode("utf-8")).hexdigest()[:16], "url": url,
            "id_angocasa": None, "tipo_negocio": negocio_fallback, "pais": "Angola", "provincia": "Luanda",
            "bairro": bairro or "Não informado", "lat": lat, "lon": lon, "headline": headline,
            "headline_completa": headline, "titulo_anuncio": headline, "tipo_imovel": tipo_imovel, "tipologia": tipologia,
            "estado": None, "area_m2": area_m2, "area_util_m2": area_m2, "area_bruta_m2": area_m2,
            "area_lote_m2": None, "preco_original": preco_extra, "moeda": moeda, "preco_akz": preco_akz,
            "preco_usd": preco_usd, "preco_m2_akz": preco_m2, "quartos": None, "wc": None,
            "mobilado": tem_kw(text, ["mobilado", "mobiliado", "furnished"]),
            "garagem": tem_kw(text, ["garagem", "estacionamento", "parking"]),
            "piscina": tem_kw(text, ["piscina", "pool"]),
            "gerador": tem_kw(text, ["gerador", "generator"]),
            "reservatorio": tem_kw(text, ["reservatório", "reservatorio", "tanque"]),
            "seguranca": tem_kw(text, ["segurança", "seguranca", "vigilância", "portaria"]),
            "churrasqueira": tem_kw(text, ["churrasqueira", "barbecue", "bbq"]),
            "ar_condicionado": tem_kw(text, ["ar condicionado", "ar-condicionado", "a/c"]),
            "internet": tem_kw(text, ["internet", "fibra", "wifi"]),
            "condominio_fechado": tem_kw(text, ["condomínio fechado", "condominio fechado", "gated", "portaria"]),
            "publicado_em": None, "descricao": text[:5000], "anunciante": None,
            "telefone_whatsapp": telefone, "url_imagem": None,
            "raw_json": json.dumps({"fonte": "RE/MAX Angola", "url": url, "fallback": "html"}, ensure_ascii=False),
        }

    def recolher(self, forcar=False) -> Tuple[List[dict], bool]:
        resultados = []
        urls_vistas = set()
        limite = CFG.get("max_anuncios_por_fonte")
        limite = int(limite) if limite else None

        for negocio_nome, uid in self.NEGOCIOS.items():
            if limite and len(resultados) >= limite:
                break
            log.info(f"  ▶ RE/MAX Angola | {negocio_nome}")
            pagina = self.db.obter_checkpoint("RE/MAX Angola", negocio_nome)
            if pagina > 1:
                log.info(f"    [RE/MAX] Retomando do checkpoint: pág {pagina}")
            paginas_sem_links = 0
            while pagina <= CFG["max_paginas"]:
                if limite and len(resultados) >= limite:
                    break
                urls_pag, total_pag = self._extrair_links_api(uid, pagina, negocio_nome)
                if not urls_pag and pagina == 1:
                    log.warning("    [RE/MAX] API sem URLs — tentando fallback HTML da primeira página")
                    urls_pag = self._extrair_links_html_fallback(uid)
                    total_pag = 1

                urls_pag = list(dict.fromkeys([u for u in urls_pag if u and "/angariacoes/" in u]))
                if not urls_pag:
                    paginas_sem_links += 1
                    log.info(f"    RE/MAX pág {pagina}: sem links úteis.")
                    if paginas_sem_links >= 2:
                        break
                    pagina += 1
                    continue
                paginas_sem_links = 0

                extraidos = 0
                for u in urls_pag:
                    if limite and len(resultados) >= limite:
                        break
                    if u in urls_vistas:
                        continue
                    urls_vistas.add(u)
                    if not forcar and self.db.anuncio_processado_recente(u):
                        log.debug(f"    Ignorado (já processado recentemente): {u}")
                        continue
                    try:
                        self._delay()
                        dados = self.extrair_detalhe(u, negocio_nome)
                        if dados:
                            resultados.append(dados)
                            extraidos += 1
                    except Exception as e:
                        log.debug(f"    Falha detalhe RE/MAX: {u} | {e}")
                        self.db.reg_erro(u, "remax_detalhe", str(e))

                log.info(f"    RE/MAX pág {pagina}/{total_pag}: {extraidos} extraídos (acum.: {len(resultados)})")
                self.db.salvar_checkpoint("RE/MAX Angola", negocio_nome, pagina + 1)
                if pagina >= total_pag:
                    self.db.limpar_checkpoint("RE/MAX Angola", negocio_nome)
                    break
                pagina += 1
                self._delay()

        log.info(f"  ✔ RE/MAX Angola: {len(resultados)} anúncios total")
        return resultados, True


# ─────────────────────────────────────────────────────────────────
# SCRAPER DEDICADO — MIA (mia.africa)
# ─────────────────────────────────────────────────────────────────

class ScraperMia:
    """
    Coletor dedicado para MIA — Mercado Imobiliário de Angola.

    Estratégia:
      - Listagem via API Supabase RPC search_properties.
      - Detalhe via REST /properties?id=eq.{uuid}&select=*.
      - Normalização para o modelo canónico PIIA.
    """

    BASE_SITE    = "https://mia.africa"
    SUPABASE_URL = "https://oxfqtppnalbtjefhuyzc.supabase.co"
    ANON_KEY     = (
        "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
        ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im94ZnF0cHBuYWxidGplZmh1eXpjIiwi"
        "cm9sZSI6ImFub24iLCJpYXQiOjE3NjgxOTA0MzIsImV4cCI6MjA4Mzc2NjQzMn0"
        ".IA2xh24fQc1SWYywgOsPBs2_1E02MUDmYR9MY2FTveA"
    )
    STORAGE_BASE = f"{SUPABASE_URL}/storage/v1/object/public/property-media"
    RPC_SEARCH   = f"{SUPABASE_URL}/rest/v1/rpc/search_properties"
    REST_PROPS   = f"{SUPABASE_URL}/rest/v1/properties"
    PER_PAGE     = 12

    TIPO_MAP = {
        "apartment":  "apartamento",
        "villa":      "vivenda",
        "house":      "vivenda",
        "land":       "terreno",
        "office":     "escritorio",
        "warehouse":  "armazem",
        "commercial": "loja",
        "retail":     "loja",
        "hotel":      "outro",
        "farm":       "quinta",
    }

    NEGOCIO_MAP = {
        "sale":      "compra",
        "rent":      "arrendamento",
        "sale_rent": "compra",
    }

    FURNISHED_MAP = {
        "furnished":       1,
        "semi-furnished":  1,
        "unfurnished":     0,
    }

    def __init__(self, db):
        self.db  = db
        self.ses = self._sessao()

    def _sessao(self) -> requests.Session:
        s = requests.Session()
        s.headers.update({
            "apikey":          self.ANON_KEY,
            "authorization":   f"Bearer {self.ANON_KEY}",
            "content-type":    "application/json",
            "content-profile": "public",
            "x-client-info":   "supabase-js-web/2.90.1",
            "Origin":          self.BASE_SITE,
            "Referer":         f"{self.BASE_SITE}/",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        })
        return s

    def _delay(self):
        time.sleep(random.uniform(CFG["delay_min"], CFG["delay_max"]))

    def _payload_listagem(self, listing_type: str, pagina: int, city: str = "Luanda", district: str = "", property_type: str = "") -> dict:
        return {
            "p_filters": {
                "city":         city,
                "district":     district,
                "propertyType": property_type,
                "listingType":  listing_type,
                "priceMin":     0,
                "priceMax":     50_000_000_000,
                "bedrooms":     "any",
                "bathrooms":    "any",
                "furnished":    False,
                "verified":     False,
                "agencyId":     "",
            },
            "p_sort":     "newest",
            "p_page":     pagina,
            "p_per_page": self.PER_PAGE,
        }

    def _salvar_debug_json(self, nome: str, dados):
        if not CFG.get("salvar_html"):
            return
        try:
            d = CFG["data_raw"]
            d.mkdir(parents=True, exist_ok=True)
            arq = d / nome
            arq.write_text(json.dumps(dados, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        except Exception as exc:
            log.debug(f"[MIA] Não foi possível salvar JSON debug: {exc}")

    def _chamar_rpc(self, listing_type: str, pagina: int, **filtros) -> Optional[dict]:
        payload = self._payload_listagem(listing_type, pagina, **filtros)
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.post(self.RPC_SEARCH, json=payload, timeout=CFG["timeout"])
                r.raise_for_status()
                dados = r.json()
                if pagina == 1:
                    self._salvar_debug_json(f"mia_rpc_{listing_type}_p{pagina}.json", dados)
                if isinstance(dados, dict) and "data" in dados:
                    return dados
                if isinstance(dados, list):
                    return {"data": dados, "total_count": len(dados)}
                log.warning(f"[MIA] Resposta inesperada p{pagina}: {str(dados)[:200]}")
                return None
            except requests.exceptions.HTTPError as e:
                log.warning(f"[MIA] HTTP {e} | listagem p{pagina} t{tentativa}")
            except requests.RequestException as e:
                log.warning(f"[MIA] Rede {e} | listagem p{pagina} t{tentativa}")
            except ValueError as e:
                log.warning(f"[MIA] JSON inválido | listagem p{pagina} t{tentativa}: {e}")
            if tentativa < CFG["max_retries"]:
                time.sleep(5 * tentativa)
        return None

    def _chamar_detalhe(self, prop_id: str) -> Optional[dict]:
        url = f"{self.REST_PROPS}?id=eq.{prop_id}&select=*"
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.get(url, timeout=CFG["timeout"])
                r.raise_for_status()
                dados = r.json()
                if isinstance(dados, list) and dados:
                    return dados[0]
                return None
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 404:
                    return None
                log.warning(f"[MIA] HTTP {e} | detalhe {prop_id} t{tentativa}")
            except requests.RequestException as e:
                log.warning(f"[MIA] Rede {e} | detalhe {prop_id} t{tentativa}")
            if tentativa < CFG["max_retries"]:
                time.sleep(5 * tentativa)
        return None

    def _url_imovel(self, prop_id: str) -> str:
        return f"{self.BASE_SITE}/imovel/{prop_id}"

    def _url_imagem(self, agent_id: str, prop_id: str, cover_path: str) -> Optional[str]:
        if cover_path:
            if str(cover_path).startswith("http"):
                return str(cover_path)
            return f"{self.STORAGE_BASE}/{cover_path}"
        if agent_id and prop_id:
            return f"{self.STORAGE_BASE}/{agent_id}/{prop_id}/"
        return None

    def _preco_e_moeda(self, det: dict) -> tuple:
        moeda_raw = (det.get("price_currency") or "AOA").upper()
        moeda = "USD" if moeda_raw in ("USD", "US$") else "AKZ"
        preco = det.get("price") or det.get("sale_price") or det.get("rent_price")
        sob_consulta = det.get("sale_price_on_consultation") or det.get("rent_price_on_consultation")
        if preco in (None, "") or sob_consulta:
            return None, moeda, None, None
        try:
            preco = float(preco)
        except (TypeError, ValueError):
            return None, moeda, None, None
        if moeda == "USD":
            return preco, "USD", usd2akz(preco), preco
        return preco, "AKZ", preco, akz2usd(preco)

    def _tipo_negocio(self, det: dict) -> str:
        lt = (det.get("listing_type") or "").lower()
        is_sale = bool(det.get("is_for_sale", False))
        is_rent = bool(det.get("is_for_rent", False))
        if lt in self.NEGOCIO_MAP:
            return self.NEGOCIO_MAP[lt]
        if is_rent and not is_sale:
            return "arrendamento"
        return "compra"

    def _tipo_imovel(self, det: dict) -> str:
        pt = (det.get("property_type") or "").lower()
        return self.TIPO_MAP.get(pt, detectar_tipo_imovel(det.get("title", ""), ""))

    def _bairro(self, det: dict) -> tuple:
        district = limpar_texto(det.get("district") or "")
        address  = limpar_texto(det.get("address")  or "")
        texto    = f"{district} {address} {det.get('title', '')}"
        b, lat, lon = detectar_bairro(texto)
        if b:
            return b, lat, lon
        bairro = district or address.split(",")[0].strip() or "Não informado"
        lat    = det.get("latitude")
        lon    = det.get("longitude")
        return bairro, lat, lon

    def _features(self, det: dict) -> dict:
        feats_raw = det.get("features") or []
        if isinstance(feats_raw, str):
            feats = [feats_raw.lower()]
        else:
            feats = [str(f).lower() for f in feats_raw]
        feats_txt = " ".join(feats)

        def feat_ou_kw(campo_bool, *kws):
            if campo_bool and det.get(campo_bool):
                return 1
            return 1 if any(k in feats_txt for k in kws) else 0

        mobilado = self.FURNISHED_MAP.get((det.get("furnished") or "").lower(), 0)
        if not mobilado and any(k in feats_txt for k in ["mobilad", "mobiliado", "furnished"]):
            mobilado = 1

        parking = det.get("parking_spaces") or 0
        try:
            parking = int(parking)
        except (TypeError, ValueError):
            parking = 0

        return {
            "mobilado":           mobilado,
            "garagem":            feat_ou_kw(None, "garagem", "estacionamento", "parking") or (1 if parking > 0 else 0),
            "piscina":            feat_ou_kw("has_pool",     "piscina", "pool"),
            "gerador":            feat_ou_kw(None,           "gerador", "generator"),
            "reservatorio":       feat_ou_kw(None,           "reservatório", "reservatorio", "tanque"),
            "seguranca":          feat_ou_kw("has_security", "segurança", "seguranca", "vigilância", "portaria"),
            "churrasqueira":      feat_ou_kw(None,           "churrasqueira", "barbecue", "bbq"),
            "ar_condicionado":    feat_ou_kw(None,           "ar condicionado", "ar-condicionado", "a/c"),
            "internet":           feat_ou_kw(None,           "internet", "fibra", "wifi", "wi-fi"),
            "condominio_fechado": feat_ou_kw(None,           "condomínio fechado", "condominio fechado", "gated", "condomínio privado"),
        }

    def _normalizar(self, lista_item: dict, det: Optional[dict]) -> dict:
        d = det if det else lista_item
        prop_id   = lista_item.get("id") or d.get("id")
        agent_id  = d.get("agent_id") or lista_item.get("agency_id", "")
        cover     = lista_item.get("cover_file_path") or d.get("cover_file_path") or ""

        p_orig, moeda, p_akz, p_usd = self._preco_e_moeda(d)
        area_m2 = d.get("area_sqm")
        lot_m2  = d.get("lot_size_sqm")
        try:
            area_float = float(area_m2) if area_m2 not in (None, "") else None
        except (TypeError, ValueError):
            area_float = None
        preco_m2 = round(p_akz / area_float, 2) if p_akz and area_float and area_float > 0 else None

        negocio = self._tipo_negocio(d)
        tipo_im = self._tipo_imovel(d)
        bairro, lat, lon = self._bairro(d)
        feats = self._features(d)
        titulo = limpar_texto(d.get("title") or lista_item.get("title") or "")
        tipologia = norm_tipologia(titulo)
        quartos = d.get("bedrooms")
        if quartos is None and tipologia:
            m = re.search(r"\d+", tipologia)
            quartos = int(m.group()) if m else None
        if not tipologia and quartos is not None:
            prefixo = "T" if tipo_im == "apartamento" else "V"
            tipologia = f"{prefixo}{quartos}"

        descricao = limpar_texto(d.get("description") or "", limite=5000)
        publicado = (d.get("published_at") or d.get("created_at") or "")[:10]
        img_url   = self._url_imagem(agent_id, prop_id, cover)
        verif = d.get("verification_status", "")
        estado = "Verificado" if verif == "verified" else None

        return {
            "fonte":              "MIA",
            "id_fonte":           prop_id,
            "url":                self._url_imovel(prop_id),
            "id_angocasa":        None,
            "tipo_negocio":       negocio,
            "pais":               "Angola",
            "provincia":          d.get("province") or d.get("city") or "Luanda",
            "bairro":             bairro,
            "lat":                lat,
            "lon":                lon,
            "headline":           titulo,
            "titulo_anuncio":     titulo,
            "tipo_imovel":        tipo_im,
            "tipologia":          tipologia,
            "estado":             estado,
            "area_m2":            area_float,
            "area_util_m2":       area_float,
            "area_bruta_m2":      None,
            "area_lote_m2":       lot_m2,
            "preco_original":     p_orig,
            "moeda":              moeda,
            "preco_akz":          p_akz,
            "preco_usd":          p_usd,
            "preco_m2_akz":       preco_m2,
            "quartos":            quartos,
            "wc":                 d.get("bathrooms"),
            **feats,
            "publicado_em":       publicado,
            "descricao":          descricao,
            "anunciante":         None,
            "telefone_whatsapp":  None,
            "url_imagem":         img_url,
            "raw_json": json.dumps({
                "fonte": "MIA",
                "prop_id": prop_id,
                "share_slug": d.get("share_slug"),
                "listing_type": d.get("listing_type"),
                "property_type": d.get("property_type"),
                "verification": verif,
                "quality_score": d.get("quality_score"),
                "has_elevator": d.get("has_elevator"),
                "has_garden": d.get("has_garden"),
                "floor_number": d.get("floor_number"),
                "total_floors": d.get("total_floors"),
                "parking_spaces": d.get("parking_spaces"),
                "year_built": d.get("year_built"),
                "price_negotiable": d.get("price_negotiable"),
                "condominium_fee": d.get("condominium_fee"),
            }, ensure_ascii=False),
        }

    def recolher_negocio(self, listing_type: str, negocio_nome: str, forcar: bool = False) -> List[dict]:
        resultados: List[dict] = []
        ids_vistos: set = set()
        pagina = self.db.obter_checkpoint("MIA", negocio_nome)
        total_paginas = max(pagina, 1)
        paginas_vazias = 0
        if pagina > 1:
            log.info(f"    [MIA] Retomando do checkpoint: pág {pagina}")

        log.info(f"  ▶ MIA | {negocio_nome}")
        while pagina <= min(CFG["max_paginas"], max(total_paginas, 1)):
            dados_rpc = self._chamar_rpc(listing_type, pagina)
            if not dados_rpc:
                paginas_vazias += 1
                log.warning(f"    [MIA] Sem resposta RPC na pág {pagina}.")
                if paginas_vazias >= 2:
                    break
                pagina += 1
                continue

            lista_pag = dados_rpc.get("data") or []
            try:
                total_count = int(dados_rpc.get("total_count") or 0)
            except (TypeError, ValueError):
                total_count = 0
            if pagina == 1 and total_count:
                total_paginas = math.ceil(total_count / self.PER_PAGE)
                log.info(f"    [MIA] {negocio_nome} — {total_count} anúncios ({total_paginas} páginas de {self.PER_PAGE})")

            if not lista_pag:
                paginas_vazias += 1
                log.info(f"    [MIA] Pág {pagina}: sem resultados.")
                if paginas_vazias >= 2:
                    break
                pagina += 1
                continue

            paginas_vazias = 0
            extraidos = 0
            for item in lista_pag:
                prop_id = item.get("id")
                if not prop_id or prop_id in ids_vistos:
                    continue
                ids_vistos.add(prop_id)
                url = self._url_imovel(prop_id)
                if not forcar and self.db.anuncio_processado_recente(url):
                    log.debug(f"    [MIA] Ignorado (recente): {url}")
                    continue
                self._delay()
                try:
                    det = self._chamar_detalhe(prop_id)
                except Exception as e:
                    log.debug(f"    [MIA] Erro detalhe {prop_id}: {e}")
                    self.db.reg_erro(url, "mia_detalhe", str(e))
                    det = None
                try:
                    norm = self._normalizar(item, det)
                    resultados.append(norm)
                    extraidos += 1
                    if CFG.get("max_anuncios_por_fonte") and len(resultados) >= int(CFG["max_anuncios_por_fonte"]):
                        break
                except Exception as e:
                    log.debug(f"    [MIA] Erro normalização {prop_id}: {e}")
                    self.db.reg_erro(url, "mia_normalizacao", str(e))

            log.info(f"    MIA pág {pagina}/{total_paginas}: {extraidos} extraídos (acum.: {len(resultados)})")
            self.db.salvar_checkpoint("MIA", negocio_nome, pagina + 1)
            if pagina >= total_paginas:
                self.db.limpar_checkpoint("MIA", negocio_nome)
                break
            pagina += 1
            self._delay()

        return resultados

    def recolher(self, forcar: bool = False) -> Tuple[List[dict], bool]:
        todos: List[dict] = []
        for listing_type, negocio_nome in [("sale", "compra"), ("rent", "arrendamento")]:
            try:
                parcial = self.recolher_negocio(listing_type, negocio_nome, forcar=forcar)
                todos.extend(parcial)
            except Exception as e:
                log.error(f"[MIA] Erro coleta {negocio_nome}: {e}")
        log.info(f"  ✔ MIA: {len(todos)} anúncios total")
        return todos, True

# ─────────────────────────────────────────────────────────────────
# SCRAPER DEDICADO — ANGOIMÓVEIS (angoimoveis.com)
# ─────────────────────────────────────────────────────────────────
class ScraperAngoImoveis:
    """
    Coletor dedicado para AngoImóveis.

    Estratégia:
      - listagem estática: /category/venda?page=N e /category/alugar?page=N;
      - detalhe HTML acessível via requests + BeautifulSoup;
      - extração estruturada por .price, .detail-line, .feature__item.js-areas,
        breadcrumb, telefone e imagem principal.
    """

    BASE = "https://angoimoveis.com"
    NEGOCIOS = {
        "compra": "/category/venda",
        "arrendamento": "/category/alugar",
    }
    TIPO_MAP = {
        "vivenda": "vivenda",
        "moradia": "vivenda",
        "apartamento": "apartamento",
        "terreno": "terreno",
        "escritório": "escritorio",
        "escritorio": "escritorio",
        "loja": "loja",
        "armazém": "armazem",
        "armazem": "armazem",
        "estabelecimento": "loja",
        "hotel": "hotel",
        "quinta": "quinta",
    }

    def __init__(self, db):
        self.db = db
        self.ses = self._sessao()

    def _sessao(self):
        s = requests.Session()
        s.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://angoimoveis.com/",
        })
        return s

    def _delay(self):
        time.sleep(random.uniform(CFG["delay_min"], CFG["delay_max"]))

    def _get(self, url: str) -> Optional[BeautifulSoup]:
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.get(url, timeout=CFG["timeout"])
                r.raise_for_status()
                if CFG["salvar_html"]:
                    slug = re.sub(r"[^\w]", "_", url)[-80:]
                    d = CFG["html_debug"]
                    d.mkdir(parents=True, exist_ok=True)
                    (d / f"angoimoveis_{slug}.html").write_text(r.text, encoding="utf-8")
                return BeautifulSoup(r.text, "lxml")
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 404:
                    return None
                log.warning(f"[AngoImóveis] HTTP {e} | tentativa {tentativa}")
            except requests.RequestException as e:
                log.warning(f"[AngoImóveis] Rede {e} | tentativa {tentativa}")
            if tentativa < CFG["max_retries"]:
                time.sleep(5 * tentativa)
        log.error(f"[AngoImóveis] Falha definitiva: {url}")
        return None

    def _url_listagem(self, path: str, pagina: int) -> str:
        base = f"{self.BASE}{path}"
        return f"{base}?page={pagina}" if pagina > 1 else base

    def _extrair_links_listagem(self, soup: BeautifulSoup) -> List[str]:
        """Extrai URLs de detalhes no padrão /slug-ID, ignorando navegação."""
        urls = []
        padrao = re.compile(r"/[a-z0-9][-a-z0-9.]+-\d+/?$", re.I)
        excluir = (
            "/search", "/login", "/register", "/contact", "/page/", "/category/",
            "/location/", "/sitemap", "/posts/", "/password/", "/blog", "/tag/",
        )
        for a in soup.select("a[href]"):
            href = (a.get("href") or "").strip()
            if not href or href.startswith(("#", "javascript:", "mailto:", "tel:")):
                continue
            if any(ex in href for ex in excluir):
                continue
            u = url_absoluta(self.BASE, href)
            if not u or dominio_base(u) != "angoimoveis.com":
                continue
            if padrao.search(urlparse(u).path):
                urls.append(u)
        seen = set()
        return [u for u in urls if not (u in seen or seen.add(u))]

    def _total_paginas(self, soup: BeautifulSoup) -> int:
        nums = []
        for a in soup.select("a[href*='page=']"):
            m = re.search(r"page=(\d+)", a.get("href", ""))
            if m:
                try:
                    nums.append(int(m.group(1)))
                except ValueError:
                    pass
        return max(nums) if nums else 1

    def extrair_detalhe(self, url: str, negocio_fallback: str) -> Optional[dict]:
        soup = self._get(url)
        if not soup:
            return None
        self._delay()

        # ID do anúncio: sufixo numérico da URL.
        id_match = re.search(r"-(\d+)/?$", urlparse(url).path.rstrip("/"))
        id_fonte = id_match.group(1) if id_match else hashlib.md5(url.encode()).hexdigest()[:12]

        # Título / headline.
        h1 = soup.select_one("h1")
        og_title = soup.select_one("meta[property='og:title']")
        titulo = limpar_texto(h1.get_text(" ", strip=True)) if h1 else ""
        if not titulo and og_title:
            titulo = limpar_texto(og_title.get("content", ""))
        if not titulo:
            titulo = urlparse(url).path.strip("/").split("/")[-1].replace("-", " ").title()

        # Negócio e tipo de imóvel via breadcrumb/categorias.
        negocio = negocio_fallback
        tipo_imovel = "outro"
        for a in soup.select("nav a[href*='/category/'], a[href*='/category/']"):
            href = (a.get("href", "") or "").lower()
            txt_link = a.get_text(" ", strip=True).lower()
            if "alugar" in href or "aluguel" in txt_link:
                negocio = "arrendamento"
            elif "venda" in href or "venda" in txt_link:
                negocio = "compra"
            for k, v in self.TIPO_MAP.items():
                if k in href or k in txt_link:
                    tipo_imovel = v
                    break

        # Preço.
        price_el = soup.select_one(".price, [class*='price']")
        preco_raw = limpar_texto(price_el.get_text(" ", strip=True)) if price_el else ""
        if not preco_raw:
            # fallback conservador: primeira ocorrência monetária no texto.
            full_short = soup.get_text(" ", strip=True)[:3000]
            m_preco = re.search(r"(?:KZ|AOA|USD|US\$|\$)?\s*[\d\s.,]{4,}\s*(?:KZ|AOA|USD|US\$|\$)?", full_short, re.I)
            preco_raw = limpar_texto(m_preco.group(0)) if m_preco else ""
        moeda = detectar_moeda(preco_raw, "AKZ")
        preco_num = norm_num(preco_raw)
        if moeda == "USD":
            preco_usd = preco_num
            preco_akz = usd2akz(preco_usd)
        else:
            moeda = "AKZ"
            preco_akz = preco_num
            preco_usd = akz2usd(preco_akz)
        preco_original = preco_num

        # Features: quartos, salas, WC.
        quartos_card = salas_card = wc_card = None
        for feat in soup.select(".feature__item.js-areas, .feature__item, [class*='feature']"):
            ft = feat.get_text(" ", strip=True).lower()
            m = re.search(r"(\d+)", ft)
            n = int(m.group(1)) if m else None
            if n is None:
                continue
            if "quarto" in ft:
                quartos_card = n
            elif "sala" in ft:
                salas_card = n
            elif "banheiro" in ft or "wc" in ft or "casa de banho" in ft:
                wc_card = n

        # Campos estruturados: .detail-line e variações.
        campos = {}
        for dl in soup.select(".detail-line, [class*='detail-line'], [class*='property-detail']"):
            texto = dl.get_text("\n", strip=True)
            partes = [p.strip() for p in texto.split("\n") if p.strip()]
            if len(partes) >= 2:
                chave = partes[0].rstrip(":").strip().lower()
                valor = " ".join(partes[1:]).strip()
                if chave and valor:
                    campos[chave] = valor
            elif len(partes) == 1 and ":" in partes[0]:
                k, _, v = partes[0].partition(":")
                campos[k.strip().lower()] = v.strip()

        def gc(*chaves):
            for c in chaves:
                c_low = c.lower()
                for k, v in campos.items():
                    if c_low in k:
                        return v
            return ""

        quartos = int(norm_num(gc("quartos")) or quartos_card or 0) or None
        wc = int(norm_num(gc("banheiro", "banheiros", "wc", "casa de banho")) or wc_card or 0) or None
        bairro_det = gc("localização", "localizacao", "bairro", "zona").strip()
        tipologia = gc("tipologia").strip() or norm_tipologia(titulo)
        estado = gc("condição", "condicao", "estado").strip() or None
        negociavel = "sim" in gc("negociável", "negociavel").lower()

        if tipologia and tipologia.isdigit():
            prefixo = "T" if tipo_imovel == "apartamento" else "V"
            tipologia = f"{prefixo}{tipologia}"
        elif tipologia:
            tipologia = norm_tipologia(tipologia) or tipologia

        # Localização.
        loc_el = soup.select_one(".card-content.mapa .card-body, .card-content.mapa, [class*='mapa']")
        loc_txt = limpar_texto(loc_el.get_text(" ", strip=True)) if loc_el else ""
        if bairro_det:
            bairro = bairro_det.split(",")[0].strip()
        elif loc_txt:
            bairro = loc_txt.split(",")[0].strip()
        else:
            bairro = ""
        provincia = loc_txt.split(",")[-1].strip() if "," in loc_txt else "Luanda"

        bairro_norm, lat, lon = detectar_bairro(f"{bairro} {titulo} {loc_txt}")
        if bairro_norm:
            bairro = bairro_norm

        # Área.
        area_m2 = norm_area(gc("área", "area", "m²", "tamanho"))
        if not area_m2:
            area_m2 = norm_area(titulo)
        if not area_m2:
            area_m2 = norm_area(soup.get_text(" ", strip=True)[:5000])
        preco_m2 = round(preco_akz / area_m2, 2) if preco_akz and area_m2 and area_m2 > 0 else None

        # Descrição.
        desc_el = soup.select_one(
            ".items-detail-text, .from-wysiwyg, [class*='description'], [class*='desc-text'], .description"
        )
        descricao = limpar_texto(desc_el.get_text(" ", strip=True) if desc_el else "", limite=5000)
        if not descricao:
            # fallback: guarda um excerto útil, evitando scripts/estilos.
            for tag in soup(["script", "style", "noscript"]):
                tag.decompose()
            descricao = limpar_texto(soup.get_text(" ", strip=True), limite=5000)

        # Anunciante.
        anunc_el = soup.select_one(
            ".card-panel .name, .seller-name, [class*='owner-name'], [class*='agent-name'], [class*='seller']"
        )
        anunciante_nome = limpar_texto(anunc_el.get_text(" ", strip=True)) if anunc_el else gc("anunciante")

        # Telefone / WhatsApp.
        tel_links = soup.select("a[href^='tel:'], a[href*='wa.me'], a[href*='whatsapp']")
        telefones = []
        for tl in tel_links:
            href = tl.get("href", "")
            num = re.sub(r"[^\d+]", "", href.replace("tel:", "").replace("https://wa.me/", ""))
            if num and num not in telefones:
                telefones.append(num)
        telefone = "; ".join(telefones[:3]) if telefones else extrair_telefone(soup.get_text(" ", strip=True))

        # Imagem principal.
        og_img = soup.select_one("meta[property='og:image']")
        img_url = og_img.get("content", "").strip() if og_img else None
        if not img_url:
            img_el = soup.select_one("img[src*='storage'], img[src*='upload'], img[src*='angoimoveis']")
            img_url = img_el.get("src", "").strip() if img_el else None
        if img_url:
            img_url = url_absoluta(self.BASE, img_url)

        full_text = limpar_texto(soup.get_text(" ", strip=True), limite=15000).lower()
        if tipo_imovel == "outro":
            tipo_imovel = detectar_tipo_imovel(titulo + " " + full_text, url)

        sinais = [preco_original, area_m2, bairro, tipologia, telefone]
        if sum(1 for x in sinais if x) < 2 and len(titulo) < 8:
            log.debug(f"[AngoImóveis] Detalhe sem dados suficientes: {url}")
            return None

        return {
            "fonte": "AngoImóveis",
            "id_fonte": id_fonte,
            "url": url,
            "id_angocasa": None,
            "tipo_negocio": negocio,
            "pais": "Angola",
            "provincia": provincia or "Luanda",
            "bairro": bairro or "Não informado",
            "lat": lat,
            "lon": lon,
            "headline": titulo,
            "titulo_anuncio": titulo,
            "tipo_imovel": tipo_imovel,
            "tipologia": tipologia,
            "estado": estado,
            "area_m2": area_m2,
            "area_util_m2": area_m2,
            "area_bruta_m2": None,
            "area_lote_m2": None,
            "preco_original": preco_original,
            "moeda": moeda,
            "preco_akz": preco_akz,
            "preco_usd": preco_usd,
            "preco_m2_akz": preco_m2,
            "quartos": quartos,
            "wc": wc,
            "mobilado": tem_kw(full_text, ["mobilado", "mobiliado", "mobilada", "furnished"]),
            "garagem": tem_kw(full_text, ["garagem", "estacionamento", "parking"]),
            "piscina": tem_kw(full_text, ["piscina", "pool"]),
            "gerador": tem_kw(full_text, ["gerador", "generator"]),
            "reservatorio": tem_kw(full_text, ["reservatório", "reservatorio", "tanque de água", "tanque de agua"]),
            "seguranca": tem_kw(full_text, ["segurança 24", "seguranca 24", "vigilância", "vigilancia", "segurança permanente"]),
            "churrasqueira": tem_kw(full_text, ["churrasqueira", "barbecue", "bbq"]),
            "ar_condicionado": tem_kw(full_text, ["ar condicionado", "ar-condicionado", "a/c", "ac "]),
            "internet": tem_kw(full_text, ["internet", "fibra", "wifi", "wi-fi"]),
            "condominio_fechado": tem_kw(full_text, [
                "condomínio fechado", "condominio fechado", "condomínio privado", "condominio privado",
                "gated", "portaria", "segurança 24",
            ]),
            "publicado_em": None,
            "descricao": descricao,
            "anunciante": anunciante_nome or None,
            "telefone_whatsapp": telefone,
            "url_imagem": img_url,
            "raw_json": json.dumps({
                "fonte": "AngoImóveis",
                "url": url,
                "id_fonte": id_fonte,
                "preco_raw": preco_raw,
                "negociavel": negociavel,
                "salas": salas_card,
            }, ensure_ascii=False),
        }

    def recolher_negocio(self, negocio: str, path: str, forcar=False) -> List[dict]:
        resultados = []
        urls_vistas = set()
        paginas_sem = 0
        total_pag = 1
        log.info(f"  ▶ AngoImóveis | {negocio}")

        pagina = self.db.obter_checkpoint("AngoImóveis", negocio)
        if pagina > 1:
            log.info(f"    [AngoImóveis] Retomando do checkpoint: pág {pagina}")
        while pagina <= CFG["max_paginas"]:
            url_pag = self._url_listagem(path, pagina)
            log.debug(f"    Pág {pagina}: {url_pag}")
            soup = self._get(url_pag)
            if not soup:
                log.warning(f"    [AngoImóveis] Sem resposta na pág {pagina} — parar.")
                break

            if pagina == 1:
                total_pag = self._total_paginas(soup)
                log.info(f"    [AngoImóveis] {negocio} — {total_pag} páginas detectadas")

            links = self._extrair_links_listagem(soup)
            if not links:
                paginas_sem += 1
                log.info(f"    [AngoImóveis] Pág {pagina}: sem links úteis.")
                if paginas_sem >= 2:
                    break
                pagina += 1
                continue
            paginas_sem = 0

            extraidos = 0
            for u in links:
                if u in urls_vistas:
                    continue
                urls_vistas.add(u)
                if not forcar and self.db.anuncio_processado_recente(u):
                    log.debug(f"    Ignorado (recente): {u}")
                    continue
                try:
                    self._delay()
                    dados = self.extrair_detalhe(u, negocio)
                    if dados:
                        resultados.append(dados)
                        extraidos += 1
                except Exception as e:
                    log.debug(f"    Falha detalhe AngoImóveis: {u} | {e}")
                    self.db.reg_erro(u, "angoimoveis_detalhe", str(e))

            log.info(
                f"    AngoImóveis pág {pagina}/{total_pag}: "
                f"{extraidos} extraídos (acum.: {len(resultados)})"
            )
            self.db.salvar_checkpoint("AngoImóveis", negocio, pagina + 1)
            if pagina >= total_pag:
                self.db.limpar_checkpoint("AngoImóveis", negocio)
                break
            pagina += 1
            self._delay()

        return resultados

    def recolher(self, forcar=False) -> Tuple[List[dict], bool]:
        todos = []
        for negocio, path in self.NEGOCIOS.items():
            try:
                parcial = self.recolher_negocio(negocio, path, forcar=forcar)
                todos.extend(parcial)
            except Exception as e:
                log.error(f"[AngoImóveis] Erro coleta {negocio}: {e}")
        log.info(f"  ✔ AngoImóveis: {len(todos)} anúncios total")
        return todos, True

# ─────────────────────────────────────────────────────────────────
# SCRAPER
# ─────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────
# SCRAPER DEDICADO — ZENKI REAL ESTATE (zenkirealestate.com)
# ─────────────────────────────────────────────────────────────────
class ScraperZenki:
    """Coletor dedicado para Zenki Real Estate: WordPress + AJAX."""

    BASE = "https://zenkirealestate.com"
    AJAX_URL = "https://zenkirealestate.com/wp-admin/admin-ajax.php"
    SEARCH_URL = "https://zenkirealestate.com/"
    ITEMS_PER_PAGE = 6
    ACTION = "property"

    NEGOCIO_MAP = {
        "VENDA": "compra",
        "VENDA POR LEILÃO": "compra",
        "ARRENDAMENTO": "arrendamento",
        "VENDA/ARREND.": "compra",
        "VENDA/ARREND": "compra",
    }
    TIPO_MAP = {
        "residencial": "apartamento", "apartamentos": "apartamento", "apartamento": "apartamento",
        "escritórios": "escritorio", "escritorios": "escritorio", "escritório": "escritorio",
        "industrial": "armazem", "armazém": "armazem", "armazem": "armazem",
        "terrenos": "terreno", "terreno": "terreno", "misto": "outro",
        "comércio": "loja", "comercio": "loja", "loja": "loja", "retail": "loja",
    }

    def __init__(self, db):
        self.db = db
        self.ses = self._sessao()

    def _sessao(self):
        s = requests.Session()
        s.headers.update({
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": f"{self.BASE}/",
        })
        return s

    def _delay(self):
        time.sleep(random.uniform(CFG["delay_min"], CFG["delay_max"]))

    def _get(self, url: str) -> Optional[BeautifulSoup]:
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.get(url, timeout=CFG["timeout"])
                r.raise_for_status()
                if CFG["salvar_html"]:
                    slug = re.sub(r"[^\w]", "_", url)[-80:]
                    CFG["html_debug"].mkdir(parents=True, exist_ok=True)
                    (CFG["html_debug"] / f"zenki_{slug}.html").write_text(r.text, encoding="utf-8")
                return BeautifulSoup(r.text, "lxml")
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 404:
                    return None
                log.warning(f"[Zenki] HTTP {e} | tentativa {tentativa}")
            except requests.RequestException as e:
                log.warning(f"[Zenki] Rede {e} | tentativa {tentativa}")
            if tentativa < CFG["max_retries"]:
                time.sleep(5 * tentativa)
        log.error(f"[Zenki] Falha definitiva: {url}")
        return None

    def _post_ajax(self, ids_vistos: List[str], pagina: int) -> Optional[BeautifulSoup]:
        data = [
            ("action", self.ACTION), ("page", str(pagina)),
            ("data_per_page", str(self.ITEMS_PER_PAGE)), ("zone", ""),
            ("typology_range", ""), ("type", ""), ("area_min", ""), ("area_max", ""),
        ]
        data.extend(("ids[]", str(pid)) for pid in ids_vistos if pid)
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.post(self.AJAX_URL, data=data, timeout=CFG["timeout"])
                r.raise_for_status()
                payload = r.json()
                if not payload.get("success"):
                    log.warning(f"[Zenki] AJAX pág {pagina}: success=False")
                    return None
                html_parcial = payload.get("data", "") or ""
                if CFG["salvar_html"]:
                    CFG["data_raw"].mkdir(parents=True, exist_ok=True)
                    (CFG["data_raw"] / f"zenki_ajax_p{pagina}_{datetime.now():%Y%m%d_%H%M%S}.html").write_text(html_parcial, encoding="utf-8")
                return BeautifulSoup(html_parcial, "lxml")
            except requests.exceptions.HTTPError as e:
                log.warning(f"[Zenki] HTTP AJAX {e} | tentativa {tentativa}")
            except requests.RequestException as e:
                log.warning(f"[Zenki] Rede AJAX {e} | tentativa {tentativa}")
            except ValueError as e:
                log.warning(f"[Zenki] JSON inválido AJAX | tentativa {tentativa}: {e}")
            if tentativa < CFG["max_retries"]:
                time.sleep(5 * tentativa)
        return None

    def _extrair_total(self, soup: BeautifulSoup) -> int:
        req = soup.select_one(".request")
        if req:
            try:
                return int(req.get("data-count", 0) or 0)
            except (TypeError, ValueError):
                return 0
        return 0

    def _extrair_items_soup(self, soup: BeautifulSoup) -> List[dict]:
        items = []
        for el in soup.select(".listing__item[data-id]"):
            post_id = el.get("data-id", "")
            url = ""
            for a in el.select("a[href]"):
                href = a.get("href") or ""
                if "/properties/" in href:
                    url = url_absoluta(self.BASE, href)
                    break
            if not url:
                continue
            ribon_el = el.select_one(".properties__ribon")
            city_el = el.select_one(".properties__address-city")
            street_el = el.select_one(".properties__address-street")
            area_card = tipologia_card = tipo_card = ""
            for ov in el.select(".properties__offer-value"):
                txt = limpar_texto(ov.get_text(" ", strip=True))
                if re.search(r"\d+\s*m\s*2", txt, re.I):
                    area_card = txt
                elif re.search(r"\bT\d", txt, re.I):
                    tipologia_card = txt
                else:
                    tipo_card = txt
            items.append({
                "url": url, "post_id": post_id,
                "negocio_raw": limpar_texto(ribon_el.get_text(" ", strip=True)).upper() if ribon_el else "",
                "cidade_card": limpar_texto(city_el.get_text(" ", strip=True)) if city_el else "",
                "zona_card": limpar_texto(street_el.get_text(" ", strip=True)) if street_el else "",
                "area_card": area_card, "tipologia_card": tipologia_card, "tipo_card": tipo_card,
            })
        return items

    def _obter_todos_ids_e_urls(self) -> List[dict]:
        search_url = f"{self.SEARCH_URL}?city=Luanda&zone=&type=&area_min=&area_max=&s=Filter&typology_range="
        soup = self._get(search_url)
        if not soup:
            log.error("[Zenki] Sem resposta na pesquisa inicial.")
            return []
        total = self._extrair_total(soup)
        todos = self._extrair_items_soup(soup)
        ids_vistos = [it["post_id"] for it in todos if it.get("post_id")]
        total_paginas = math.ceil(total / self.ITEMS_PER_PAGE) if total > 0 else CFG["max_paginas"]
        log.info(f"  [Zenki] Total anunciado: {total} | pág 1: {len(todos)} cards")
        pagina = 2
        vazias = 0
        while pagina <= min(total_paginas + 1, CFG["max_paginas"]):
            self._delay()
            soup_ajax = self._post_ajax(ids_vistos, pagina)
            if not soup_ajax:
                vazias += 1
                if vazias >= 2:
                    break
                pagina += 1
                continue
            novos = self._extrair_items_soup(soup_ajax)
            if not novos:
                vazias += 1
                if vazias >= 2:
                    break
                pagina += 1
                continue
            vazias = 0
            for it in novos:
                if it.get("post_id") and it["post_id"] not in ids_vistos:
                    todos.append(it)
                    ids_vistos.append(it["post_id"])
            log.info(f"  [Zenki] AJAX pág {pagina}: {len(novos)} cards | acum.: {len(todos)}")
            if len(novos) < self.ITEMS_PER_PAGE:
                break
            pagina += 1
        return todos

    def _parse_header_info(self, header_txt: str) -> Tuple[str, Optional[float], Optional[str]]:
        parts = header_txt.split("|", 1)
        tipo_raw = limpar_texto(parts[0]).lower()
        area_m2 = None
        tipologia = None
        if len(parts) > 1:
            resto = limpar_texto(parts[1])
            area_m2 = norm_area(resto) or norm_area(resto.replace("M2", " m2"))
            tipologia = norm_tipologia(resto)
            if not tipologia:
                m = re.search(r"\b(T\d[+\-]?\d?)\b", resto, re.I)
                if m:
                    tipologia = m.group(1).upper()
        return tipo_raw, area_m2, tipologia

    def _parse_features(self, params_txt: str) -> dict:
        t = (params_txt or "").lower()
        feats = {"quartos": None, "wc": None}
        m_wc = re.search(r"wc\s*[-–]\s*(\d+)", t)
        if m_wc:
            feats["wc"] = int(m_wc.group(1))
        m_q = re.search(r"quartos?\s*[-–]\s*(\d+)", t)
        if m_q:
            feats["quartos"] = int(m_q.group(1))
        else:
            m_s = re.search(r"suite?s?\s*[-–]\s*(\d+)", t)
            if m_s:
                feats["quartos"] = int(m_s.group(1))
        feats.update({
            "mobilado": tem_kw(t, ["mobilado", "mobiliado", "furnished"]),
            "garagem": tem_kw(t, ["garagem", "estacionamento", "parking"]),
            "piscina": tem_kw(t, ["piscina", "pool"]),
            "gerador": tem_kw(t, ["gerador", "generator", "kva"]),
            "reservatorio": tem_kw(t, ["reservatório", "reservatorio", "tanque", "eta"]),
            "seguranca": tem_kw(t, ["segurança", "seguranca", "vigilância", "portaria"]),
            "churrasqueira": tem_kw(t, ["churrasqueira", "barbecue", "bbq"]),
            "ar_condicionado": tem_kw(t, ["ar condicionado", "ar-condicionado", "a/c"]),
            "internet": tem_kw(t, ["internet", "fibra", "wifi"]),
            "condominio_fechado": tem_kw(t, ["condomínio fechado", "condominio fechado", "gated", "portaria", "condomínio privado"]),
            "ginasio": 1 if "ginásio" in t or "ginasio" in t else 0,
        })
        return feats

    def extrair_detalhe(self, url: str, card: dict) -> Optional[dict]:
        soup = self._get(url)
        if not soup:
            return None
        self._delay()
        body_class = soup.body.get("class", []) if soup.body else []
        body_class_str = " ".join(body_class) if isinstance(body_class, list) else str(body_class)
        post_id_m = re.search(r"postid-(\d+)", body_class_str)
        post_id = post_id_m.group(1) if post_id_m else card.get("post_id", "")
        h1 = soup.select_one("h1")
        h2 = soup.select_one("h2")
        titulo = limpar_texto(h1.get_text(" ", strip=True)) if h1 else ""
        loc_raw = limpar_texto(h2.get_text(" ", strip=True)) if h2 else ""
        partes_loc = [p.strip() for p in loc_raw.split(",")]
        provincia = partes_loc[0].title() if partes_loc else "Luanda"
        bairro_raw = partes_loc[1].title() if len(partes_loc) > 1 else ""
        bairro_norm, lat, lon = detectar_bairro(f"{bairro_raw} {titulo}")
        bairro = bairro_norm or bairro_raw or card.get("zona_card", "") or "Não informado"
        header_el = soup.select_one(".property__header")
        header_txt = limpar_texto(header_el.get_text(" ", strip=True)) if header_el else ""
        header_txt = re.sub(r"PARTILHAR.*$", "", header_txt, flags=re.I).strip()
        tipo_raw, area_m2, tipologia = self._parse_header_info(header_txt)
        if not tipologia:
            tipologia = norm_tipologia((card.get("tipologia_card", "") or "") + " " + titulo)
        tipo_imovel = self.TIPO_MAP.get(tipo_raw) or detectar_tipo_imovel(f"{titulo} {tipo_raw}", url)
        q_from_tip = None
        if tipologia:
            m = re.search(r"\d+", tipologia)
            q_from_tip = int(m.group()) if m else None
        ribon_el = soup.select_one(".properties__ribon")
        ribon_txt = limpar_texto(ribon_el.get_text(" ", strip=True)).upper() if ribon_el else ""
        negocio = self.NEGOCIO_MAP.get(ribon_txt) or detectar_negocio(card.get("negocio_raw", ""), url, fallback="compra")
        desc_el = soup.select_one(".property__description")
        descricao = limpar_texto(desc_el.get_text(" ", strip=True), limite=5000) if desc_el else ""
        params_el = soup.select_one(".property__params-list")
        params_txt = limpar_texto(params_el.get_text("\n", strip=True)) if params_el else ""
        feats = self._parse_features(params_txt)
        quartos = feats.pop("quartos") or q_from_tip
        wc = feats.pop("wc")
        ginasio = feats.pop("ginasio", 0)
        body_text = limpar_texto(soup.get_text(" ", strip=True), limite=12000)
        preco_orig, moeda = extrair_preco_texto(body_text)
        if moeda == "USD":
            preco_usd = preco_orig; preco_akz = usd2akz(preco_usd)
        elif moeda == "AKZ" and preco_orig:
            preco_akz = preco_orig; preco_usd = akz2usd(preco_akz)
        else:
            preco_orig = preco_akz = preco_usd = None; moeda = "AKZ"
        preco_m2 = round(preco_akz / area_m2, 2) if preco_akz and area_m2 and area_m2 > 0 else None
        agent_el = soup.select_one(".fn")
        anunciante = limpar_texto(agent_el.get_text(" ", strip=True)) if agent_el else None
        tels = [a.get("href", "").replace("tel:", "").strip() for a in soup.select("a[href^='tel:']") if a.get("href", "").replace("tel:", "").strip()]
        telefone = "; ".join(list(dict.fromkeys(tels))[:3]) if tels else extrair_telefone(body_text)
        mail_el = soup.select_one("a[href^='mailto:']")
        email_agente = mail_el.get("href", "").replace("mailto:", "").strip() if mail_el else None
        og_img = soup.select_one("meta[property='og:image']")
        url_imagem = og_img.get("content", "").strip() if og_img else None
        if not url_imagem:
            img_el = soup.select_one(".js-dapi-gallery img, img[src*='zenkirealestate']")
            url_imagem = url_absoluta(self.BASE, img_el.get("src", "")) if img_el else None
        sinais = [titulo, bairro, tipo_imovel, tipologia, area_m2, descricao]
        if sum(1 for x in sinais if x and x not in ("outro", "Não informado")) < 2:
            return None
        segmento = "Comercial" if tipo_imovel in ("escritorio", "loja") else ("Industrial" if tipo_imovel == "armazem" else ("Terreno" if tipo_imovel == "terreno" else "Residencial"))
        return {
            "fonte": "Zenki Real Estate", "id_fonte": post_id, "url": url, "id_angocasa": None,
            "tipo_negocio": negocio, "pais": "Angola", "provincia": provincia or "Luanda", "bairro": bairro,
            "lat": lat, "lon": lon, "headline": titulo, "titulo_anuncio": titulo, "tipo_imovel": tipo_imovel,
            "tipologia": tipologia, "estado": None, "area_m2": area_m2, "area_util_m2": None,
            "area_bruta_m2": area_m2, "area_lote_m2": None, "preco_original": preco_orig, "moeda": moeda,
            "preco_akz": preco_akz, "preco_usd": preco_usd, "preco_m2_akz": preco_m2, "quartos": quartos, "wc": wc,
            **feats, "publicado_em": None, "descricao": descricao, "anunciante": anunciante,
            "telefone_whatsapp": telefone, "url_imagem": url_imagem,
            "segmento_imobiliario": segmento,
            "email_agente": email_agente,
            "id_zenki": post_id,
            "imovel_corporativo": 1 if tipo_imovel in ("escritorio", "armazem", "loja") else 0,
            "preco_disponivel": 1 if preco_orig else 0,
            "url_brochura": None,
            "pdf_brochura": None,
            "raw_json": json.dumps({"fonte": "Zenki Real Estate", "post_id": post_id, "ribon": ribon_txt,
                                    "tipo_raw": tipo_raw, "header": header_txt, "params": params_txt,
                                    "ginasio": ginasio, "email_agente": email_agente,
                                    "segmento_imobiliario": segmento, "preco_disponivel": 1 if preco_orig else 0}, ensure_ascii=False),
        }

    def recolher(self, forcar: bool = False) -> Tuple[List[dict], bool]:
        resultados = []
        urls_vistas = set()
        log.info("  ▶ Zenki Real Estate | Luanda")
        cards = self._obter_todos_ids_e_urls()
        if not cards:
            log.warning("  [Zenki] Sem cards recolhidos.")
            return [], False
        for card in cards:
            u = card.get("url", "")
            if not u or u in urls_vistas:
                continue
            urls_vistas.add(u)
            if not forcar and self.db.anuncio_processado_recente(u):
                continue
            try:
                self._delay()
                dados = self.extrair_detalhe(u, card)
                if dados:
                    resultados.append(dados)
                    if CFG.get("max_anuncios_por_fonte") and len(resultados) >= int(CFG["max_anuncios_por_fonte"]):
                        break
            except Exception as e:
                log.debug(f"    [Zenki] Falha detalhe: {u} | {e}")
                self.db.reg_erro(u, "zenki_detalhe", str(e))
        log.info(f"  ✔ Zenki Real Estate: {len(resultados)} anúncios extraídos de {len(cards)} cards")
        return resultados, True


# ─────────────────────────────────────────────────────────────────
# SCRAPER DEDICADO — REE-GERA (ree-gera.com)
# ─────────────────────────────────────────────────────────────────
class ScraperReeGera:
    """Coletor dedicado para REE-GERA Imobiliária (WordPress/wpResidence)."""

    BASE = "https://ree-gera.com"
    NEGOCIOS = {"compra": "/a%C3%A7ao/vendas/", "arrendamento": "/a%C3%A7ao/aluguel/"}
    FT2_TO_M2 = 0.092903
    TIPO_MAP = {
        "apartamentos": "apartamento", "apartamento": "apartamento", "casas": "vivenda",
        "vivendas": "vivenda", "vivenda": "vivenda", "moradia": "vivenda", "moradias": "vivenda",
        "terrenos": "terreno", "terreno": "terreno", "escritórios": "escritorio",
        "escritório": "escritorio", "escritorios": "escritorio", "escritorio": "escritorio",
        "lojas": "loja", "loja": "loja", "armazém": "armazem", "armazéns": "armazem",
        "armazem": "armazem", "duplexes": "apartamento", "duplex": "apartamento",
        "condomínios": "outro", "condominio": "outro", "empreendimento": "outro",
        "hotel": "outro", "guest house": "outro", "industrial": "armazem", "colégio": "outro",
    }

    def __init__(self, db):
        self.db = db
        self.ses = self._sessao()

    def _sessao(self):
        s = requests.Session()
        s.headers.update({
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://ree-gera.com/",
        })
        return s

    def _delay(self):
        time.sleep(random.uniform(CFG["delay_min"], CFG["delay_max"]))

    def _get(self, url: str) -> Optional[BeautifulSoup]:
        for tentativa in range(1, CFG["max_retries"] + 1):
            try:
                r = self.ses.get(url, timeout=CFG["timeout"])
                r.raise_for_status()
                if CFG["salvar_html"]:
                    slug = re.sub(r"[^\w]", "_", url)[-80:]
                    CFG["html_debug"].mkdir(parents=True, exist_ok=True)
                    (CFG["html_debug"] / f"reegera_{slug}.html").write_text(r.text, encoding="utf-8")
                return BeautifulSoup(r.text, "lxml")
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 404:
                    return None
                log.warning(f"[REE-GERA] HTTP {e} | tentativa {tentativa}")
            except requests.RequestException as e:
                log.warning(f"[REE-GERA] Rede {e} | tentativa {tentativa}")
            if tentativa < CFG["max_retries"]:
                time.sleep(5 * tentativa)
        log.error(f"[REE-GERA] Falha definitiva: {url}")
        return None

    def _url_listagem(self, path: str, pagina: int) -> str:
        base = f"{self.BASE}{path}"
        return base.rstrip("/") + f"/page/{pagina}/" if pagina > 1 else base

    def _extrair_links_listagem(self, soup: BeautifulSoup) -> List[str]:
        urls, seen = [], set()
        for card in soup.select(".property_listing.property_card_default"):
            link = card.select_one("a[href*='/propriedades/']")
            if link:
                u = url_absoluta(self.BASE, link.get("href", ""))
                if u and u not in seen:
                    seen.add(u); urls.append(u)
        return urls

    def _total_paginas(self, soup: BeautifulSoup) -> int:
        nums = []
        for a in soup.select("a[href*='/page/']"):
            m = re.search(r"/page/(\d+)/", a.get("href", ""))
            if m:
                nums.append(int(m.group(1)))
        return max(nums) if nums else 1

    def _ft2_para_m2(self, txt: str) -> Optional[float]:
        if not txt:
            return None
        m = re.search(r"([\d\s,\.]+)\s*m[²2]", txt, re.I)
        if m:
            return norm_num(m.group(1))
        m = re.search(r"([\d\s,\.]+)\s*ft[²2]?", txt, re.I)
        if m:
            v = norm_num(m.group(1)); return round(v * self.FT2_TO_M2, 2) if v else None
        v = norm_num(txt)
        return round(v * self.FT2_TO_M2, 2) if v else None

    def _extrair_campos(self, soup: BeautifulSoup) -> Dict[str, str]:
        campos = {}
        for el in soup.select(".listing_detail"):
            strong = el.select_one("strong")
            if strong:
                lbl = strong.get_text(strip=True).rstrip(":").strip().lower()
                val = el.get_text(" ", strip=True).replace(strong.get_text(strip=True), "", 1).strip()
                if lbl:
                    campos[lbl] = val
        return campos

    def extrair_detalhe(self, url: str, negocio_fallback: str) -> Optional[dict]:
        soup = self._get(url)
        if not soup:
            return None
        self._delay()
        labels = [a.get_text(strip=True).lower() for a in soup.select(".single_property_labels a")]
        negocio = negocio_fallback
        tipo_imovel = "outro"
        for lbl in labels:
            if lbl in ("arrendamento", "aluguel", "arrendar"):
                negocio = "arrendamento"
            elif lbl in ("vendas", "venda", "compra"):
                negocio = "compra"
            elif lbl in self.TIPO_MAP:
                tipo_imovel = self.TIPO_MAP[lbl]
        h1 = soup.select_one("h1")
        titulo = limpar_texto(h1.get_text(" ", strip=True)) if h1 else ""
        if not titulo:
            og = soup.select_one("meta[property='og:title']")
            titulo = limpar_texto(og.get("content", "")) if og else ""
        if not titulo:
            titulo = urlparse(url).path.strip("/").split("/")[-1].replace("-", " ").title()
        if tipo_imovel == "outro":
            tipo_imovel = detectar_tipo_imovel(titulo, url)
        campos = self._extrair_campos(soup)
        def gc(*chaves):
            for c in chaves:
                for k, v in campos.items():
                    if c.lower() in k:
                        return v
            return ""
        id_fonte = gc("property id", "property_id").strip() or hashlib.md5(url.encode("utf-8")).hexdigest()[:12]
        # Preço REE-GERA: tema WordPress/wpResidence pode publicar em vários seletores ou JSON-LD.
        preco_raw = ""
        for sel in (".price_area", ".price_tag", ".listing_unit_price_wrapper", ".price", "[itemprop='price']", ".property_price"):
            el = soup.select_one(sel)
            if el:
                preco_raw = limpar_texto(el.get_text(" ", strip=True))
                if preco_raw:
                    break
        if not preco_raw:
            preco_raw = gc("price", "preço", "preco", "valor")
        if not re.search(r"\d", preco_raw or ""):
            for s_ld in soup.select("script[type='application/ld+json']"):
                try:
                    data_ld = json.loads(s_ld.string or "{}")
                except Exception:
                    continue
                nodes = data_ld if isinstance(data_ld, list) else [data_ld]
                for node in nodes:
                    if not isinstance(node, dict):
                        continue
                    off = node.get("offers") or {}
                    if isinstance(off, list):
                        off = off[0] if off else {}
                    if isinstance(off, dict):
                        p = off.get("price") or off.get("lowPrice")
                        if p:
                            preco_raw = f"{p} {off.get('priceCurrency','')}"
                            break
                if re.search(r"\d", preco_raw or ""):
                    break
        sob_consulta = bool(re.search(r"sob\s+consulta|consult|on\s+request|preço\s+a", (preco_raw or ""), re.I))
        preco_original = preco_usd = preco_akz = None
        moeda = "AKZ"
        if not sob_consulta and re.search(r"\d", preco_raw or ""):
            moeda = detectar_moeda(preco_raw, "AKZ")
            preco_num = norm_num(re.sub(r"/[Mm][êe]s|/[Mm]onth|/[Aa]no|/year", "", preco_raw, flags=re.I))
            limiar = 50 if re.search(r"arrend|rent|/m[êe]s|/month", (preco_raw or "").lower()) else 1000
            if preco_num and preco_num >= limiar:
                preco_original = preco_num
                if moeda == "USD":
                    preco_usd = preco_num; preco_akz = usd2akz(preco_num)
                else:
                    moeda = "AKZ"; preco_akz = preco_num; preco_usd = akz2usd(preco_num)
        area_raw = gc("property size", "area", "size", "m²", "m2")
        area_m2 = self._ft2_para_m2(area_raw) if area_raw else norm_area(titulo)
        preco_m2 = round(preco_akz / area_m2, 2) if preco_akz and area_m2 and area_m2 > 0 else None
        area_el = soup.select_one(".wpresidence-detail-area")
        bairro_raw = area_el.get_text(" ", strip=True).replace("Area:", "").strip() if area_el else gc("area")
        city_el = soup.select_one(".wpresidence-detail-city")
        cidade = city_el.get_text(" ", strip=True).replace("City:", "").strip() if city_el else "Luanda"
        bairro_norm, lat, lon = detectar_bairro(f"{bairro_raw} {titulo}")
        bairro = bairro_norm or bairro_raw or "Não informado"
        if not lat:
            map_el = soup.select_one("[data-lat], [data-lng]")
            if map_el:
                try:
                    lat = float(map_el.get("data-lat", 0) or 0) or None
                    lon = float(map_el.get("data-lng", 0) or 0) or None
                except (ValueError, TypeError):
                    pass
        quartos = int(norm_num(gc("rooms", "quartos", "bedrooms")) or 0) or None
        wc = int(norm_num(gc("bathrooms", "casas de banho", "wc", "banheiro")) or 0) or None
        tipologia = norm_tipologia(titulo)
        if not tipologia and quartos:
            tipologia = f"{'T' if tipo_imovel == 'apartamento' else 'V'}{quartos}"
        estado = gc("estado", "condição", "condicao", "estado de conservação").strip() or None
        publicado_em = gc("updated on", "publicado", "data").strip() or None
        feature_items = [el.get_text(" ", strip=True).lower() for el in soup.select(".listing_detail") if not el.select_one("strong")]
        features_txt = " ".join(feature_items) + " " + soup.get_text(" ", strip=True).lower()
        flags = {
            "mobilado": tem_kw(features_txt, ["mobilado", "mobiliado", "furnished"]),
            "garagem": tem_kw(features_txt, ["garagem", "estacionamento", "parking", "estacionamento privado"]),
            "piscina": tem_kw(features_txt, ["piscina", "área social - piscina", "pool"]),
            "gerador": tem_kw(features_txt, ["gerador", "generator"]),
            "reservatorio": tem_kw(features_txt, ["reservatório", "reservatorio", "tanque de água", "tanque de agua"]),
            "seguranca": tem_kw(features_txt, ["porteiro", "segurança 24", "seguranca 24", "vigilância", "security", "portaria"]),
            "churrasqueira": tem_kw(features_txt, ["churrasqueira", "barbecue", "bbq"]),
            "ar_condicionado": tem_kw(features_txt, ["ar condicionado", "ar-condicionado", "air ac", "a/c"]),
            "internet": tem_kw(features_txt, ["internet", "fibra", "wifi", "wi-fi"]),
            "condominio_fechado": tem_kw(features_txt, ["condomínio fechado", "condominio fechado", "condomínio privado", "condominio privado", "gated", "portaria", "porteiro"]),
        }
        desc_el = soup.select_one(".listing_details.the_grid_view, .listing_details, .property-description")
        descricao = limpar_texto(desc_el.get_text(" ", strip=True) if desc_el else "", limite=5000)
        if not descricao:
            descricao = limpar_texto(soup.get_text(" ", strip=True), limite=5000)
        agent_el = soup.select_one(".agent_detail .name, .agent-name, [class*='agent'] .name")
        anunciante = limpar_texto(agent_el.get_text(" ", strip=True)) if agent_el else None
        tels = []
        for tl in soup.select("a[href^='tel:'], a[href*='wa.me']"):
            href = tl.get("href", "")
            num = re.sub(r"[^\d+]", "", href.replace("tel:", "").replace("https://wa.me/", ""))
            if num and num not in tels:
                tels.append(num)
        telefone = "; ".join(tels[:3]) if tels else extrair_telefone(soup.get_text(" ", strip=True))
        og_img = soup.select_one("meta[property='og:image']")
        img_url = og_img.get("content", "").strip() if og_img else None
        if not img_url:
            img_el = soup.select_one(".listing-unit-img-wrapper img, .property-img img")
            img_url = url_absoluta(self.BASE, img_el.get("src", "")) if img_el else None
        sinais = [preco_original, area_m2, bairro not in (None, "Não informado"), tipologia, telefone]
        if sum(1 for x in sinais if x) < 2 and len(titulo) < 8:
            return None
        return {
            "fonte": "REE-GERA", "id_fonte": id_fonte, "url": url, "id_angocasa": None,
            "tipo_negocio": negocio, "pais": "Angola", "provincia": cidade or "Luanda", "bairro": bairro,
            "lat": lat, "lon": lon, "headline": titulo, "titulo_anuncio": titulo, "tipo_imovel": tipo_imovel,
            "tipologia": tipologia, "estado": estado, "area_m2": area_m2, "area_util_m2": area_m2,
            "area_bruta_m2": None, "area_lote_m2": None, "preco_original": preco_original, "moeda": moeda,
            "preco_akz": preco_akz, "preco_usd": preco_usd, "preco_m2_akz": preco_m2, "quartos": quartos,
            "wc": wc, **flags, "publicado_em": publicado_em, "descricao": descricao,
            "anunciante": anunciante, "telefone_whatsapp": telefone, "url_imagem": img_url,
            "preco_disponivel": 0 if sob_consulta else (1 if preco_original else 0),
            "raw_json": json.dumps({"fonte": "REE-GERA", "url": url, "id_fonte": id_fonte, "labels": labels,
                                    "preco_raw": preco_raw, "sob_consulta": sob_consulta, "area_raw": area_raw}, ensure_ascii=False),
        }

    def recolher_negocio(self, negocio: str, path: str, forcar: bool = False) -> List[dict]:
        resultados, urls_vistas = [], set()
        total_pag = 1
        paginas_sem = 0
        log.info(f"  ▶ REE-GERA | {negocio}")
        pagina = self.db.obter_checkpoint("REE-GERA", negocio)
        if pagina > 1:
            log.info(f"    [REE-GERA] Retomando do checkpoint: pág {pagina}")
        while pagina <= CFG["max_paginas"]:
            soup = self._get(self._url_listagem(path, pagina))
            if not soup:
                break
            if pagina == 1:
                total_pag = self._total_paginas(soup)
                log.info(f"    [REE-GERA] {negocio} — {total_pag} página(s) detectada(s)")
            links = self._extrair_links_listagem(soup)
            if not links:
                paginas_sem += 1
                if paginas_sem >= 2:
                    break
                pagina += 1
                continue
            paginas_sem = 0
            extraidos = 0
            for u in links:
                if u in urls_vistas:
                    continue
                urls_vistas.add(u)
                if not forcar and self.db.anuncio_processado_recente(u):
                    continue
                try:
                    self._delay()
                    dados = self.extrair_detalhe(u, negocio)
                    if dados:
                        resultados.append(dados); extraidos += 1
                        if CFG.get("max_anuncios_por_fonte") and len(resultados) >= int(CFG["max_anuncios_por_fonte"]):
                            break
                except Exception as e:
                    log.debug(f"    Falha detalhe REE-GERA: {u} | {e}")
                    self.db.reg_erro(u, "reegera_detalhe", str(e))
            log.info(f"    REE-GERA pág {pagina}/{total_pag}: {extraidos} extraídos (acum.: {len(resultados)})")
            self.db.salvar_checkpoint("REE-GERA", negocio, pagina + 1)
            if pagina >= total_pag:
                self.db.limpar_checkpoint("REE-GERA", negocio)
                break
            pagina += 1
            self._delay()
        return resultados

    def recolher(self, forcar: bool = False) -> Tuple[List[dict], bool]:
        todos = []
        for negocio, path in self.NEGOCIOS.items():
            try:
                todos.extend(self.recolher_negocio(negocio, path, forcar=forcar))
            except Exception as e:
                log.error(f"[REE-GERA] Erro coleta {negocio}: {e}")
        log.info(f"  ✔ REE-GERA: {len(todos)} anúncios total")
        return todos, True

class Scraper:
    """
    Seletores confirmados ao vivo em 05/06/2026:
      Card   : a.listedcar[href] | .details h3 | .dealName | .areaSize
               .price .value | .price .currency | .price .sec-value | .description
      Detalhe: .col.params dl dt+dd | .text (h2 + texto)
      Paginação:
        Pág 1: /anuncios/{neg}/angola-luanda{-slug}/mostrar-50/ordenar-normal/
        Pág N: /anuncios/{neg}/angola-luanda{-slug}/mostrar-50/pagina-{N}/ordenar-normal/
    """
    BASE="https://www.angocasa.com"

    def __init__(self,db):
        self.db=db
        self.ses=self._sessao()

    def _sessao(self):
        s=requests.Session()
        s.headers.update({
            "User-Agent":("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/124.0.0.0 Safari/537.36"),
            "Accept-Language":"pt-PT,pt;q=0.9,en;q=0.8",
            "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer":"https://www.angocasa.com/",
        })
        return s

    def _delay(self):
        time.sleep(random.uniform(CFG["delay_min"],CFG["delay_max"]))

    def _get(self, url):
        for t in range(1,CFG["max_retries"]+1):
            try:
                r=self.ses.get(url,timeout=CFG["timeout"])
                r.raise_for_status()
                if CFG["salvar_html"]:
                    slug=re.sub(r"[^\w]","_",url)[-80:]
                    d=CFG["html_debug"]; d.mkdir(parents=True,exist_ok=True)
                    (d/f"{slug}.html").write_text(r.text,encoding="utf-8")
                return BeautifulSoup(r.text,"lxml")
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code==404: return None
                log.warning(f"HTTP {e} | t{t}")
            except requests.RequestException as e:
                log.warning(f"Rede {e} | t{t}")
            if t<CFG["max_retries"]: time.sleep(5*t)
        log.error(f"Falha definitiva: {url}")
        return None

    def construir_url(self, neg_slug, bairro_slug, pagina=1):
        """
        neg_slug  : 'venda' ou 'arrenda'
        bairro_slug: slug do bairro ou '' para toda Luanda
        """
        base=f"{self.BASE}/anuncios/{neg_slug}/angola-luanda"
        if bairro_slug: base+=f"-{bairro_slug}"
        base+=f"/mostrar-{CFG['items_pag']}/"
        if pagina>1: base+=f"pagina-{pagina}/"
        base+="ordenar-normal/"
        return base

    def extrair_card(self, el):
        """Extrai dados básicos de um card/anúncio da listagem.

        Observação: o AngoCasa pode devolver links absolutos ou relativos.
        Esta função normaliza sempre a URL antes de gravar.
        """
        href = (el.get("href") or "").strip()

        if not href:
            return None

        # Normalizar URL do anúncio
        href = urljoin(self.BASE, href)
        if href.startswith("http://"):
            href = href.replace("http://", "https://", 1)

        # Evitar links que não sejam de anúncios/imóveis
        href_lower = href.lower()
        if "angocasa.com" not in href_lower:
            return None

        titulo = (
            el.select_one(".details h3")
            or el.select_one("h3")
            or el.select_one("h2")
            or el.select_one(".title")
            or el.select_one(".headline")
        )
        titulo_txt = titulo.get_text(" ", strip=True) if titulo else ""

        negocio_el = el.select_one(".dealName") or el.select_one(".deal-name")
        negocio_txt = negocio_el.get_text(" ", strip=True) if negocio_el else ""

        area_el = el.select_one(".areaSize") or el.select_one(".area-size")
        area_txt = area_el.get_text(" ", strip=True) if area_el else ""

        val_el = el.select_one(".price .value") or el.select_one(".price")
        val_txt = val_el.get_text(" ", strip=True) if val_el else ""

        cur_el = el.select_one(".price .currency") or el.select_one(".currency")
        cur_txt = cur_el.get_text(" ", strip=True) if cur_el else "AKZ"

        sec_el = el.select_one(".price .sec-value") or el.select_one(".sec-value")
        sec_txt = sec_el.get_text(" ", strip=True) if sec_el else ""

        desc_el = el.select_one(".description") or el.select_one("p")
        desc_txt = desc_el.get_text(" ", strip=True) if desc_el else ""

        # Se o título vier vazio, usar parte da URL como fallback mínimo
        if not titulo_txt:
            titulo_txt = href.rstrip("/").split("/")[-1].replace("-", " ").title()

        return {
            "url": href,
            "headline": titulo_txt,
            "tipo_imovel": norm_tipo_url(href),
            "tipologia": norm_tipologia(titulo_txt),
            "tipo_negocio": norm_negocio(negocio_txt),
            "area_card": area_txt,
            "val_txt": val_txt,
            "moeda_card": cur_txt.upper().replace("KZ", "AKZ").replace("KZS", "AKZ"),
            "sec_txt": sec_txt,
            "descricao_card": desc_txt,
        }

    def extrair_preco(self, val_txt, moeda, sec_txt):
        """
        Retorna (preco_original, moeda_norm, preco_akz, preco_usd)
        val_txt  : '1 500 000 AKZ'  ou  '500 000 USD'
        sec_txt  : '414 000 000 Kz' (preço secundário em Kz quando primário é USD)
        """
        preco_orig = norm_num(val_txt)
        moeda_n = moeda.upper().replace("KZ","AKZ").replace("KZS","AKZ")
        if not moeda_n or moeda_n in ("","N/A"): moeda_n="AKZ"

        if moeda_n in ("AKZ","AOA","KZ","KZS"):
            moeda_n="AKZ"
            p_akz=preco_orig
            p_usd=akz2usd(p_akz)
        elif moeda_n in ("USD","US$","$","DÓLARES","DOLARES"):
            moeda_n="USD"
            p_usd=preco_orig
            # tentar pegar Kz do sec_txt
            p_akz_sec=norm_num(re.sub(r"[Kk][Zz]","",sec_txt)) if sec_txt else None
            p_akz = p_akz_sec if p_akz_sec else usd2akz(p_usd)
        else:
            moeda_n="AKZ"; p_akz=preco_orig; p_usd=akz2usd(p_akz)

        return preco_orig, moeda_n, p_akz, p_usd

    def extrair_detalhe(self, url, dados_card):
        """
        Visita página de detalhe e enriquece os dados do card.
        Melhorias v2.8:
          - área flexível mesmo quando o site informa número sem m²;
          - quartos/WC vindos de parâmetros estruturados;
          - telefone, anunciante e imagem principal;
          - bairro com fallback por texto completo e slug.
        """
        soup=self._get(url)
        if not soup: return None
        self._delay()

        # ── Título/headline completo ─────────────────────────
        titulo_completo = extrair_headline_completa(soup, dados_card.get("headline", ""))

        # ── Campos estruturados: dl dt+dd ──────────────────────
        campos={}
        for dl in soup.select(".col.params dl, dl"):
            dts=dl.find_all("dt")
            dds=dl.find_all("dd")
            for dt,dd in zip(dts,dds):
                k=limpar_texto(dt.get_text(" ", strip=True)).rstrip(":").strip().lower()
                v=limpar_texto(dd.get_text(" ", strip=True))
                if k and v:
                    campos[k]=v

        # ── Função auxiliar para procurar chave parcial ────────
        def gc(partes):
            for p in partes:
                p = p.lower()
                for k,v in campos.items():
                    if p in k:
                        return v
            return ""

        # ── Texto completo útil ────────────────────────────────
        body_text = limpar_texto(soup.get_text(" ", strip=True), limite=20000)
        body_low  = body_text.lower()

        # ── Extrair ID Angocasa ────────────────────────────────
        id_raw=gc(["id angocasa","id_angocasa","id"])
        id_an=id_raw.replace("#","").strip() if id_raw else None
        if not id_an:
            m_id = re.search(r"/(\d{5,})/?$", urlparse(url).path)
            id_an = m_id.group(1) if m_id else None

        # ── Extrair preço (pode ser diferente do card) ─────────
        preco_raw=gc(["preço","preco","price"])
        moeda_det="AKZ"
        if "USD" in preco_raw.upper() or "$" in preco_raw: moeda_det="USD"
        elif "KZ" in preco_raw.upper() or "AKZ" in preco_raw.upper() or "AOA" in preco_raw.upper(): moeda_det="AKZ"

        if not preco_raw:
            preco_raw=dados_card.get("val_txt","")
            moeda_det=dados_card.get("moeda_card","AKZ")

        p_orig,moeda_n,p_akz,p_usd = self.extrair_preco(preco_raw, moeda_det, dados_card.get("sec_txt",""))

        # ── Áreas: aceitar '350', '350 m²', '350 m2' ───────────
        area_util  = norm_area_flex(gc(["área útil","area util","útil", "area útil"]))
        area_bruta = norm_area_flex(gc(["área bruta","area bruta","bruta"]))
        area_lote  = norm_area_flex(gc(["área lote","area lote","lote","implantação", "implantacao"]))
        area_card  = norm_area_flex(dados_card.get("area_card",""))
        area_txt_fallback = re.search(r"(?:área|area|m2|m²)\D{0,20}([\d\s.,]{2,})", body_text, flags=re.I)
        area_body = norm_area_flex(area_txt_fallback.group(1)) if area_txt_fallback else None
        area_m2 = area_util or area_bruta or area_lote or area_card or area_body

        preco_m2=round(p_akz/area_m2,2) if p_akz and area_m2 and area_m2>0 else None

        # ── Tipologia e tipo de imóvel ─────────────────────────
        tip_det = gc(["tipologia"])
        tip = norm_tipologia(tip_det) or dados_card.get("tipologia") or norm_tipologia(titulo_completo + " " + body_text[:1000])
        tipo_det = gc(["tipo de imóvel", "tipo de imovel", "tipo"]).lower().strip()
        tipo_imovel = tipo_det if tipo_det else dados_card.get("tipo_imovel","outro")
        tipo_map={"vivenda":"vivenda","moradia":"vivenda","apartamento":"apartamento","terreno":"terreno",
                  "loja":"loja","escritório":"escritorio","escritorio":"escritorio",
                  "armazém":"armazem","armazem":"armazem", "prédio":"outro", "predio":"outro"}
        tipo_imovel=tipo_map.get(tipo_imovel, detectar_tipo_imovel(titulo_completo + " " + body_text[:1000], url) if tipo_imovel in ("", "outro") else tipo_imovel)

        # ── Negócio ───────────────────────────────────────────
        neg_det = gc(["tipo de negócio","tipo de negocio","negócio","negocio"])
        negocio = norm_negocio(neg_det) or dados_card.get("tipo_negocio","") or detectar_negocio(titulo_completo, url, "")

        # ── Localização ───────────────────────────────────────
        bairro  = gc(["localidade","zona","bairro"]).strip()
        regiao  = gc(["região","regiao", "província", "provincia"]).strip() or "Luanda"
        pais_d  = gc(["país","pais"]).strip() or "Angola"
        b_det, lat, lon = detectar_bairro(f"{bairro} {titulo_completo} {body_text[:1500]} {url}")
        if b_det:
            bairro = b_det
        if not bairro:
            bairro = "Não informado"

        # ── Características ───────────────────────────────────
        q_raw    = gc(["quartos", "quarto", "dormitórios", "dormitorios"])
        wc_raw   = gc(["casas de banho","casa de banho","wc","banheiro", "banheiros"])
        quartos  = int(norm_num(q_raw) or 0) or None
        wc_n     = int(norm_num(wc_raw) or 0) or None
        estado   = gc(["estado", "condição", "condicao"]).strip()
        mobilado_raw = gc(["mobilad","mobiliado","furnished"])
        mobilado = 0 if "não" in mobilado_raw.lower() or "nao" in mobilado_raw.lower() else (1 if mobilado_raw else 0)

        ext  = gc(["exterior"]).lower()
        intr = gc(["interior"]).lower()
        extr = gc(["extras","outros", "características", "caracteristicas"]).lower()
        tudo = " ".join([ext,intr,extr,dados_card.get("descricao_card","").lower(), body_low[:5000]])

        pub = gc(["publicado","published","data"]).strip()

        # ── Descrição ─────────────────────────────────────────
        text_el=soup.select_one(".text")
        descricao=""
        if text_el:
            h_el=text_el.find(["h2","h3"])
            titulo_an=h_el.get_text(" ", strip=True) if h_el else titulo_completo
            if h_el: h_el.decompose()
            descricao=limpar_texto(text_el.get_text(" ",strip=True), limite=5000)
        else:
            titulo_an=titulo_completo
            descricao=limpar_texto(body_text, limite=5000)

        # Quartos: ordem estruturado > tipologia > descrição
        if quartos is None and tip:
            q_tip=re.search(r"\d+",tip)
            if q_tip: quartos=int(q_tip.group())
        if quartos is None:
            q_match=re.search(r"(\d+)\s*(?:quarto|suite|su[ií]te|dormit)",tudo+" "+descricao.lower())
            quartos=int(q_match.group(1)) if q_match else None
        if wc_n is None:
            wc_match=re.search(r"(\d+)\s*(?:wc|casa(?:s)? de banho|banheiro)",tudo+" "+descricao.lower())
            wc_n=int(wc_match.group(1)) if wc_match else None

        # ── Telefone / WhatsApp ───────────────────────────────
        telefones=[]
        for a in soup.select("a[href^='tel:'], a[href*='wa.me'], a[href*='whatsapp']"):
            href=a.get("href","")
            num=re.sub(r"[^\d+]", "", href.replace("tel:", "").replace("https://wa.me/", "").replace("http://wa.me/", ""))
            if num and len(re.sub(r"\D", "", num)) >= 9 and num not in telefones:
                telefones.append(num)
        tel_text = extrair_telefone(body_text)
        if tel_text:
            for t in tel_text.split(";"):
                t=t.strip()
                if t and t not in telefones:
                    telefones.append(t)
        telefone = "; ".join(telefones[:3]) if telefones else None

        # ── Anunciante / agência ──────────────────────────────
        anunciante = None
        for sel in [".agent .name", ".agent-name", ".agency", ".agency-name", ".seller", ".seller-name", ".user-name", ".contact .name", ".author"]:
            obj=soup.select_one(sel)
            if obj:
                txt=limpar_texto(obj.get_text(" ", strip=True))
                if txt and len(txt) > 2:
                    anunciante=txt
                    break
        if not anunciante:
            anunciante = gc(["anunciante", "agente", "imobiliária", "imobiliaria", "contacto"])
            anunciante = limpar_texto(anunciante) or None

        # ── Imagem principal ──────────────────────────────────
        url_imagem=None
        og=soup.select_one("meta[property='og:image'], meta[name='twitter:image']")
        if og and og.get("content"):
            url_imagem=url_absoluta(self.BASE, og.get("content"))
        if not url_imagem:
            img=soup.select_one("img[src*='angocasa'], img[src*='upload'], img[src*='imoveis'], img[src]")
            if img and img.get("src"):
                url_imagem=url_absoluta(self.BASE, img.get("src"))

        return {
            "fonte":       "AngoCasa",
            "id_fonte":    id_an,
            "url":          url,
            "id_angocasa":  id_an,
            "tipo_negocio": negocio,
            "pais":         pais_d or "Angola",
            "provincia":    regiao or "Luanda",
            "bairro":       bairro,
            "lat":          lat,
            "lon":          lon,
            "headline":     titulo_completo,
            "headline_completa": titulo_completo,
            "titulo_anuncio": titulo_an or titulo_completo,
            "tipo_imovel":  tipo_imovel,
            "tipologia":    tip,
            "estado":       estado,
            "area_m2":      area_m2,
            "area_util_m2": area_util,
            "area_bruta_m2":area_bruta,
            "area_lote_m2": area_lote,
            "preco_original":p_orig,
            "moeda":        moeda_n,
            "preco_akz":    p_akz,
            "preco_usd":    p_usd,
            "preco_m2_akz": preco_m2,
            "quartos":      quartos,
            "wc":           wc_n,
            "mobilado":     mobilado,
            "garagem":      tem_kw(tudo,["garagem","estacionamento", "parking"]),
            "piscina":      tem_kw(tudo,["piscina"]),
            "gerador":      tem_kw(tudo,["gerador"]),
            "reservatorio": tem_kw(tudo,["reservatório","reservatorio","tanque de água","tanque de agua"]),
            "seguranca":    tem_kw(tudo,["segurança 24","seguranca 24","vigilância","vigilancia", "portaria"]),
            "churrasqueira":tem_kw(tudo,["churrasqueira"]),
            "ar_condicionado":tem_kw(tudo,["ar condicionado","ar-condicionado","a/c"]),
            "internet":     tem_kw(tudo,["internet","fibra","wifi","wi-fi"]),
            "condominio_fechado":tem_kw(tudo,["condomínio fechado","condominio fechado","condomínio privado","condominio privado", "gated", "portaria"]),
            "publicado_em": pub,
            "descricao":    descricao.strip(),
            "anunciante": anunciante,
            "telefone_whatsapp": telefone,
            "url_imagem": url_imagem,
            "raw_json": json.dumps({"fonte":"AngoCasa", "campos": campos}, ensure_ascii=False),
        }


    def recolher_listagem(self, neg_slug, neg_nome, bairro_nome, bairro_slug, forcar=False):
        """
        Percorre todas as páginas de um bairro+negócio.
        Retorna (lista_dados, coleta_completa).
        """
        resultados=[]
        urls_vistas=set()
        contexto = bairro_nome or "Luanda geral"
        pagina=self.db.obter_checkpoint("AngoCasa", neg_nome, contexto)
        if pagina > 1:
            log.info(f"    [AngoCasa] Retomando do checkpoint: {contexto} | pág {pagina}")
        coleta_completa=False
        paginas_sem_extracao=0
        log.info(f"  ▶ {bairro_nome} | {neg_nome}")

        while pagina<=CFG["max_paginas"]:
            url_pag=self.construir_url(neg_slug,bairro_slug,pagina)
            log.debug(f"    Pág {pagina}: {url_pag}")
            soup=self._get(url_pag)
            if not soup:
                log.warning(f"    Sem resposta na pág {pagina} — parar coleta.")
                break

            cards=soup.select("a.listedcar[href]")
            if not cards:
                # Fallback: se o CSS do site mudar, procurar links prováveis de imóveis/anúncios.
                candidatos = soup.select("a[href]")
                cards = [a for a in candidatos if any(x in (a.get("href") or "").lower() for x in ["/anuncio/", "/anuncios/", "/imovel/", "/apartamento/", "/vivenda/", "/terreno/"])]
            if not cards:
                log.info(f"    Sem cards na pág {pagina} — fim da paginação.")
                coleta_completa=True
                break

            novos_nesta_pag=0
            for el in cards:
                dados_card=self.extrair_card(el)
                if not dados_card: continue
                u=dados_card["url"]
                if u in urls_vistas:
                    continue
                urls_vistas.add(u)

                # Retomada inteligente: se já foi gravado há poucas horas, não reprocessar
                if not forcar and self.db.anuncio_processado_recente(u):
                    log.debug(f"    Ignorado (já processado recentemente): {u}")
                    continue

                # Visitar página de detalhe
                self._delay()
                dados=self.extrair_detalhe(u,dados_card)
                if not dados:
                    self.db.reg_erro(u,"detalhe_falhou","sem resposta")
                    continue

                # Preencher bairro a partir do slug se não veio do detalhe
                if not dados.get("bairro") or dados["bairro"]=="":
                    dados["bairro"]=bairro_nome

                # Garantir coordenadas
                if not dados.get("lat"):
                    for b in CFG["bairros"]:
                        if b["nome"]==bairro_nome:
                            dados["lat"]=b["lat"]; dados["lon"]=b["lon"]; break

                resultados.append(dados)
                novos_nesta_pag+=1
                if CFG.get("max_anuncios_por_fonte") and len(resultados) >= int(CFG["max_anuncios_por_fonte"]):
                    log.info(f"    Limite de teste atingido: {len(resultados)} anúncios ({contexto}/{neg_nome})")
                    break

            log.info(f"    Pág {pagina}: {novos_nesta_pag} anúncios extraídos "
                     f"(total acum.: {len(resultados)})")
            self.db.salvar_checkpoint("AngoCasa", neg_nome, pagina + 1, contexto)

            if CFG.get("max_anuncios_por_fonte") and len(resultados) >= int(CFG["max_anuncios_por_fonte"]):
                coleta_completa = False
                break

            if novos_nesta_pag == 0:
                paginas_sem_extracao += 1
            else:
                paginas_sem_extracao = 0

            # Evita percorrer dezenas de páginas quando o site responde, mas o seletor não encontra anúncios úteis.
            if paginas_sem_extracao >= 3:
                log.warning("    3 páginas consecutivas sem extração útil — parar para evitar varredura vazia.")
                coleta_completa = False
                break

            # Verificar se há próxima página
            prox=soup.select_one(f'a[data-page="{pagina}"]')
            if not prox and pagina>1:
                coleta_completa=True
                break
            if not soup.select_one('a[data-page]'):
                coleta_completa=True
                break

            pagina+=1
            self._delay()

        if coleta_completa:
            self.db.limpar_checkpoint("AngoCasa", neg_nome, contexto)
        log.info(f"  ✔ {bairro_nome}/{neg_nome}: {len(resultados)} anúncios "
                 f"| completa={coleta_completa}")
        return resultados, coleta_completa


    def recolher_listagem_e_gravar(self, db, stats, neg_slug, neg_nome, bairro_nome, bairro_slug, forcar=False):
        """
        Versão incremental do coletor AngoCasa.
        Grava cada página imediatamente no SQLite, em vez de esperar terminar todo o bloco.
        Isso evita perda de progresso em interrupções e acelera retomadas.
        """
        urls_ativas = []
        urls_vistas = set()
        contexto = bairro_nome or "Luanda geral"
        pagina = self.db.obter_checkpoint("AngoCasa", neg_nome, contexto)
        if pagina > 1:
            log.info(f"    [AngoCasa] Retomando do checkpoint: {contexto} | pág {pagina}")
        coleta_completa = False
        paginas_sem_extracao = 0
        total_extraidos = 0
        log.info(f"  ▶ {bairro_nome} | {neg_nome}")

        while pagina <= CFG["max_paginas"]:
            url_pag = self.construir_url(neg_slug, bairro_slug, pagina)
            log.debug(f"    Pág {pagina}: {url_pag}")
            soup = self._get(url_pag)
            if not soup:
                log.warning(f"    Sem resposta na pág {pagina} — parar coleta.")
                break

            cards = soup.select("a.listedcar[href]")
            if not cards:
                candidatos = soup.select("a[href]")
                cards = [a for a in candidatos if any(x in (a.get("href") or "").lower()
                         for x in ["/anuncio/", "/anuncios/", "/imovel/", "/apartamento/", "/vivenda/", "/terreno/"])]

            if not cards:
                log.info(f"    Sem cards na pág {pagina} — fim da paginação.")
                coleta_completa = True
                break

            dados_pagina = []
            novos_nesta_pag = 0

            for el in cards:
                dados_card = self.extrair_card(el)
                if not dados_card:
                    continue
                u = dados_card["url"]
                if u in urls_vistas:
                    continue
                urls_vistas.add(u)

                if not forcar and self.db.anuncio_processado_recente(u):
                    log.debug(f"    Ignorado (já processado recentemente): {u}")
                    continue

                self._delay()
                dados = self.extrair_detalhe(u, dados_card)
                if not dados:
                    self.db.reg_erro(u, "detalhe_falhou", "sem resposta")
                    continue

                if not dados.get("bairro") or dados["bairro"] == "":
                    dados["bairro"] = bairro_nome

                if not dados.get("lat"):
                    for b in CFG["bairros"]:
                        if b["nome"] == bairro_nome:
                            dados["lat"] = b["lat"]
                            dados["lon"] = b["lon"]
                            break

                dados_pagina.append(dados)
                novos_nesta_pag += 1

            # Grava imediatamente esta página. É aqui que a v2.6 perdia progresso ao interromper.
            if dados_pagina:
                urls_ativas.extend(self._processar_dados_lista(db, dados_pagina, stats))

            total_extraidos += novos_nesta_pag
            log.info(f"    Pág {pagina}: {novos_nesta_pag} anúncios extraídos "
                     f"(total acum.: {total_extraidos})")

            self.db.salvar_checkpoint("AngoCasa", neg_nome, pagina + 1, contexto)

            if CFG.get("max_anuncios_por_fonte") and len(resultados) >= int(CFG["max_anuncios_por_fonte"]):
                coleta_completa = False
                break

            if novos_nesta_pag == 0:
                paginas_sem_extracao += 1
            else:
                paginas_sem_extracao = 0

            if paginas_sem_extracao >= 3:
                log.warning("    3 páginas consecutivas sem extração útil — parar para evitar varredura vazia.")
                coleta_completa = False
                break

            prox = soup.select_one(f'a[data-page="{pagina}"]')
            if not prox and pagina > 1:
                coleta_completa = True
                break
            if not soup.select_one('a[data-page]'):
                coleta_completa = True
                break

            pagina += 1
            self._delay()

        if coleta_completa:
            self.db.limpar_checkpoint("AngoCasa", neg_nome, contexto)

        log.info(f"  ✔ {bairro_nome}/{neg_nome}: {total_extraidos} anúncios "
                 f"| completa={coleta_completa}")
        return urls_ativas, coleta_completa

    def _processar_dados_lista(self, db, dados_lista, stats):
        """Grava anúncios extraídos e atualiza estatísticas de execução."""
        urls_ativas = []
        for d in dados_lista:
            try:
                d = enriquecer_campos_padrao(d)
                aid, ev = db.upsert(d)
                urls_ativas.append(d["url"])
                stats["n_coletados"] += 1
                if ev == "novo":
                    stats["n_novos"] += 1
                elif ev == "atualizado":
                    stats["n_atualizados"] += 1
                elif ev == "preco_alterado":
                    stats["n_atualizados"] += 1
                    stats["n_preco_alt"] += 1
            except Exception as e:
                log.error(f"Erro upsert {d.get('url','?')}: {e}")
                db.reg_erro(d.get("url", "?"), "upsert_erro", str(e))
                stats["n_erros"] += 1
        return urls_ativas


    # ─────────────────────────────────────────────────────────
    # COLETOR GENÉRICO MULTI-FONTE
    # ─────────────────────────────────────────────────────────
    def _urls_paginas_fonte(self, cfg_fonte: Dict, negocio: str, pagina: int) -> List[str]:
        """Gera URLs candidatas para listagem em fontes não-AngoCasa."""
        base = cfg_fonte["base"].rstrip("/")
        paths = cfg_fonte.get("paths", {}).get(negocio, ["/"])
        urls = []
        for path in paths:
            path = path if path.startswith("/") else f"/{path}"
            u = f"{base}{path}"
            if pagina == 1:
                urls.append(u)
            else:
                # Opção 1: paginação por querystring.
                for param in cfg_fonte.get("page_params", ["page"]):
                    parsed = urlparse(u)
                    q = dict(parse_qsl(parsed.query))
                    q[param] = str(pagina)
                    urls.append(urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", urlencode(q), "")))
                # Opção 2: paginação por caminho.
                for patt in cfg_fonte.get("page_path_patterns", []):
                    urls.append(f"{u.rstrip('/')}{patt.format(pagina=pagina)}")
        # Remover duplicados mantendo ordem
        seen=set(); out=[]
        for u in urls:
            if u not in seen:
                seen.add(u); out.append(u)
        return out

    def _link_parece_anuncio(self, href: str, texto: str, cfg_fonte: Dict) -> bool:
        """Filtro conservador para separar links de imóveis de menus, redes sociais e assets."""
        if not href:
            return False
        h = href.lower()
        if any(x in h for x in ["#", "javascript:", "mailto:", "tel:", "whatsapp", "facebook", "instagram", "linkedin"]):
            return False
        if any(h.endswith(ext) for ext in [".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf", ".css", ".js"]):
            return False
        texto_full = (h + " " + (texto or "").lower())
        kws = cfg_fonte.get("link_keywords", []) + ["t1", "t2", "t3", "t4", "venda", "arrendar", "rent", "sale"]
        return any(k.lower() in texto_full for k in kws)

    def extrair_urls_listagem_generica(self, soup, cfg_fonte: Dict) -> List[str]:
        base = cfg_fonte["base"]
        dominio = dominio_base(base)
        urls=[]
        for a in soup.select("a[href]"):
            href = a.get("href") or ""
            texto = a.get_text(" ", strip=True)
            u = url_absoluta(base, href)
            if not u or dominio not in dominio_base(u):
                continue
            if self._link_parece_anuncio(u, texto, cfg_fonte):
                urls.append(u)
        # Remover links genéricos de listagens/categorias que tendem a gerar ruído.
        filtradas=[]; seen=set()
        for u in urls:
            path = urlparse(u).path.strip("/").lower()
            if path in ["", "comprar", "venda", "arrendar", "alugar", "imoveis", "properties", "property", "for-sale", "for-rent"]:
                continue
            if u not in seen:
                seen.add(u); filtradas.append(u)
        return filtradas[:120]

    def extrair_detalhe_generico(self, url: str, fonte_nome: str, negocio_fallback: str):
        """Extrai dados de detalhe de fontes com HTML variado."""
        soup = self._get(url)
        if not soup:
            return None
        self._delay()

        for tag in soup(["script", "style", "noscript", "svg"]):
            tag.decompose()

        titulo_candidates=[]
        for sel in ["h1", "h2", ".title", ".property-title", ".listing-title", "title", "meta[property='og:title']"]:
            obj=soup.select_one(sel)
            if not obj:
                continue
            txt = obj.get("content") if obj.name == "meta" else obj.get_text(" ", strip=True)
            txt = limpar_texto(txt)
            if txt:
                titulo_candidates.append(txt)
        body_text = limpar_texto(soup.get_text(" ", strip=True), limite=12000)
        headline = max(titulo_candidates, key=len) if titulo_candidates else limpar_texto(urlparse(url).path.split("/")[-1].replace("-", " ").title())

        preco_orig, moeda = extrair_preco_texto(body_text + " " + headline)
        if moeda == "USD":
            preco_usd = preco_orig
            preco_akz = usd2akz(preco_usd)
        else:
            moeda = "AKZ"
            preco_akz = preco_orig
            preco_usd = akz2usd(preco_akz)

        area_m2 = norm_area(body_text)
        preco_m2 = round(preco_akz / area_m2, 2) if preco_akz and area_m2 and area_m2 > 0 else None
        bairro, lat, lon = detectar_bairro(body_text + " " + headline)
        tipo_imovel = detectar_tipo_imovel(body_text + " " + headline, url)
        tipologia = norm_tipologia(body_text + " " + headline)
        negocio = detectar_negocio(body_text + " " + headline, url, negocio_fallback)

        q_match = re.search(r"(\d+)\s*(?:quarto|bedroom|suite)", body_text.lower())
        quartos = int(q_match.group(1)) if q_match else None
        if quartos is None and tipologia:
            q_tip = re.search(r"\d+", tipologia)
            quartos = int(q_tip.group()) if q_tip else None

        wc_match = re.search(r"(\d+)\s*(?:wc|casa(?:s)? de banho|banheiro|bathroom)", body_text.lower())
        wc = int(wc_match.group(1)) if wc_match else None
        telefone = extrair_telefone(body_text)
        imagem = None
        og_img = soup.select_one("meta[property='og:image']")
        if og_img and og_img.get("content"):
            imagem = url_absoluta(url, og_img.get("content"))

        id_fonte = hashlib.md5(url.encode("utf-8")).hexdigest()[:16]
        descricao = body_text[:5000]

        # Só grava se houver pelo menos um sinal imobiliário relevante.
        sinais = [preco_orig, area_m2, bairro, tipologia, telefone, tipo_imovel not in (None, "outro")]
        if sum(1 for x in sinais if x) < 2 and len(headline) < 8:
            return None

        return {
            "fonte": fonte_nome,
            "id_fonte": id_fonte,
            "url": url,
            "id_angocasa": None,
            "tipo_negocio": negocio,
            "pais": "Angola",
            "provincia": "Luanda",
            "bairro": bairro or "Não informado",
            "lat": lat,
            "lon": lon,
            "headline": headline,
            "titulo_anuncio": headline,
            "tipo_imovel": tipo_imovel,
            "tipologia": tipologia,
            "estado": None,
            "area_m2": area_m2,
            "area_util_m2": None,
            "area_bruta_m2": area_m2,
            "area_lote_m2": None,
            "preco_original": preco_orig,
            "moeda": moeda,
            "preco_akz": preco_akz,
            "preco_usd": preco_usd,
            "preco_m2_akz": preco_m2,
            "quartos": quartos,
            "wc": wc,
            "mobilado": tem_kw(body_text, ["mobilado", "mobiliado", "furnished"]),
            "garagem": tem_kw(body_text, ["garagem", "estacionamento", "parking"]),
            "piscina": tem_kw(body_text, ["piscina", "pool"]),
            "gerador": tem_kw(body_text, ["gerador", "generator"]),
            "reservatorio": tem_kw(body_text, ["reservatório", "reservatorio", "tanque de água", "tanque de agua"]),
            "seguranca": tem_kw(body_text, ["segurança 24", "seguranca 24", "vigilância", "vigilancia", "security"]),
            "churrasqueira": tem_kw(body_text, ["churrasqueira", "barbecue", "bbq"]),
            "ar_condicionado": tem_kw(body_text, ["ar condicionado", "ar-condicionado", "air conditioning", "a/c"]),
            "internet": tem_kw(body_text, ["internet", "fibra", "wifi", "wi-fi"]),
            "condominio_fechado": tem_kw(body_text, ["condomínio fechado", "condominio fechado", "condomínio privado", "condominio privado", "gated", "portaria", "segurança 24"]),
            "publicado_em": None,
            "descricao": descricao,
            "anunciante": None,
            "telefone_whatsapp": telefone,
            "url_imagem": imagem,
            "raw_json": json.dumps({"fonte": fonte_nome, "url": url}, ensure_ascii=False),
        }

    def recolher_fonte_generica(self, chave_fonte: str, negocio: str, forcar=False):
        cfg_fonte = FONTES_EXTRA[chave_fonte]
        fonte_nome = cfg_fonte["nome"]
        resultados=[]
        urls_vistas=set()
        paginas_sem_links=0
        log.info(f"  ▶ {fonte_nome} | {negocio}")

        pagina_ini = self.db.obter_checkpoint(fonte_nome, negocio)
        if pagina_ini > 1:
            log.info(f"    [{fonte_nome}] Retomando do checkpoint: pág {pagina_ini}")
        for pagina in range(pagina_ini, CFG["max_paginas"] + 1):
            urls_paginas = self._urls_paginas_fonte(cfg_fonte, negocio, pagina)
            links_pagina=[]
            for url_pag in urls_paginas:
                soup = self._get(url_pag)
                if not soup:
                    continue
                links = self.extrair_urls_listagem_generica(soup, cfg_fonte)
                if links:
                    links_pagina.extend(links)
                    # Se uma URL candidata funcionou para esta página, não precisa testar todas.
                    break
                self._delay()

            # Remover duplicados da página
            links_pagina = list(dict.fromkeys(links_pagina))
            if not links_pagina:
                paginas_sem_links += 1
                log.info(f"    {fonte_nome} pág {pagina}: sem links úteis.")
                if paginas_sem_links >= 2:
                    break
                continue
            paginas_sem_links = 0

            extraidos=0
            for u in links_pagina:
                if u in urls_vistas:
                    continue
                urls_vistas.add(u)
                if not forcar and self.db.anuncio_processado_recente(u):
                    continue
                try:
                    dados = self.extrair_detalhe_generico(u, fonte_nome, negocio)
                    if not dados:
                        continue
                    resultados.append(dados)
                    extraidos += 1
                except Exception as e:
                    log.debug(f"    Falha detalhe {fonte_nome}: {u} | {e}")
                    self.db.reg_erro(u, "detalhe_generico", str(e))
            log.info(f"    {fonte_nome} pág {pagina}: {extraidos} anúncios extraídos (total acum.: {len(resultados)})")
            self.db.salvar_checkpoint(fonte_nome, negocio, pagina + 1)
            self._delay()

        return resultados, False     
     
    def executar_fonte(self, fonte, forcar=False,  bairro_filtro=None, por_bairros=None):

        try:
            if fonte == "angocasa":
                db = DB()
                st = self.scraping_completo( db, bairro_filtro=bairro_filtro, por_bairros=por_bairros, forcar=forcar)
                return fonte, st

            if fonte == "remax":
                db = DB()
                scraper = ScraperRemax(db)
            elif fonte == "angoimoveis":
                db = DB()
                scraper = ScraperAngoImoveis(db)
            elif fonte == "mia":
                db = DB()
                scraper = ScraperMia(db)
            elif fonte == "zenki":
                db = DB()
                scraper = ScraperZenki(db)
            elif fonte == "ree_gera":
                db = DB()
                scraper = ScraperReeGera(db)
            else:
                return fonte, []

            dados, _ = scraper.recolher(forcar=forcar)
            return fonte, dados
        except Exception as e:
            log.error(f'error {e}')
        

    def scraping_multifonte(self, db, bairro_filtro=None, por_bairros=False, forcar=False, fontes=None):
        """Executa AngoCasa + fontes extras selecionadas."""
        fontes = fontes or ["angocasa"]
        if "todas" in fontes or "all" in fontes:
            fontes = FONTES_DISPONIVEIS
        fontes = [f.lower().strip() for f in fontes if f]

        stats_total = {
            "n_novos": 0, "n_atualizados": 0, "n_preco_alt": 0,
            "n_coletados": 0, "n_removidos": 0, "n_erros": 0,
            "bairros_proc": "",
        }
        partes_proc=[]
        try:
            with ThreadPoolExecutor(max_workers=len(fontes)) as executor:
                futures = [executor.submit(self.executar_fonte, fonte, bairro_filtro, por_bairros, forcar) for fonte in fontes]

                for future in as_completed(futures):
                    fonte, result = future.result()

                # AngoCasa
                if fonte == "angocasa":
                    for k in stats_total:

                        if (k in result and isinstance(stats_total[k], int)):
                            stats_total[k] += result[k]

                    partes_proc.append(result.get("bairros_proc", "AngoCasa" ))

                # Other scrapers
                else:

                    self._processar_dados_lista(db, result, stats_total)

                    partes_proc.append(fonte)
        except Exception as e:
            log.error(f"Error coleta: {e}")
            stats_total["n_erros"] += 1
        stats_total["bairros_proc"] = "; ".join(partes_proc)
        return stats_total

        """
        if "angocasa" in fontes:
            db=DB()
            st = self.scraping_completo(db, bairro_filtro=bairro_filtro, por_bairros=por_bairros, forcar=forcar)
            for k in stats_total:
                if k in st and isinstance(stats_total[k], int):
                    stats_total[k] += st[k]
            partes_proc.append(st.get("bairros_proc", "AngoCasa"))

        # RE/MAX: scraper dedicado via API interna + detalhe SSR.
        # Não usar o coletor genérico, pois a listagem é React SPA.
        if "remax" in fontes:
            try:
                scraper_remax = ScraperRemax(db)
                dados_lista, _ = scraper_remax.recolher(forcar=forcar)
                self._processar_dados_lista(db, dados_lista, stats_total)
                partes_proc.append("RE/MAX Angola")
            except Exception as e:
                log.error(f"Erro coleta RE/MAX: {e}")
                stats_total["n_erros"] += 1

        # AngoImóveis: scraper dedicado com URLs /category/venda e /category/alugar.
        # Não usar o coletor genérico, pois os caminhos e seletores são específicos.
        if "angoimoveis" in fontes:
            try:
                scraper_ai = ScraperAngoImoveis(db)
                dados_ai, _ = scraper_ai.recolher(forcar=forcar)
                self._processar_dados_lista(db, dados_ai, stats_total)
                partes_proc.append("AngoImóveis")
            except Exception as e:
                log.error(f"Erro coleta AngoImóveis: {e}")
                stats_total["n_erros"] += 1

        # MIA: scraper dedicado via API Supabase (RPC search_properties + REST properties).
        # Não usar o coletor genérico, pois o site é SPA e a listagem não é HTML estática.
        if "mia" in fontes:
            try:
                scraper_mia = ScraperMia(db)
                dados_mia, _ = scraper_mia.recolher(forcar=forcar)
                self._processar_dados_lista(db, dados_mia, stats_total)
                partes_proc.append("MIA")
            except Exception as e:
                log.error(f"Erro coleta MIA: {e}")
                stats_total["n_erros"] += 1

        # Zenki Real Estate: scraper dedicado via AJAX WordPress + detalhe SSR.
        if "zenki" in fontes:
            try:
                scraper_zenki = ScraperZenki(db)
                dados_zenki, _ = scraper_zenki.recolher(forcar=forcar)
                self._processar_dados_lista(db, dados_zenki, stats_total)
                partes_proc.append("Zenki Real Estate")
            except Exception as e:
                log.error(f"Erro coleta Zenki: {e}")
                stats_total["n_erros"] += 1

        # REE-GERA: scraper dedicado WordPress/wpResidence.
        if "ree_gera" in fontes:
            try:
                scraper_rg = ScraperReeGera(db)
                dados_rg, _ = scraper_rg.recolher(forcar=forcar)
                self._processar_dados_lista(db, dados_rg, stats_total)
                partes_proc.append("REE-GERA")
            except Exception as e:
                log.error(f"Erro coleta REE-GERA: {e}")
                stats_total["n_erros"] += 1

        negocios = ["compra", "arrendamento"]
        for f in fontes:
            if f in ("angocasa", "remax", "angoimoveis", "mia", "zenki", "ree_gera"):
                continue
            if f not in FONTES_EXTRA:
                log.warning(f"Fonte ignorada: {f}. Fontes disponíveis: {', '.join(FONTES_DISPONIVEIS)}")
                continue
            for negocio in negocios:
                try:
                    dados_lista, _ = self.recolher_fonte_generica(f, negocio, forcar=forcar)
                    self._processar_dados_lista(db, dados_lista, stats_total)
                    partes_proc.append(f"{FONTES_EXTRA[f]['nome']}:{negocio}")
                except Exception as e:
                    log.error(f"Erro coleta {f}/{negocio}: {e}")
                    stats_total["n_erros"] += 1

        stats_total["bairros_proc"] = "; ".join([p for p in partes_proc if p])
        """

    def scraping_completo(self, db, bairro_filtro=None, por_bairros=False, forcar=False):
        """
        Ponto de entrada principal do scraping.

        Modo padrão, sem bairro e sem --por-bairros:
            1) coleta toda Luanda em compra/venda;
            2) coleta toda Luanda em arrendamento;
            3) bairro/localidade fica para filtrar depois no Excel.

        Modo com --bairro:
            coleta apenas o bairro informado.

        Modo com --por-bairros:
            percorre todos os bairros configurados, um a um.
        """
        negocios = [("venda", "compra"), ("arrenda", "arrendamento")]
        apenas = (CFG.get("apenas_negocio") or "todos").lower().strip()
        if apenas in ("compra", "arrendamento"):
            negocios = [n for n in negocios if n[1] == apenas]
        stats = {
            "n_novos": 0,
            "n_atualizados": 0,
            "n_preco_alt": 0,
            "n_coletados": 0,     
            "n_removidos": 0,
            "n_erros": 0,
        }
        bairros_proc = []

        # 1) Bairro específico: mantém a funcionalidade antiga sob comando explícito.
        if bairro_filtro:
            bairros = [b for b in CFG["bairros"] if b["nome"].lower() == bairro_filtro.lower()]
            if not bairros:
                log.error(f"Bairro '{bairro_filtro}' não encontrado na configuração.")
                return stats

            for b in bairros:
                bairros_proc.append(b["nome"])
                for neg_slug, neg_nome in negocios:
                    try:
                        urls_ativas, completa = self.recolher_listagem_e_gravar(
                            db, stats, neg_slug, neg_nome, b["nome"], b["slug"], forcar=forcar
                        )
                        db.marcar_inativos(b["nome"], neg_nome, urls_ativas, completa)

                        conn = db._con()
                        try:
                            c2 = conn.cursor()
                            c2.execute(
                                "SELECT COUNT(*) FROM anuncios_removidos WHERE bairro=? "
                                "AND tipo_negocio=? AND date(data_remocao)=date('now')",
                                (b["nome"], neg_nome),
                            )
                            stats["n_removidos"] += c2.fetchone()[0]
                        finally:
                            conn.close()
                    except Exception as e:
                        log.error(f"Erro coleta {b['nome']}/{neg_nome}: {e}")
                        stats["n_erros"] += 1

            stats["bairros_proc"] = ", ".join(bairros_proc)
            return stats

        # 2) Modo antigo opcional: varrer bairro por bairro.
        if por_bairros:
            bairros = CFG["bairros"]
            for b in bairros:
                bairros_proc.append(b["nome"])
                for neg_slug, neg_nome in negocios:
                    try:
                        urls_ativas, completa = self.recolher_listagem_e_gravar(
                            db, stats, neg_slug, neg_nome, b["nome"], b["slug"], forcar=forcar
                        )
                        db.marcar_inativos(b["nome"], neg_nome, urls_ativas, completa)
                    except Exception as e:
                        log.error(f"Erro coleta {b['nome']}/{neg_nome}: {e}")
                        stats["n_erros"] += 1

            stats["bairros_proc"] = ", ".join(bairros_proc)
            return stats

        # 3) Novo modo padrão: Luanda inteira, sem filtro de bairro.
        bairros_proc.append("Luanda geral")
        for neg_slug, neg_nome in negocios:
            try:
                # bairro_slug vazio gera URL: /anuncios/{neg}/angola-luanda/...
                urls_ativas, completa = self.recolher_listagem_e_gravar(
                    db, stats, neg_slug, neg_nome, "Luanda geral", "", forcar=forcar
                )

                # Não marcar inativos no modo geral quando max_paginas pode limitar a coleta.
                # A remoção por ausência deve ser implementada apenas quando houver varredura total confiável.
                if completa:
                    log.info("  Coleta geral completa; inativação automática desativada por segurança.")
            except Exception as e:
                log.error(f"Erro coleta Luanda geral/{neg_nome}: {e}")
                stats["n_erros"] += 1

        stats["bairros_proc"] = ", ".join(bairros_proc)
        return stats


# ─────────────────────────────────────────────────────────────────
# EXPORTAÇÃO EXCEL
# ─────────────────────────────────────────────────────────────────
def _estilo_cabecalho(ws, linha=1):
    """Aplica estilo ao cabeçalho de uma folha."""
    if not EXCEL_OK: return
    fill  = PatternFill("solid", fgColor="C0392B")
    fonte = Font(bold=True, color="FFFFFF", size=11)
    borda = Border(
        bottom=Side(style="medium"),
        right=Side(style="thin")
    )
    for cell in ws[linha]:
        cell.fill=fill; cell.font=fonte
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=borda

def _auto_largura(ws, max_col=50):
    """Ajusta largura automática das colunas, com limites para manter legibilidade."""
    if not EXCEL_OK:
        return
    for col in ws.columns:
        max_w = 10
        letra = get_column_letter(col[0].column)
        cabecalho = str(col[0].value or "").strip().lower()
        for cell in col:
            try:
                valor = str(cell.value) if cell.value is not None else ""
                max_w = max(max_w, min(len(valor) + 2, 80))
            except Exception:
                pass
        if cabecalho in ("headline", "titulo_anuncio", "descricao", "url"):
            max_w = min(max_w, 90 if cabecalho in ("headline", "titulo_anuncio") else 80)
        else:
            max_w = min(max_w, 28)
        ws.column_dimensions[letra].width = max_w

def _zebra(ws, linha, par_cor="F2F2F2", impar_cor="FFFFFF"):
    """Aplica cor alternada às linhas."""
    if not EXCEL_OK:
        return
    cor = par_cor if linha % 2 == 0 else impar_cor
    fill = PatternFill("solid", fgColor=cor)
    for cell in ws[linha]:
        cell.fill = fill

def _formatar_numero(cell, casas_decimais=2):
    """
    Formata números sem símbolo.
    Usa o padrão nativo do Excel (#,##0.00 / #,##0), que no Excel em PT
    aparece como 8 500 000,00 e 20 000, respeitando a localidade do sistema.
    """
    if casas_decimais == 0:
        cell.number_format = '# ##0'
    else:
        cell.number_format = '# ##0,00'
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _aplicar_formatacao_padrao(ws):
    """Formatação padrão: números centralizados, textos preservados e cabeçalho fixo."""
    if not EXCEL_OK or ws.max_row < 1:
        return

    campos_valor = {
        "preco_original", "preco_akz", "preco_usd", "preco_m2_akz",
        "preco_akz_ant", "variacao_akz", "variacao_pct", "valor",
        "renda_media", "renda_med", "renda_m2", "yield_bruto",
        "yield_liquido", "payback_bruto", "payback_liquido", "cambio",
        "taxa câmbio akz/usd"
    }
    ampos_quantidade = {
        "id", "anuncio_id", "id_angocasa", "area_m2", "area_util_m2",
        "area_bruta_m2", "area_lote_m2", "quartos", "wc", "mobilado",
        "garagem", "piscina", "gerador", "reservatorio", "seguranca",
        "churrasqueira", "ar_condicionado", "internet", "condominio_fechado",
        "ativo", "n_atualizacoes", "hist_precos", "preco_alt", "removidos",
        "erros", "bairros", "total", "bairros configurados", "items por página",
        "máx. páginas", "imovel_corporativo", "preco_disponivel"
    }
    campos_texto = {"headline", "headline_completa", "titulo_anuncio", "descricao", "url", "observacao", "detalhes",
                    "email_agente", "segmento_imobiliario", "hash_dedupe", "id_zenki", "url_brochura", "pdf_brochura"}

    headers = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(1, col).value
        headers[col] = str(valor).strip().lower() if valor is not None else ""

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            campo = headers.get(cell.column, "")
            if cell.value is None:
                continue
            if campo in campos_valor:
                _formatar_numero(cell, 2)
            elif campo in campos_quantidade:
                _formatar_numero(cell, 0)
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                casas = 0 if float(cell.value).is_integer() else 2
                _formatar_numero(cell, casas)
            elif campo in campos_texto:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    _auto_largura(ws)
    # Linhas um pouco mais altas para permitir leitura de headlines longas
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 24


def _hyperlinkar_url(ws, nome_coluna_url="url", adicionar_coluna_abrir=True):
    """
    Torna a coluna URL clicável e, opcionalmente, adiciona a coluna 'Abrir'
    com um hiperlink curto no início da folha.
    """
    if not EXCEL_OK or ws.max_row < 2:
        return
    url_col = None
    for j in range(1, ws.max_column + 1):
        if str(ws.cell(1, j).value or "").strip().lower() == nome_coluna_url:
            url_col = j
            break
    if not url_col:
        return
    link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
    for i in range(2, ws.max_row + 1):
        u = ws.cell(i, url_col).value
        if isinstance(u, str) and u.startswith("http"):
            cell = ws.cell(i, url_col)
            cell.hyperlink = u
            cell.font = link_font
    if adicionar_coluna_abrir:
        ws.insert_cols(1)
        ws.cell(1, 1, "Abrir")
        for i in range(2, ws.max_row + 1):
            u = ws.cell(i, url_col + 1).value
            if isinstance(u, str) and u.startswith("http"):
                c = ws.cell(i, 1, "Abrir")
                c.hyperlink = u
                c.font = link_font
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 10

def _adicionar_tabela(ws, nome_base):
    """Transforma a área usada em tabela nativa do Excel, quando possível."""
    if not EXCEL_OK or ws.max_row < 2 or ws.max_column < 1:
        return
    nome = re.sub(r"[^A-Za-z0-9_]", "_", nome_base)[:240]
    if not nome or nome[0].isdigit():
        nome = f"Tabela_{nome}"
    ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
    try:
        tabela = Table(displayName=nome, ref=ref)
        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
    except Exception as exc:
        log.debug(f"Não foi possível criar tabela em {ws.title}: {exc}")

def _valor_para_excel(campo, valor):
    """Converte valores binários de características para Sim/Não na exportação."""
    campos_sim_nao = {
        "mobilado", "garagem", "piscina", "gerador", "reservatorio",
        "seguranca", "churrasqueira", "ar_condicionado", "internet",
        "condominio_fechado", "ativo"
    }
    if campo in campos_sim_nao:
        if valor in (1, True, "1", "Sim", "sim", "SIM"):
            return "Sim"
        if valor in (0, False, "0", "Não", "Nao", "não", "nao", "NÃO", "NAO"):
            return "Não"
    return valor

def exportar_excel(db):
    """
    Gera ficheiro Excel com 5 abas:
      1. Base Anúncios
      2. Histórico de Preços
      3. Alterações da Semana
      4. Parâmetros
      5. Logs (stats da BD)
    """
    if not EXCEL_OK:
        log.error("openpyxl não instalado. Execute: pip install openpyxl")
        return None

    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    out = CFG["excel_dir"]
    out.mkdir(parents=True,exist_ok=True)
    ficheiro = out / f"PIIA_Fase1_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remover folha vazia

    # ── ABA 1: Base Anúncios ──────────────────────────────────
    ws1=wb.create_sheet("Base Anúncios")
    anuncios=db.obter_todos()
    cols1=[
        "fonte", "id_fonte", "id_angocasa", "tipo_negocio", "headline", "headline_completa", "bairro", "provincia",
        "tipo_imovel", "tipologia", "estado", "area_m2", "area_util_m2",
        "area_bruta_m2", "area_lote_m2", "preco_original", "moeda",
        "preco_akz", "preco_usd", "preco_m2_akz", "quartos", "wc",
        "mobilado", "garagem", "piscina", "gerador", "reservatorio",
        "seguranca", "churrasqueira", "ar_condicionado", "internet",
        "condominio_fechado", "publicado_em", "data_extracao",
        "data_ultima_atualizacao", "ativo", "n_atualizacoes", "anunciante",
        "telefone_whatsapp", "email_agente", "segmento_imobiliario",
        "imovel_corporativo", "preco_disponivel", "hash_dedupe", "id_zenki",
        "url_brochura", "pdf_brochura", "url_imagem", "url", "id"
    ]
    ws1.append(cols1)
    _estilo_cabecalho(ws1)
    for i,a in enumerate(anuncios,2):
        ws1.append([_valor_para_excel(c, (a.get("titulo_anuncio") if c == "headline" and a.get("titulo_anuncio") else a.get(c))) for c in cols1])
        _zebra(ws1,i)
    _hyperlinkar_url(ws1, "url")
    _estilo_cabecalho(ws1)
    _aplicar_formatacao_padrao(ws1)
    _adicionar_tabela(ws1, "Tabela_Base_Anuncios")

    # ── ABA 2: Histórico de Preços ────────────────────────────
    ws2=wb.create_sheet("Histórico de Preços")
    hist=db.obter_historico()
    cols2=[
        "id","anuncio_id","id_angocasa","bairro","tipo_negocio","tipo_imovel",
        "tipologia","preco_original","moeda","preco_akz","preco_usd",
        "preco_m2_akz","preco_akz_ant","variacao_akz","variacao_pct","evento",
        "data_registo","url"
    ]
    ws2.append(cols2)
    _estilo_cabecalho(ws2)
    # Colorir alterações de preço
    fill_up  =PatternFill("solid",fgColor="FADBD8")  # vermelho claro (subiu)
    fill_down=PatternFill("solid",fgColor="D5F5E3")  # verde claro (desceu)
    for i,h in enumerate(hist,2):
        ws2.append([h.get(c) for c in cols2])
        if h.get("evento")=="preco_alterado":
            vp=h.get("variacao_pct") or 0
            fill=fill_up if vp>0 else fill_down
            for cell in ws2[i]: cell.fill=fill
        else:
            _zebra(ws2,i)
    _hyperlinkar_url(ws2, "url")
    _estilo_cabecalho(ws2)
    _aplicar_formatacao_padrao(ws2)
    _adicionar_tabela(ws2, "Tabela_Historico_Precos")

    # ── ABA 3: Alterações da Semana ───────────────────────────
    ws3=wb.create_sheet("Alterações da Semana")
    alt=db.obter_alteracoes_semana()
    cols3=[
        "id_angocasa","bairro","tipo_negocio","tipo_imovel","tipologia",
        "headline","area_m2","preco_akz_ant","preco_akz","variacao_akz",
        "variacao_pct","moeda","data_registo","url"
    ]
    ws3.append(cols3)
    _estilo_cabecalho(ws3)
    for i,a in enumerate(alt,2):
        ws3.append([a.get(c) for c in cols3])
        vp=a.get("variacao_pct") or 0
        fill=fill_up if vp>0 else fill_down
        for cell in ws3[i]: cell.fill=fill
    _hyperlinkar_url(ws3, "url")
    _estilo_cabecalho(ws3)
    _aplicar_formatacao_padrao(ws3)
    _adicionar_tabela(ws3, "Tabela_Alteracoes_Semana")

    # ── ABA 4: Parâmetros ─────────────────────────────────────
    ws4=wb.create_sheet("Parâmetros")
    ws4.append(["Parâmetro","Valor","Descrição"])
    _estilo_cabecalho(ws4)
    params=[
        ("Versão","2.0 Fase 1","PIIA — Plataforma de Inteligência Imobiliária Angola"),
        ("Data Exportação",datetime.now().strftime("%Y-%m-%d %H:%M"),""),
        ("Taxa Câmbio AKZ/USD",CFG["cambio"],"Atualizar no .env: TAXA_CAMBIO_AKZ_USD"),
        ("Delay Scraping (s)",f"{CFG['delay_min']}–{CFG['delay_max']}","Intervalo entre pedidos"),
        ("Máx. Páginas",CFG["max_paginas"],"Por pesquisa"),
        ("Items por Página",CFG["items_pag"],""),
        ("PROJECT_ROOT",str(PROJECT_ROOT),"Raiz do projecto"),
        ("Base de Dados",str(CFG["db"]),""),
        ("Bairros Configurados",len(CFG["bairros"]),""),
    ]
    for i,(k,v,d) in enumerate(params,2):
        ws4.append([k,v,d]); _zebra(ws4,i)
    ws4.append([])
    ws4.append(["Bairro","Slug","Lat","Lon"])
    row_b=ws4.max_row
    _estilo_cabecalho(ws4,row_b)
    for b in CFG["bairros"]:
        ws4.append([b["nome"],b["slug"],b["lat"],b["lon"]])
    _aplicar_formatacao_padrao(ws4)
    # Não criar tabela nativa nesta aba porque há duas secções e uma coluna D sem cabeçalho na primeira linha.

    # ── ABA 5: Logs / Stats ───────────────────────────────────
    ws5=wb.create_sheet("Logs")
    s=db.stats()
    ws5.append(["Métrica","Valor"])
    _estilo_cabecalho(ws5)
    metricas=[
        ("Total Anúncios Ativos",      s.get("total",0)),
        ("→ Compra",                   s.get("compra",0)),
        ("→ Arrendamento",             s.get("arrendamento",0)),
        ("Bairros com Dados",          s.get("bairros",0)),
        ("Registos Histórico Preços",  s.get("hist_precos",0)),
        ("Alterações de Preço",        s.get("preco_alt",0)),
        ("Anúncios Removidos",         s.get("removidos",0)),
        ("Erros de Scraping",          s.get("erros",0)),
        ("Última Extração",            s.get("ultima_extracao","N/A")),
        ("Exportado em",               datetime.now().isoformat()),
    ]
    for i,(k,v) in enumerate(metricas,2):
        ws5.append([k,v]); _zebra(ws5,i)
    _aplicar_formatacao_padrao(ws5)
    _adicionar_tabela(ws5, "Tabela_Logs")

    wb.save(str(ficheiro))
    log.info(f"✅ Excel exportado: {ficheiro.name}")
    return ficheiro


# ─────────────────────────────────────────────────────────────────
# PLACEHOLDERS PARA FASES FUTURAS (não implementados na Fase 1)
# ─────────────────────────────────────────────────────────────────
def fase2_analise_mercado(db):
    """FASE 2 — preço/m², mediana por bairro/tipologia, compra vs arrendamento."""
    raise NotImplementedError("Fase 2 — não implementado ainda.")

def fase2_yield_payback(db):
    """FASE 2 — yield bruto/líquido, payback, comparação."""
    raise NotImplementedError("Fase 2 — não implementado ainda.")

def fase3_scoring(db):
    """FASE 3 — score de oportunidade (0–100), score estratégico."""
    raise NotImplementedError("Fase 3 — não implementado ainda.")

def fase3_valuation(db):
    """FASE 3 — valor justo, margem de segurança."""
    raise NotImplementedError("Fase 3 — não implementado ainda.")

def fase4_pdfs(db):
    """FASE 4 — processamento de PDFs, OCR, extração de insights."""
    raise NotImplementedError("Fase 4 — não implementado ainda.")

def fase4_anthropic(db):
    """FASE 4 — análise com IA Anthropic (requer USAR_ANTHROPIC=TRUE no .env)."""
    raise NotImplementedError("Fase 4 — não implementado ainda.")

def fase5_gis(db):
    """FASE 5 — mapas de calor, folium, GeoPandas."""
    raise NotImplementedError("Fase 5 — não implementado ainda.")

def fase5_ml(db):
    """FASE 5 — modelos de Machine Learning para previsão de preços."""
    raise NotImplementedError("Fase 5 — não implementado ainda.")


# ─────────────────────────────────────────────────────────────────
# MAIN / ARGPARSE
# ─────────────────────────────────────────────────────────────────
def main():
    parser=argparse.ArgumentParser(
        prog="piia",
        description="PIIA v2.0 Fase 1 — Plataforma de Inteligência Imobiliária Angola",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python piia.py --setup
  python piia.py --scraping                         (AngoCasa: Luanda inteira, compra + arrendamento)
  python piia.py --scraping --fontes todas        (multi-fonte)
  python piia.py --scraping --bairro Talatona --max-paginas 3
  python piia.py --scraping --por-bairros --max-paginas 1
  python piia.py --excel
  python piia.py --debug
  python piia.py                       (executa tudo: setup + scraping + excel)
        """
    )
    parser.add_argument("--setup",       action="store_true",
                        help="Criar estrutura de pastas e .env.example")
    parser.add_argument("--scraping",    action="store_true",
                        help="Recolher anúncios imobiliários")
    parser.add_argument("--excel",       action="store_true",
                        help="Exportar Excel com dados actuais")
    parser.add_argument("--bairro",      type=str, default=None, metavar="NOME",
                        help="Limitar scraping a um bairro (ex: Talatona)")
    parser.add_argument("--por-bairros", action="store_true",
                        help="Percorrer bairro por bairro usando a lista configurada. Por padrão coleta Luanda inteira.")
    parser.add_argument("--max-paginas", type=int, default=None, metavar="N",
                        help="Número máximo de páginas por pesquisa (sobrepõe config)")
    parser.add_argument("--debug",       action="store_true",
                        help="Log verbose + guardar HTML das páginas")
    parser.add_argument("--forcar",      action="store_true",
                        help="Forçar reprocessamento dos anúncios, ignorando a janela de retomada inteligente")
    parser.add_argument("--fontes",      type=str, default="angocasa", metavar="LISTA",
                        help="Fontes separadas por vírgula: angocasa, angoimoveis, mia, remax, zenki, ree_gera ou todas")
    parser.add_argument("--max-anuncios", type=int, default=None, metavar="N",
                        help="Limite máximo de anúncios por fonte/negócio para testes rápidos.")
    parser.add_argument("--apenas-negocio", type=str, default="todos",
                        choices=["todos", "compra", "arrendamento"],
                        help="Limita a coleta a compra ou arrendamento. Padrão: todos.")
    parser.add_argument("--rapido",      action="store_true",
                        help="Modo rápido: reduz delays, timeout e retries. Use com moderação.")
    parser.add_argument("--delay-min",   type=float, default=None,
                        help="Delay mínimo entre pedidos, em segundos.")
    parser.add_argument("--delay-max",   type=float, default=None,
                        help="Delay máximo entre pedidos, em segundos.")
    parser.add_argument("--timeout",     type=int, default=None,
                        help="Timeout HTTP em segundos.")
    args=parser.parse_args()

    # Nenhum argumento → executar tudo
    executar_tudo = not (args.setup or args.scraping or args.excel)

    # Carregar .env antes de configurar log
    carregar_env()

    # Aplicar overrides de CLI
    if args.debug:
        CFG["salvar_html"]=True
    if args.max_paginas:
        CFG["max_paginas"]=args.max_paginas
    CFG["apenas_negocio"] = args.apenas_negocio
    CFG["max_anuncios_por_fonte"] = args.max_anuncios
    if args.rapido:
        CFG["delay_min"] = 0.3
        CFG["delay_max"] = 1.0
        CFG["timeout"] = 20
        CFG["max_retries"] = 2
    if args.delay_min is not None:
        CFG["delay_min"] = args.delay_min
    if args.delay_max is not None:
        CFG["delay_max"] = args.delay_max
    if args.timeout is not None:
        CFG["timeout"] = args.timeout

    # Configurar logging
    configurar_log(debug=args.debug)

    log.info("═"*60)
    log.info("  PIIA v2.10 — Fase 1")
    log.info(f"  Início: {datetime.now():%Y-%m-%d %H:%M:%S}")
    log.info(f"  ROOT  : {PROJECT_ROOT}")
    log.info("═"*60)

    # ── SETUP ────────────────────────────────────────────────
    if args.setup or executar_tudo:
        log.info("▶ SETUP — criar estrutura de pastas")
        criar_estrutura()

    # Inicializar BD (sempre necessário)
    db=DB()

    # ── SCRAPING ─────────────────────────────────────────────
    if args.scraping or executar_tudo:
        log.info("▶ SCRAPING — recolher anúncios")
        rid=db.iniciar_run("scraping")
        scraper=Scraper(db)
        try:
            fontes = [f.strip().lower() for f in args.fontes.split(",") if f.strip()]
            stats=scraper.scraping_multifonte(db,bairro_filtro=args.bairro, por_bairros=args.por_bairros, forcar=args.forcar, fontes=fontes)
            if stats:
                log.info("─"*50)
                log.info(f"  Coletados    : {stats['n_coletados']}")
                log.info(f"  Novos        : {stats['n_novos']}")
                log.info(f"  Atualizados  : {stats['n_atualizados']}")
                log.info(f"  Preço alterado: {stats['n_preco_alt']}")
                log.info(f"  Removidos    : {stats['n_removidos']}")
                log.info(f"  Erros        : {stats['n_erros']}")
                log.info("─"*50)
                db.finalizar_run(rid,{
                    "n_coletados": stats["n_coletados"],
                    "n_novos": stats["n_novos"],
                    "n_atualizados":stats["n_atualizados"],
                    "n_preco_alt": stats["n_preco_alt"],
                    "n_removidos": stats["n_removidos"],
                    "n_erros":     stats["n_erros"],
                    "bairros_proc":stats.get("bairros_proc",""),
                })
                db.backup()
        except KeyboardInterrupt:
            log.warning("Scraping interrompido pelo utilizador.")
            db.finalizar_run(rid,{"status":"interrompido","n_erros":1})
        except Exception as e:
            log.error(f"Erro fatal no scraping: {e}")
            db.finalizar_run(rid,{"status":"erro","detalhes":str(e)})
            raise

    # ── EXCEL ─────────────────────────────────────────────────
    if args.excel or executar_tudo:
        log.info("▶ EXCEL — exportar relatório")
        f=exportar_excel(db)
        if f:
            log.info(f"  Ficheiro: {f}")

    # ── STATS FINAIS ──────────────────────────────────────────
    s=db.stats()
    log.info("═"*60)
    log.info("  RESUMO BASE DE DADOS")
    log.info(f"  Total anúncios activos : {s['total']}")
    log.info(f"  → Compra               : {s['compra']}")
    log.info(f"  → Arrendamento         : {s['arrendamento']}")
    log.info(f"  Bairros com dados      : {s['bairros']}")
    log.info(f"  Hist. alterações preço : {s['preco_alt']}")
    log.info(f"  Última extração        : {s['ultima_extracao']}")
    log.info("═"*60)
    log.info("✅ PIIA Fase 1 concluída.")


if __name__ == "__main__":
    main()
