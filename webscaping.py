from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as Ec
from openpyxl import Workbook
from openpyxl import load_workbook
import os


sheet_location=r'C:/Users/Admins/Downloads/WebScrabed.xlsx'
xpath="//*[contains(text(),'Published')]"
datetext="Published:"

url='https://www.aiimsexams.ac.in/'

def storedata(*args):
    row=list(args)
    print("type of result....", row)  
    if not os.path.exists(sheet_location):
        wb=Workbook()
        ws=wb.active
        ws.title='Web scrabed'
        ws.append(row)
    else:
        wb=load_workbook(sheet_location)
        ws=wb.active
    ws.append(row)
    wb.save(sheet_location)




    

driver=webdriver.Chrome()
driver.get(url)
driver.set_page_load_timeout(30)
driver.maximize_window()
WebDriverWait(driver,30).until(Ec.presence_of_element_located((By.XPATH, xpath)))

#page title
print(driver.title)
html=driver.page_source  


#Find all links the website
links=driver.find_elements(By.TAG_NAME, 'a')
print("Total number of links........",len(links))
for lnk in links:
    linktext=lnk.text
    linkurl=lnk.get_attribute('href')
    storedata(linktext, linkurl)

#Getting published date
def publisheddate(xpath, datetext):
    pubdates=driver.find_elements(By.XPATH, xpath)
    print(type(pubdates))
    for a in pubdates:
        textword=a.text.split( )[1]
        dates=a.text.split(datetext)[-1].strip()
        storedata(textword, dates)
      
publisheddate(xpath, datetext)